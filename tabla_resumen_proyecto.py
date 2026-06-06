from __future__ import annotations

import unicodedata
import calendar
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LOCALIDADES_PROYECTO = [
    (1, "USAQUÉN"),
    (2, "CHAPINERO"),
    (3, "SANTA  FE"),
    (4, "SAN CRISTÓBAL"),
    (5, "USME"),
    (6, "TUNJUELITO"),
    (7, "BOSA"),
    (8, "KENNEDY"),
    (9, "FONTIBON"),
    (10, "ENGATIVÁ"),
    (11, "SUBA"),
    (12, "BARRIOS UNIDOS"),
    (13, "TEUSAQUILLO"),
    (14, "LOS MÁRTIRES"),
    (15, "ANTONIO NARIÑO"),
    (16, "PUENTE  ARANDA"),
    (17, "LA  CANDELARIA"),
    (18, "RAFAEL URIBE URIBE"),
    (19, "CIUDAD BOLÍVAR"),
    (20, "SUMAPAZ"),
]

MESES_ES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

FILL_TITULO_GRIS = PatternFill("solid", fgColor="D9D9D9")
FILL_BLANCO = PatternFill("solid", fgColor="FFFFFF")
FORMATO_PESOS = '_-"$"* #,##0_-;\\-"$"* #,##0_-;_-"$"* "-"_-;_-@_-'


@dataclass(frozen=True)
class FilaLibFen:
    numero: int
    localidad: str
    liberacion: float
    fenecimiento: float

    @property
    def total(self) -> float:
        return self.liberacion + self.fenecimiento


@dataclass(frozen=True)
class FilaConPerdida:
    numero: int
    localidad: str
    contratista: str
    numero_contrato: object
    vigencia: object
    fecha_finalizacion: datetime
    fecha_perdida_competencia: datetime
    monto: float


@dataclass(frozen=True)
class FilaBogdataMatriz:
    numero: int
    localidad: str
    apropiacion_matriz: float
    giros_matriz: float
    saldo_final_matriz: float


@dataclass(frozen=True)
class FilaDepurados:
    numero: int
    localidad: str
    cantidad_inicial: int
    depurados: int


@dataclass(frozen=True)
class FilaCpsPn:
    numero: int
    localidad: str
    cantidad_inicial: int
    depurados: int


@dataclass(frozen=True)
class FilaPorDepurarVigencia:
    numero: int
    localidad: str
    conteos_por_vigencia: dict[int, tuple[int, int]]


def normalizar(texto: object) -> str:
    try:
        if pd.isna(texto):
            return ""
    except TypeError:
        pass
    s = str(texto or "").strip().lower()
    s = "".join(
        ch for ch in unicodedata.normalize("NFD", s)
        if unicodedata.category(ch) != "Mn"
    )
    return " ".join(s.split())


LOCALIDADES_NUMERO = {normalizar(nombre): numero for numero, nombre in LOCALIDADES_PROYECTO}
LOCALIDADES_NOMBRE_REPORTE = {numero: nombre for numero, nombre in LOCALIDADES_PROYECTO}


def numero_localidad(localidad: str) -> int:
    numero = LOCALIDADES_NUMERO.get(normalizar(localidad))
    if numero is None:
        raise ValueError(f"Localidad sin número asignado para tabla resumen: {localidad}")
    return numero


def nombre_localidad_reporte(localidad: str) -> str:
    return LOCALIDADES_NOMBRE_REPORTE[numero_localidad(localidad)]


def _columna(df: pd.DataFrame, *partes: str) -> str | None:
    partes_norm = [normalizar(p) for p in partes if normalizar(p)]
    for col in df.columns:
        n = normalizar(col)
        if all(p in n for p in partes_norm):
            return col
    return None


def _es_celda_vacia(valor: object) -> bool:
    if valor is None:
        return True
    try:
        if pd.isna(valor):
            return True
    except TypeError:
        pass
    return not str(valor).strip()


def _mascara_fila_real(df: pd.DataFrame) -> pd.Series:
    col_nombre = _columna(df, "nombre", "contratista")
    col_contrato = _columna(df, "numero", "contrato")
    if col_nombre:
        return ~df[col_nombre].map(_es_celda_vacia)
    if col_contrato:
        return ~df[col_contrato].map(_es_celda_vacia)
    return pd.Series(True, index=df.index)


def _fecha_datetime(valor: object) -> datetime | None:
    if valor is None or _es_celda_vacia(valor):
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime(valor.year, valor.month, valor.day)
    parsed = pd.to_datetime(valor, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def _sumar_meses(fecha: datetime, meses: int) -> datetime:
    total = fecha.month - 1 + meses
    anio = fecha.year + total // 12
    mes = total % 12 + 1
    dia = min(fecha.day, calendar.monthrange(anio, mes)[1])
    return datetime(anio, mes, dia)


def _nombre_mes_es(fecha: datetime) -> str:
    return MESES_ES.get(fecha.month, fecha.strftime("%B").upper())


def _mes_objetivo_perdida(fecha_corte: datetime | date) -> tuple[int, int]:
    f = _fecha_datetime(fecha_corte) or datetime.now()
    objetivo = _sumar_meses(f, -30)
    return objetivo.year, objetivo.month


def _rango_proximos_a_perder(fecha_corte: datetime | date) -> tuple[datetime, datetime]:
    f = _fecha_datetime(fecha_corte) or datetime.now()
    inicio = _sumar_meses(datetime(f.year, f.month, 1), 1)
    mes_final = _sumar_meses(inicio, 5)
    fin = datetime(
        mes_final.year,
        mes_final.month,
        calendar.monthrange(mes_final.year, mes_final.month)[1],
    )
    return inicio, fin


def _estado_perdida_reportable(estado: object, acciones: object = "") -> bool:
    del acciones
    n = normalizar(estado)
    if not n:
        return False
    if "liquidado" in n:
        return False
    return (
        ("revision" in n and "entes de control" in n)
        or ("terminado" in n and "no se liquida" in n)
        or ("terminado" in n and "proceso de liquidacion" in n)
        or ("presunto incumplimiento" in n)
        or ("demanda" in n)
    )


def fila_lib_y_fen_desde_matriz(df_matriz: pd.DataFrame, localidad: str) -> FilaLibFen:
    """Suma Liberación y Fenecimiento desde la Matriz OXP de una localidad."""
    col_tipo = _columna(df_matriz, "tipo", "anul")
    col_monto = _columna(df_matriz, "monto", "liber") or _columna(
        df_matriz, "monto", "fenec"
    )
    if not col_tipo or not col_monto:
        raise ValueError(
            "La Matriz no tiene las columnas Tipo de Anulación y "
            "Monto Liberación/Fenecimiento."
        )

    reales = _mascara_fila_real(df_matriz)
    tipos = df_matriz[col_tipo].map(normalizar)
    montos = pd.to_numeric(df_matriz[col_monto], errors="coerce").fillna(0)

    # Filas reales con tipo vacío se tratan como Liberación; se excluyen filas de total
    # sin contratista para no duplicar montos de la Matriz.
    liberacion = montos[reales & ((tipos == "liberacion") | (tipos == ""))].sum()
    fenecimiento = montos[reales & (tipos == "fenecimiento")].sum()

    numero = numero_localidad(localidad)

    return FilaLibFen(
        numero=numero,
        localidad=nombre_localidad_reporte(localidad),
        liberacion=float(liberacion),
        fenecimiento=float(fenecimiento),
    )


def fila_bogdata_matriz_desde_matriz(
    df_matriz: pd.DataFrame,
    localidad: str,
) -> FilaBogdataMatriz:
    """Suma columnas de Matriz para la hoja BOGDATA VS MATRIZ."""
    col_apropiacion = _columna(df_matriz, "apropiacion")
    col_giros = _columna(df_matriz, "giros")
    col_saldo = _columna(df_matriz, "saldo", "final")
    if not col_apropiacion or not col_giros or not col_saldo:
        raise ValueError(
            "La Matriz no tiene las columnas Apropiación, Giros y Saldo Final."
        )

    reales = _mascara_fila_real(df_matriz)
    apropiacion = pd.to_numeric(
        df_matriz[col_apropiacion], errors="coerce"
    ).fillna(0)
    giros = pd.to_numeric(df_matriz[col_giros], errors="coerce").fillna(0)
    saldo = pd.to_numeric(df_matriz[col_saldo], errors="coerce").fillna(0)
    numero = numero_localidad(localidad)

    return FilaBogdataMatriz(
        numero=numero,
        localidad=nombre_localidad_reporte(localidad),
        apropiacion_matriz=float(apropiacion[reales].sum()),
        giros_matriz=float(giros[reales].sum()),
        saldo_final_matriz=float(saldo[reales].sum()),
    )


def fila_depurados_desde_matriz(
    df_matriz: pd.DataFrame,
    localidad: str,
) -> FilaDepurados:
    """Cuenta contratos iniciales y depurados desde Año Suscripción y Saldo Final."""
    col_vigencia = _columna(df_matriz, "ano", "suscripcion")
    col_saldo = _columna(df_matriz, "saldo", "final")
    if not col_vigencia or not col_saldo:
        raise ValueError(
            "La Matriz no tiene las columnas Año Suscripción y Saldo Final."
        )

    filas_con_vigencia = ~df_matriz[col_vigencia].map(_es_celda_vacia)
    saldo_vacio = df_matriz[col_saldo].map(_es_celda_vacia)
    saldo_cero = pd.to_numeric(df_matriz[col_saldo], errors="coerce").fillna(0).eq(0)
    depurados = filas_con_vigencia & (saldo_vacio | saldo_cero)
    numero = numero_localidad(localidad)

    return FilaDepurados(
        numero=numero,
        localidad=nombre_localidad_reporte(localidad),
        cantidad_inicial=int(filas_con_vigencia.sum()),
        depurados=int(depurados.sum()),
    )


def fila_cps_pn_desde_matriz(
    df_matriz: pd.DataFrame,
    localidad: str,
) -> FilaCpsPn:
    """Cuenta CPS persona natural iniciales y depurados desde Clasificación y Saldo Final."""
    col_clasificacion = _columna(df_matriz, "clasificacion")
    col_saldo = _columna(df_matriz, "saldo", "final")
    if not col_clasificacion or not col_saldo:
        raise ValueError(
            "La Matriz no tiene las columnas Clasificación y Saldo Final."
        )

    clasificaciones = df_matriz[col_clasificacion].map(normalizar)
    es_cps_pn = clasificaciones.map(
        lambda valor: (
            "prestacion de servicios" in valor
            and "persona natural" in valor
        )
    )
    saldo_vacio = df_matriz[col_saldo].map(_es_celda_vacia)
    saldo_cero = pd.to_numeric(df_matriz[col_saldo], errors="coerce").fillna(0).eq(0)
    depurados = es_cps_pn & (saldo_vacio | saldo_cero)
    numero = numero_localidad(localidad)

    return FilaCpsPn(
        numero=numero,
        localidad=nombre_localidad_reporte(localidad),
        cantidad_inicial=int(es_cps_pn.sum()),
        depurados=int(depurados.sum()),
    )


def fila_por_depurar_vigencia_desde_matriz(
    df_matriz: pd.DataFrame,
    localidad: str,
) -> FilaPorDepurarVigencia:
    """Cuenta inicial y con saldo por Año Suscripción."""
    col_vigencia = _columna(df_matriz, "ano", "suscripcion")
    col_saldo = _columna(df_matriz, "saldo", "final")
    if not col_vigencia or not col_saldo:
        raise ValueError(
            "La Matriz no tiene las columnas Año Suscripción y Saldo Final."
        )

    vigencias = pd.to_numeric(df_matriz[col_vigencia], errors="coerce")
    saldo_vacio = df_matriz[col_saldo].map(_es_celda_vacia)
    saldos = pd.to_numeric(df_matriz[col_saldo], errors="coerce")
    tiene_saldo = (~saldo_vacio) & saldos.fillna(0).ne(0)
    conteos: dict[int, tuple[int, int]] = {}

    for vigencia in sorted(int(v) for v in vigencias.dropna().unique()):
        filas_vigencia = vigencias.eq(vigencia)
        inicial = int(filas_vigencia.sum())
        con_saldo = int((filas_vigencia & tiene_saldo).sum())
        conteos[vigencia] = (inicial, con_saldo)

    numero = numero_localidad(localidad)
    return FilaPorDepurarVigencia(
        numero=numero,
        localidad=nombre_localidad_reporte(localidad),
        conteos_por_vigencia=conteos,
    )


def _filas_perdida_desde_matriz(
    df_matriz: pd.DataFrame,
    localidad: str,
    incluir_fechas,
    mensaje_columnas: str,
) -> list[FilaConPerdida]:
    columnas = {
        "contratista": _columna(df_matriz, "nombre", "contratista"),
        "contrato": _columna(df_matriz, "numero", "contrato"),
        "vigencia": _columna(df_matriz, "ano", "suscripcion"),
        "fecha_fin": _columna(df_matriz, "fecha", "final"),
        "saldo": _columna(df_matriz, "saldo", "final"),
        "estado": _columna(df_matriz, "estado", "actual"),
        "acciones": _columna(df_matriz, "acciones"),
    }
    requeridas = ("contratista", "contrato", "vigencia", "fecha_fin", "saldo", "estado")
    if any(not columnas[c] for c in requeridas):
        raise ValueError(mensaje_columnas)

    numero = numero_localidad(localidad)
    nombre_loc = nombre_localidad_reporte(localidad)
    saldos = pd.to_numeric(df_matriz[columnas["saldo"]], errors="coerce").fillna(0)
    grupos: dict[tuple[str, str, str, datetime], FilaConPerdida] = {}

    for idx, row in df_matriz.iterrows():
        if saldos.loc[idx] <= 0:
            continue
        fecha_fin = _fecha_datetime(row[columnas["fecha_fin"]])
        if fecha_fin is None:
            continue
        fecha_perdida = _sumar_meses(fecha_fin, 30)
        if not incluir_fechas(fecha_fin, fecha_perdida):
            continue
        acciones = row[columnas["acciones"]] if columnas.get("acciones") else ""
        if not _estado_perdida_reportable(row[columnas["estado"]], acciones):
            continue
        contratista = row[columnas["contratista"]]
        if _es_celda_vacia(contratista):
            continue
        contratista_txt = str(contratista).strip()
        numero_contrato = row[columnas["contrato"]]
        vigencia = row[columnas["vigencia"]]
        llave = (
            normalizar(contratista_txt),
            normalizar(numero_contrato),
            normalizar(vigencia),
            datetime(fecha_fin.year, fecha_fin.month, fecha_fin.day),
        )
        existente = grupos.get(llave)
        if existente is None:
            grupos[llave] = FilaConPerdida(
                numero=numero,
                localidad=nombre_loc,
                contratista=contratista_txt,
                numero_contrato=numero_contrato,
                vigencia=vigencia,
                fecha_finalizacion=fecha_fin,
                fecha_perdida_competencia=fecha_perdida,
                monto=float(saldos.loc[idx]),
            )
        else:
            grupos[llave] = FilaConPerdida(
                numero=existente.numero,
                localidad=existente.localidad,
                contratista=existente.contratista,
                numero_contrato=existente.numero_contrato,
                vigencia=existente.vigencia,
                fecha_finalizacion=existente.fecha_finalizacion,
                fecha_perdida_competencia=existente.fecha_perdida_competencia,
                monto=existente.monto + float(saldos.loc[idx]),
            )

    return sorted(
        grupos.values(),
        key=lambda f: (
            f.numero,
            _clave_numero_contrato(f.numero_contrato),
            f.fecha_finalizacion,
            f.contratista,
        ),
    )


def filas_con_perdida_desde_matriz(
    df_matriz: pd.DataFrame,
    localidad: str,
    fecha_corte: datetime | date,
) -> list[FilaConPerdida]:
    """Contratos que cumplen 30 meses desde fecha de finalización en el mes de corte."""
    anio_obj, mes_obj = _mes_objetivo_perdida(fecha_corte)

    def incluir_fechas(fecha_fin: datetime, fecha_perdida: datetime) -> bool:
        del fecha_perdida
        return fecha_fin.year == anio_obj and fecha_fin.month == mes_obj

    return _filas_perdida_desde_matriz(
        df_matriz,
        localidad,
        incluir_fechas,
        "La Matriz no tiene las columnas necesarias para CON PÉRDIDA.",
    )


def filas_proximos_a_perder_desde_matriz(
    df_matriz: pd.DataFrame,
    localidad: str,
    fecha_corte: datetime | date,
) -> list[FilaConPerdida]:
    """Contratos que perderán competencia en los seis meses posteriores al corte."""
    inicio, fin = _rango_proximos_a_perder(fecha_corte)

    def incluir_fechas(fecha_fin: datetime, fecha_perdida: datetime) -> bool:
        del fecha_fin
        return inicio <= fecha_perdida <= fin

    return _filas_perdida_desde_matriz(
        df_matriz,
        localidad,
        incluir_fechas,
        "La Matriz no tiene las columnas necesarias para PRÓXIMOS A PERDER.",
    )


def ordenar_filas_lib_y_fen(filas: Iterable[FilaLibFen | dict]) -> list[FilaLibFen]:
    normalizadas: list[FilaLibFen] = []
    for fila in filas:
        if isinstance(fila, FilaLibFen):
            normalizadas.append(fila)
        else:
            normalizadas.append(
                FilaLibFen(
                    numero=int(fila["numero"]),
                    localidad=str(fila["localidad"]),
                    liberacion=float(fila.get("liberacion", 0) or 0),
                    fenecimiento=float(fila.get("fenecimiento", 0) or 0),
                )
            )
    return sorted(normalizadas, key=lambda f: f.numero)


def _fila_con_perdida_desde_dict(fila: FilaConPerdida | dict) -> FilaConPerdida:
    if isinstance(fila, FilaConPerdida):
        return fila
    fecha_fin = _fecha_datetime(fila.get("fecha_finalizacion"))
    fecha_perdida = _fecha_datetime(fila.get("fecha_perdida_competencia"))
    if fecha_fin is None or fecha_perdida is None:
        raise ValueError("Fila CON PÉRDIDA sin fechas válidas.")
    return FilaConPerdida(
        numero=int(fila["numero"]),
        localidad=str(fila["localidad"]),
        contratista=str(fila["contratista"]),
        numero_contrato=fila.get("numero_contrato"),
        vigencia=fila.get("vigencia"),
        fecha_finalizacion=fecha_fin,
        fecha_perdida_competencia=fecha_perdida,
        monto=float(fila.get("monto", 0) or 0),
    )


def _clave_numero_contrato(valor: object) -> tuple[int, float | str, str]:
    texto = str(valor or "").strip()
    if not texto:
        return (2, "", "")
    numero = pd.to_numeric(texto, errors="coerce")
    if pd.notna(numero):
        return (0, float(numero), texto)
    return (1, normalizar(texto), texto)


def ordenar_filas_con_perdida(
    filas: Iterable[FilaConPerdida | dict],
) -> list[FilaConPerdida]:
    normalizadas = [_fila_con_perdida_desde_dict(fila) for fila in filas]
    return sorted(
        normalizadas,
        key=lambda f: (
            f.numero,
            _clave_numero_contrato(f.numero_contrato),
            f.fecha_finalizacion,
            f.contratista,
        ),
    )


def _fila_bogdata_matriz_desde_dict(
    fila: FilaBogdataMatriz | dict,
) -> FilaBogdataMatriz:
    if isinstance(fila, FilaBogdataMatriz):
        return fila
    return FilaBogdataMatriz(
        numero=int(fila["numero"]),
        localidad=str(fila["localidad"]),
        apropiacion_matriz=float(fila.get("apropiacion_matriz", 0) or 0),
        giros_matriz=float(fila.get("giros_matriz", 0) or 0),
        saldo_final_matriz=float(fila.get("saldo_final_matriz", 0) or 0),
    )


def ordenar_filas_bogdata_matriz(
    filas: Iterable[FilaBogdataMatriz | dict],
) -> list[FilaBogdataMatriz]:
    normalizadas = [_fila_bogdata_matriz_desde_dict(fila) for fila in filas]
    return sorted(normalizadas, key=lambda f: f.numero)


def _fila_depurados_desde_dict(fila: FilaDepurados | dict) -> FilaDepurados:
    if isinstance(fila, FilaDepurados):
        return fila
    return FilaDepurados(
        numero=int(fila["numero"]),
        localidad=str(fila["localidad"]),
        cantidad_inicial=int(fila.get("cantidad_inicial", 0) or 0),
        depurados=int(fila.get("depurados", 0) or 0),
    )


def ordenar_filas_depurados(filas: Iterable[FilaDepurados | dict]) -> list[FilaDepurados]:
    normalizadas = [_fila_depurados_desde_dict(fila) for fila in filas]
    return sorted(normalizadas, key=lambda f: f.numero)


def _fila_cps_pn_desde_dict(fila: FilaCpsPn | dict) -> FilaCpsPn:
    if isinstance(fila, FilaCpsPn):
        return fila
    return FilaCpsPn(
        numero=int(fila["numero"]),
        localidad=str(fila["localidad"]),
        cantidad_inicial=int(fila.get("cantidad_inicial", 0) or 0),
        depurados=int(fila.get("depurados", 0) or 0),
    )


def ordenar_filas_cps_pn(filas: Iterable[FilaCpsPn | dict]) -> list[FilaCpsPn]:
    normalizadas = [_fila_cps_pn_desde_dict(fila) for fila in filas]
    return sorted(normalizadas, key=lambda f: f.numero)


def _fila_por_depurar_vigencia_desde_dict(
    fila: FilaPorDepurarVigencia | dict,
) -> FilaPorDepurarVigencia:
    if isinstance(fila, FilaPorDepurarVigencia):
        return fila
    conteos_raw = fila.get("conteos_por_vigencia", {}) or {}
    conteos: dict[int, tuple[int, int]] = {}
    for vigencia, valores in conteos_raw.items():
        if isinstance(valores, dict):
            inicial = valores.get("inicial", 0)
            con_saldo = valores.get("con_saldo", 0)
        else:
            inicial, con_saldo = valores
        conteos[int(vigencia)] = (int(inicial or 0), int(con_saldo or 0))
    return FilaPorDepurarVigencia(
        numero=int(fila["numero"]),
        localidad=str(fila["localidad"]),
        conteos_por_vigencia=conteos,
    )


def ordenar_filas_por_depurar_vigencia(
    filas: Iterable[FilaPorDepurarVigencia | dict],
) -> list[FilaPorDepurarVigencia]:
    normalizadas = [_fila_por_depurar_vigencia_desde_dict(fila) for fila in filas]
    return sorted(normalizadas, key=lambda f: f.numero)


def crear_excel_tabla_resumen_proyecto_lib_y_fen(filas: Iterable[FilaLibFen | dict]) -> bytes:
    """Crea el primer Excel Tabla Resumen Proyecto con la pestaña LIB Y FEN."""
    wb = Workbook()
    ws = wb.active
    ws.title = "LIB Y FEN"
    _llenar_hoja_lib_y_fen(ws, filas)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def crear_excel_tabla_resumen_proyecto(
    filas_lib_y_fen: Iterable[FilaLibFen | dict],
    filas_con_perdida: Iterable[FilaConPerdida | dict],
    fecha_corte: datetime | date,
    filas_proximos_a_perder: Iterable[FilaConPerdida | dict] | None = None,
    filas_bogdata_matriz: Iterable[FilaBogdataMatriz | dict] | None = None,
    filas_depurados: Iterable[FilaDepurados | dict] | None = None,
    filas_por_depurar_vigencia: Iterable[FilaPorDepurarVigencia | dict] | None = None,
    filas_cps_pn: Iterable[FilaCpsPn | dict] | None = None,
) -> bytes:
    """Crea el Excel Tabla Resumen Proyecto con las pestañas ya definidas."""
    wb = Workbook()
    ws_lib = wb.active
    ws_lib.title = "LIB Y FEN"
    _llenar_hoja_lib_y_fen(ws_lib, filas_lib_y_fen)
    _crear_hoja_con_perdida(
        wb,
        ordenar_filas_lib_y_fen(filas_lib_y_fen),
        filas_con_perdida,
        fecha_corte,
    )
    _crear_hoja_proximos_a_perder(
        wb,
        ordenar_filas_lib_y_fen(filas_lib_y_fen),
        filas_proximos_a_perder or [],
        fecha_corte,
    )
    _crear_hoja_bogdata_vs_matriz(
        wb,
        filas_bogdata_matriz or [],
        fecha_corte,
    )
    _crear_hoja_depurados(
        wb,
        filas_depurados or [],
        fecha_corte,
    )
    _crear_hoja_por_depurar_vigencia(
        wb,
        filas_por_depurar_vigencia or [],
        fecha_corte,
    )
    _crear_hoja_cps_pn(
        wb,
        filas_cps_pn or [],
        fecha_corte,
    )
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _llenar_hoja_lib_y_fen(ws, filas: Iterable[FilaLibFen | dict]) -> None:
    headers = [
        "COD",
        "LOCALIDAD",
        "LIBERACION",
        "FENECIMIENTO",
        "TOTAL",
        "% PARTICIPACION",
        "APROPIACION DISPONIBLE",
    ]
    ws.append([""] * len(headers))
    ws.append([""] * len(headers))
    ws.append(headers)

    filas_ordenadas = ordenar_filas_lib_y_fen(filas)
    for idx, fila in enumerate(filas_ordenadas, start=4):
        total = fila.total
        apropiacion = 0
        ws.cell(idx, 1).value = fila.numero
        ws.cell(idx, 2).value = fila.localidad.upper()
        ws.cell(idx, 3).value = fila.liberacion
        ws.cell(idx, 4).value = fila.fenecimiento
        ws.cell(idx, 5).value = total
        ws.cell(idx, 6).value = total / apropiacion if apropiacion else ""
        ws.cell(idx, 7).value = apropiacion

    ultima = 4 + len(filas_ordenadas)
    ws.cell(ultima, 2).value = "TOTAL"
    ws.cell(ultima, 3).value = sum(f.liberacion for f in filas_ordenadas)
    ws.cell(ultima, 4).value = sum(f.fenecimiento for f in filas_ordenadas)
    ws.cell(ultima, 5).value = sum(f.total for f in filas_ordenadas)
    ws.cell(ultima, 6).value = ""
    ws.cell(ultima, 7).value = 0

    _formatear_hoja_lib_y_fen(ws, ultima)


def _formatear_hoja_lib_y_fen(ws, ultima_fila: int) -> None:
    borde = Side(style="thin", color="808080")
    border = Border(left=borde, right=borde, top=borde, bottom=borde)
    fuente_header = Font(bold=True, color="000000")
    fuente_total = Font(bold=True, color="000000")

    for cell in ws[3]:
        cell.font = fuente_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        cell.fill = FILL_TITULO_GRIS

    for row in ws.iter_rows(min_row=4, max_row=ultima_fila, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = FILL_BLANCO
            if cell.row == ultima_fila:
                cell.font = fuente_total

    for col in ("C", "D", "E", "G"):
        for row in range(4, ultima_fila + 1):
            ws[f"{col}{row}"].number_format = FORMATO_PESOS
    for row in range(4, ultima_fila + 1):
        ws[f"F{row}"].number_format = "0.00%"

    widths = {
        "A": 8,
        "B": 22,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 16,
        "G": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(3, ultima_fila + 1):
        ws.row_dimensions[row].height = 22
    ws.freeze_panes = "A4"


def _crear_hoja_bogdata_vs_matriz(
    wb,
    filas_bogdata_matriz: Iterable[FilaBogdataMatriz | dict],
    fecha_corte: datetime | date,
) -> None:
    ws = wb.create_sheet("BOGDATA VS MATRIZ")
    filas_ordenadas = ordenar_filas_bogdata_matriz(filas_bogdata_matriz)
    f = _fecha_datetime(fecha_corte) or datetime.now()
    titulo = f"BOGDATA VS MATRIZ A {f.day} DE {_nombre_mes_es(f)} DE {f.year}"

    ws.merge_cells("A1:K1")
    ws["A1"] = titulo
    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:E2")
    ws.merge_cells("F2:H2")
    ws.merge_cells("I2:K2")
    ws["A2"] = "Nº"
    ws["B2"] = "LOCALIDAD"
    ws["C2"] = "APROPIACIÓN DISPONIBLE"
    ws["F2"] = "AUTORIZACIÓN GIROS ACUMULADO"
    ws["I2"] = "SALDO FINAL"
    headers = {
        "C3": "BOGDATA",
        "D3": "MATRIZ FDL",
        "E3": "DIFERENCIA APROPIACIÓN DISPONIBLE",
        "F3": "BOGDATA",
        "G3": "MATRIZ FDL",
        "H3": "DIFERENCIA AUTORIZACIÓN GIROS ACUMULADOS",
        "I3": "BOGDATA",
        "J3": "MATRIZ FDL",
        "K3": "DIFERENCIA SALDOS FINAL",
    }
    for coord, valor in headers.items():
        ws[coord] = valor

    for idx, fila in enumerate(filas_ordenadas, start=4):
        ws.cell(idx, 1).value = fila.numero
        ws.cell(idx, 2).value = " ".join(fila.localidad.title().split())
        ws.cell(idx, 4).value = fila.apropiacion_matriz
        ws.cell(idx, 7).value = fila.giros_matriz
        ws.cell(idx, 10).value = fila.saldo_final_matriz

    total_row = 4 + len(filas_ordenadas)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    ws.cell(total_row, 1).value = "TOTAL"
    first_data = 4
    last_data = max(first_data, total_row - 1)
    for col in (4, 7, 10):
        letter = get_column_letter(col)
        ws.cell(total_row, col).value = (
            f"=SUM({letter}{first_data}:{letter}{last_data})"
            if filas_ordenadas
            else 0
        )

    note_row = total_row + 2
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row + 1, end_column=10)
    ws.cell(note_row, 2).value = (
        "Las cifras de BogData las alimento yo cuando se descargue la ejecución "
        "al final de mes"
    )
    _formatear_hoja_bogdata_vs_matriz(ws, total_row, note_row)


def _formatear_hoja_bogdata_vs_matriz(ws, total_row: int, note_row: int) -> None:
    borde = Side(style="thin", color="808080")
    border = Border(left=borde, right=borde, top=borde, bottom=borde)
    fuente_header = Font(bold=True, color="000000")
    fuente_titulo = Font(bold=True, color="000000", size=12)

    ws["A1"].font = fuente_titulo
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = FILL_TITULO_GRIS
    ws["A1"].border = border

    for row in range(2, 4):
        for col in range(1, 12):
            cell = ws.cell(row, col)
            cell.font = fuente_header
            cell.fill = FILL_TITULO_GRIS
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(4, total_row + 1):
        for col in range(1, 12):
            cell = ws.cell(row, col)
            cell.fill = FILL_BLANCO
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row == total_row:
                cell.font = fuente_header
            if col >= 3:
                cell.number_format = FORMATO_PESOS

    ws.cell(note_row, 2).alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    ws.cell(note_row, 2).font = Font(italic=True, color="666666")

    widths = {
        "A": 6,
        "B": 22,
        "C": 18,
        "D": 18,
        "E": 22,
        "F": 18,
        "G": 18,
        "H": 24,
        "I": 18,
        "J": 18,
        "K": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 42
    for row in range(4, total_row + 1):
        ws.row_dimensions[row].height = 24
    ws.row_dimensions[note_row].height = 28
    ws.freeze_panes = "A4"


def _crear_hoja_depurados(
    wb,
    filas_depurados: Iterable[FilaDepurados | dict],
    fecha_corte: datetime | date,
) -> None:
    ws = wb.create_sheet("DEPURADOS")
    f = _fecha_datetime(fecha_corte) or datetime.now()
    filas_ordenadas = ordenar_filas_depurados(filas_depurados)
    headers = [
        "N°",
        "FDL",
        "CANTIDAD INICIAL DE CONTRATOS",
        f"DEPURADOS A {f.day} DE {_nombre_mes_es(f)}",
        "PENDIENTES POR DEPURAR",
        "% DEPURADO",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(2, col).value = header

    for idx, fila in enumerate(filas_ordenadas, start=3):
        ws.cell(idx, 1).value = fila.numero
        ws.cell(idx, 2).value = fila.localidad
        ws.cell(idx, 3).value = fila.cantidad_inicial
        ws.cell(idx, 4).value = fila.depurados
        ws.cell(idx, 5).value = f"=C{idx}-D{idx}"
        ws.cell(idx, 6).value = f"=D{idx}/C{idx}"

    total_row = 3 + len(filas_ordenadas)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    ws.cell(total_row, 1).value = "TOTAL"
    first_data = 3
    last_data = max(first_data, total_row - 1)
    if filas_ordenadas:
        ws.cell(total_row, 3).value = f"=SUM(C{first_data}:C{last_data})"
        ws.cell(total_row, 4).value = f"=SUM(D{first_data}:D{last_data})"
        ws.cell(total_row, 5).value = f"=SUM(E{first_data}:E{last_data})"
        ws.cell(total_row, 6).value = f"=D{total_row}/C{total_row}"
    else:
        for col in range(3, 7):
            ws.cell(total_row, col).value = 0

    _formatear_hoja_depurados(ws, total_row)


def _formatear_hoja_depurados(ws, total_row: int) -> None:
    borde = Side(style="thin", color="808080")
    border = Border(left=borde, right=borde, top=borde, bottom=borde)
    fuente_header = Font(bold=True, color="000000")

    for col in range(1, 7):
        cell = ws.cell(2, col)
        cell.font = fuente_header
        cell.fill = FILL_TITULO_GRIS
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(3, total_row + 1):
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.fill = FILL_BLANCO
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row == total_row:
                cell.font = fuente_header
            if col == 6:
                cell.number_format = "0.00%"

    widths = {
        "A": 7,
        "B": 24,
        "C": 20,
        "D": 20,
        "E": 20,
        "F": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[2].height = 34
    for row in range(3, total_row + 1):
        ws.row_dimensions[row].height = 22
    ws.freeze_panes = "A3"


def _crear_hoja_cps_pn(
    wb,
    filas_cps_pn: Iterable[FilaCpsPn | dict],
    fecha_corte: datetime | date,
) -> None:
    ws = wb.create_sheet("CPS PN")
    f = _fecha_datetime(fecha_corte) or datetime.now()
    filas_ordenadas = ordenar_filas_cps_pn(filas_cps_pn)
    headers = [
        "N°",
        "FDL",
        "CANTIDAD INICIAL DE CPS PN",
        f"CPS PN DEPURADOS A {f.day} DE {_nombre_mes_es(f)}",
        "PENDIENTES POR DEPURAR",
        "% DEPURADO",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(2, col).value = header

    for idx, fila in enumerate(filas_ordenadas, start=3):
        ws.cell(idx, 1).value = fila.numero
        ws.cell(idx, 2).value = fila.localidad
        ws.cell(idx, 3).value = fila.cantidad_inicial
        ws.cell(idx, 4).value = fila.depurados
        ws.cell(idx, 5).value = f"=C{idx}-D{idx}"
        ws.cell(idx, 6).value = f"=D{idx}/C{idx}"

    total_row = 3 + len(filas_ordenadas)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    ws.cell(total_row, 1).value = "TOTAL"
    first_data = 3
    last_data = max(first_data, total_row - 1)
    if filas_ordenadas:
        ws.cell(total_row, 3).value = f"=SUM(C{first_data}:C{last_data})"
        ws.cell(total_row, 4).value = f"=SUM(D{first_data}:D{last_data})"
        ws.cell(total_row, 5).value = f"=SUM(E{first_data}:E{last_data})"
        ws.cell(total_row, 6).value = f"=D{total_row}/C{total_row}"
    else:
        for col in range(3, 7):
            ws.cell(total_row, col).value = 0

    _formatear_hoja_depurados(ws, total_row)


def _crear_hoja_por_depurar_vigencia(
    wb,
    filas_por_depurar_vigencia: Iterable[FilaPorDepurarVigencia | dict],
    fecha_corte: datetime | date,
) -> None:
    ws = wb.create_sheet("POR DEPURAR x VIGENCIA")
    f = _fecha_datetime(fecha_corte) or datetime.now()
    filas_ordenadas = ordenar_filas_por_depurar_vigencia(filas_por_depurar_vigencia)
    vigencias = list(range(2011, f.year))

    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws["A1"] = "N°"
    ws["B1"] = "FDL"

    col = 3
    columnas_inicial: list[int] = []
    columnas_saldo: list[int] = []
    for vigencia in vigencias:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws.cell(1, col).value = vigencia
        ws.cell(2, col).value = "INICIAL"
        ws.cell(2, col + 1).value = f
        columnas_inicial.append(col)
        columnas_saldo.append(col + 1)
        col += 2

    col_sin_info = col
    ws.merge_cells(start_row=1, start_column=col_sin_info, end_row=1, end_column=col_sin_info + 1)
    ws.cell(1, col_sin_info).value = "Sin info"
    ws.cell(2, col_sin_info).value = "INICIAL"
    ws.cell(2, col_sin_info + 1).value = f
    columnas_inicial.append(col_sin_info)
    columnas_saldo.append(col_sin_info + 1)

    col_total = col_sin_info + 2
    ws.merge_cells(start_row=1, start_column=col_total, end_row=1, end_column=col_total + 1)
    ws.cell(1, col_total).value = "TOTAL"
    ws.cell(2, col_total).value = "INICIAL"
    ws.cell(2, col_total + 1).value = f

    for row_idx, fila in enumerate(filas_ordenadas, start=3):
        ws.cell(row_idx, 1).value = fila.numero
        ws.cell(row_idx, 2).value = fila.localidad
        tiene_inicial = False
        tiene_saldo = False
        for idx, vigencia in enumerate(vigencias):
            col_ini = 3 + (idx * 2)
            inicial, con_saldo = fila.conteos_por_vigencia.get(vigencia, (0, 0))
            ws.cell(row_idx, col_ini).value = inicial or None
            ws.cell(row_idx, col_ini + 1).value = con_saldo or None
            tiene_inicial = tiene_inicial or inicial > 0
            tiene_saldo = tiene_saldo or con_saldo > 0
        inicial_refs = [f"{get_column_letter(c)}{row_idx}" for c in columnas_inicial]
        saldo_refs = [f"{get_column_letter(c)}{row_idx}" for c in columnas_saldo]
        if tiene_inicial:
            ws.cell(row_idx, col_total).value = "=" + "+".join(inicial_refs)
        if tiene_saldo:
            ws.cell(row_idx, col_total + 1).value = "=" + "+".join(saldo_refs)

    total_row = 3 + len(filas_ordenadas)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    ws.cell(total_row, 1).value = "TOTAL"
    first_data = 3
    last_data = max(first_data, total_row - 1)
    for col_idx in range(3, col_total + 2):
        letter = get_column_letter(col_idx)
        columna_tiene_datos = any(
            not _es_celda_vacia(ws.cell(row_idx, col_idx).value)
            for row_idx in range(first_data, total_row)
        )
        if filas_ordenadas and columna_tiene_datos:
            ws.cell(total_row, col_idx).value = (
                f"=SUM({letter}{first_data}:{letter}{last_data})"
            )

    _formatear_hoja_por_depurar_vigencia(ws, total_row, col_total + 1)


def _formatear_hoja_por_depurar_vigencia(
    ws,
    total_row: int,
    ultima_columna: int,
) -> None:
    borde = Side(style="thin", color="808080")
    border = Border(left=borde, right=borde, top=borde, bottom=borde)
    fuente_header = Font(bold=True, color="000000")

    for row in range(1, 3):
        for col in range(1, ultima_columna + 1):
            cell = ws.cell(row, col)
            cell.font = fuente_header
            cell.fill = FILL_TITULO_GRIS
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row == 2 and col >= 4 and (col - 4) % 2 == 0:
                cell.number_format = "dd/mm/yyyy"

    for row in range(3, total_row + 1):
        for col in range(1, ultima_columna + 1):
            cell = ws.cell(row, col)
            cell.fill = FILL_BLANCO
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row == total_row:
                cell.font = fuente_header

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 24
    for col in range(3, ultima_columna + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 30
    for row in range(3, total_row + 1):
        ws.row_dimensions[row].height = 22
    ws.freeze_panes = "C3"


def _crear_hoja_con_perdida(
    wb,
    filas_localidades: Iterable[FilaLibFen | dict],
    filas_con_perdida: Iterable[FilaConPerdida | dict],
    fecha_corte: datetime | date,
) -> None:
    f = _fecha_datetime(fecha_corte) or datetime.now()
    mes = _nombre_mes_es(f)
    titulo = f"CONTRATOS CON PÉRDIDA DE COMPETENCIA EN {mes} DE {f.year}"
    _crear_hoja_perdida_competencia(
        wb,
        "CON PÉRDIDA",
        titulo,
        filas_localidades,
        filas_con_perdida,
        combinar_localidad=True,
    )


def _crear_hoja_proximos_a_perder(
    wb,
    filas_localidades: Iterable[FilaLibFen | dict],
    filas_proximos_a_perder: Iterable[FilaConPerdida | dict],
    fecha_corte: datetime | date,
) -> None:
    inicio, fin = _rango_proximos_a_perder(fecha_corte)
    titulo = (
        "CONTRATOS PRÓXIMOS A PERDER COMPETENCIA "
        f"{_nombre_mes_es(inicio)} {inicio.year} - {_nombre_mes_es(fin)} {fin.year}"
    )
    _crear_hoja_perdida_competencia(
        wb,
        "PRÓXIMOS A PERDER",
        titulo,
        filas_localidades,
        filas_proximos_a_perder,
        combinar_localidad=True,
    )


def _crear_hoja_perdida_competencia(
    wb,
    nombre_hoja: str,
    titulo: str,
    filas_localidades: Iterable[FilaLibFen | dict],
    filas_perdida: Iterable[FilaConPerdida | dict],
    combinar_localidad: bool = False,
) -> None:
    ws = wb.create_sheet(nombre_hoja)
    ws.merge_cells("A2:H2")
    ws["A2"] = titulo
    headers = [
        "N°",
        "FDL",
        "CONTRATISTA",
        "No. CTO",
        "VIGENCIA ",
        "FECHA  FINALIZACION ",
        "FECHA DE  PERDIDA DE  COMPETENCIA ",
        "MONTO ",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(3, col).value = header

    por_localidad: dict[int, list[FilaConPerdida]] = {}
    for fila in ordenar_filas_con_perdida(filas_perdida):
        por_localidad.setdefault(fila.numero, []).append(fila)

    fila_excel = 4
    contratos_reportados = 0
    monto_total = 0.0
    for loc in ordenar_filas_lib_y_fen(filas_localidades):
        filas_loc = por_localidad.get(loc.numero, [])
        if not filas_loc:
            ws.cell(fila_excel, 1).value = loc.numero
            ws.cell(fila_excel, 2).value = loc.localidad
            for col in range(3, 9):
                ws.cell(fila_excel, col).value = "N/A"
            fila_excel += 1
            continue
        inicio_localidad = fila_excel
        for i, fila in enumerate(filas_loc):
            if i == 0:
                ws.cell(fila_excel, 1).value = fila.numero
                ws.cell(fila_excel, 2).value = fila.localidad
            ws.cell(fila_excel, 3).value = fila.contratista
            ws.cell(fila_excel, 4).value = fila.numero_contrato
            ws.cell(fila_excel, 5).value = fila.vigencia
            ws.cell(fila_excel, 6).value = fila.fecha_finalizacion
            ws.cell(fila_excel, 7).value = fila.fecha_perdida_competencia
            ws.cell(fila_excel, 8).value = fila.monto
            contratos_reportados += 1
            monto_total += fila.monto
            fila_excel += 1
        fin_localidad = fila_excel - 1
        if combinar_localidad and fin_localidad > inicio_localidad:
            ws.merge_cells(
                start_row=inicio_localidad,
                start_column=1,
                end_row=fin_localidad,
                end_column=1,
            )
            ws.merge_cells(
                start_row=inicio_localidad,
                start_column=2,
                end_row=fin_localidad,
                end_column=2,
            )
            ws.cell(inicio_localidad, 1).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            ws.cell(inicio_localidad, 2).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    total_row = fila_excel
    ws.cell(total_row, 1).value = "TOTAL"
    ws.cell(total_row, 4).value = contratos_reportados
    ws.cell(total_row, 8).value = monto_total
    _formatear_hoja_con_perdida(ws, total_row)


def _formatear_hoja_con_perdida(ws, ultima_fila: int) -> None:
    borde = Side(style="thin", color="808080")
    border = Border(left=borde, right=borde, top=borde, bottom=borde)
    fuente_header = Font(bold=True, color="000000")
    fuente_titulo = Font(bold=True, color="000000", size=12)

    ws["A2"].font = fuente_titulo
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].fill = FILL_TITULO_GRIS
    ws["A2"].border = border
    for col in range(1, 9):
        cell = ws.cell(3, col)
        cell.font = fuente_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = FILL_TITULO_GRIS
        cell.border = border

    for row in ws.iter_rows(min_row=4, max_row=ultima_fila, min_col=1, max_col=8):
        for cell in row:
            cell.fill = FILL_BLANCO
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row == ultima_fila:
                cell.font = fuente_header

    for row in range(4, ultima_fila + 1):
        ws[f"F{row}"].number_format = "dd/mm/yyyy"
        ws[f"G{row}"].number_format = "dd/mm/yyyy"
        ws[f"H{row}"].number_format = FORMATO_PESOS
    widths = {
        "A": 8,
        "B": 20,
        "C": 42,
        "D": 13,
        "E": 12,
        "F": 18,
        "G": 22,
        "H": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 38
    for row in range(4, ultima_fila + 1):
        ws.row_dimensions[row].height = 24
    ws.freeze_panes = "A4"
