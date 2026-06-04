"""Pestaña Liquidados con saldo: saldo mensual desde Matriz con apropiación (k4)."""

from __future__ import annotations

from datetime import date, datetime

from openpyxl.styles import PatternFill

from constantes import HOJAS_LIQUIDADOS_CON_SALDO
from cxp_cruce import (
    _celda_para_escribir,
    _celda_tiene_formula,
    _copiar_estilo_celda,
    _copiar_estilo_celda_sin_relleno,
    _fila_encabezado_hoja_datos,
    _fila_inicio_datos_hoja,
    _fila_tiene_contratista,
    _hoja_tiene_filas_contratista,
    _indice_columna_en_hoja,
    _normalizar,
    _ultima_columna_con_datos,
    clave_cuatro,
    _fecha_datetime,
    preparar_indice_matriz,
)
from hoja_suspendidos import (
    _aplicar_encabezados_saldo_mes_alternos,
    _aplicar_estilo_total_saldo,
    _aplicar_tope_mes_anterior,
    _columna_saldo_mes_en_hoja,
    _es_columna_saldo_mes,
    _escribir_total_estado,
    _listar_columnas_saldo_mes,
    _normalizar_titulos_mes_cortos,
    _rango_filas_datos_seguimiento,
    _ultima_fila_contratista,
    _autoajustar_columna_suspendidos,
    _copiar_ancho_columna,
    _copiar_estilo_columna,
    titulo_saldo_corte,
    titulo_saldo_suspendidos,
)


def _fila_suma_liquidados(ws) -> int:
    """Primera fila de totales al pie: suma de todos los saldos del mes."""
    ultima = _ultima_fila_contratista(ws)
    if ultima is not None:
        return ultima + 1
    return _fila_inicio_datos_hoja(ws)


def _fila_conteo_liquidados(ws) -> int:
    """Debajo de la suma: cantidad de contratos con saldo distinto de $0."""
    return _fila_suma_liquidados(ws) + 1


def _limpiar_totales_legacy_arriba(ws, col_saldo: int) -> None:
    """Quita suma/conteo antiguos en filas 1-2 si la plantilla los tenía arriba."""
    from cxp_cruce import FILA_CONTEO_CONTRATOS, FILA_SUMA_CONTRATOS

    for fila in (FILA_SUMA_CONTRATOS, FILA_CONTEO_CONTRATOS):
        celda = _celda_para_escribir(ws, fila, col_saldo)
        if _celda_tiene_formula(celda):
            continue
        celda.value = None
        celda.fill = PatternFill(fill_type=None)


def resolver_hoja_liquidados_con_saldo(nombres_hojas: list[str]) -> str | None:
    for candidato in HOJAS_LIQUIDADOS_CON_SALDO:
        if candidato in nombres_hojas:
            return candidato
    for nombre in nombres_hojas:
        norm = _normalizar(nombre)
        if norm in ("liquidados con saldo", "liquidado con saldo"):
            return nombre
    return None


def preparar_mapa_k4_saldo_matriz(df_matriz, localidad: str) -> dict[str, float | None]:
    """Saldo por nombre+contrato+año+apropiación (suma si hay varias filas k4 en Matriz)."""
    from cxp_cruce import _suma_saldos_grupo_matriz

    df_loc, _, _ = preparar_indice_matriz(df_matriz, localidad)
    mapa: dict[str, float | None] = {}
    for k4, grupo in df_loc.groupby("_k4", sort=False):
        mapa[k4] = _suma_saldos_grupo_matriz(grupo["_saldo"])
    return mapa


def _columna_saldo_mes_anterior(
    ws,
    fecha: datetime | date,
    col_saldo: int,
) -> int | None:
    """
    Columna de saldo del mes inmediatamente anterior (a la izquierda en la hoja).
    """
    del fecha
    cols = _listar_columnas_saldo_mes(ws)
    for i, (col, _) in enumerate(cols):
        if col == col_saldo and i > 0:
            return cols[i - 1][0]

    fila_hdr = _fila_encabezado_hoja_datos(ws)
    for col_cand in range(col_saldo - 1, 0, -1):
        val = ws.cell(fila_hdr, col_cand).value
        if val is None or not str(val).strip():
            continue
        if _es_columna_saldo_mes(str(val).strip()):
            return col_cand
    return None


def _leer_conteo_reportado_pie_liquidados(
    ws,
    col_saldo: int | None,
) -> float | None:
    """Valor del conteo al pie (fila debajo de la suma) en esa columna."""
    if not col_saldo:
        return None
    celda = _celda_para_escribir(ws, _fila_conteo_liquidados(ws), col_saldo)
    if celda.value in (None, ""):
        return None
    try:
        return float(celda.value)
    except (TypeError, ValueError):
        return None


def _asegurar_columna_saldo_mes(
    ws,
    fecha: datetime | date,
) -> tuple[int, int | None]:
    """Devuelve (col_saldo_actual, col_saldo_mes_anterior)."""
    _normalizar_titulos_mes_cortos(ws, fecha)
    col_saldo = _columna_saldo_mes_en_hoja(ws, fecha)
    col_prev = None
    if col_saldo is not None:
        col_prev = _columna_saldo_mes_anterior(ws, fecha, col_saldo)

    if col_saldo is None:
        siguiente = _ultima_columna_con_datos(ws) + 1
        col_saldo = siguiente
        col_prev = _columna_saldo_mes_anterior(ws, fecha, col_saldo)

    fila_hdr = _fila_encabezado_hoja_datos(ws)

    fila_suma = _fila_suma_liquidados(ws)
    fila_conteo = _fila_conteo_liquidados(ws)

    if col_prev:
        _copiar_estilo_columna(
            ws,
            col_prev,
            col_saldo,
            sin_relleno_desde_fila=_fila_inicio_datos_hoja(ws),
        )
        _copiar_estilo_celda(
            ws.cell(fila_suma, col_prev),
            _celda_para_escribir(ws, fila_suma, col_saldo),
        )
        _copiar_estilo_celda(
            ws.cell(fila_conteo, col_prev),
            _celda_para_escribir(ws, fila_conteo, col_saldo),
        )
    else:
        col_sf = _indice_columna_en_hoja(ws, "SALDO FINAL", "Saldo Final")
        if col_sf:
            _copiar_estilo_columna(ws, col_sf, col_saldo)
            _copiar_estilo_celda(
                ws.cell(fila_suma, col_sf),
                _celda_para_escribir(ws, fila_suma, col_saldo),
            )
            _copiar_estilo_celda(
                ws.cell(fila_conteo, col_sf),
                _celda_para_escribir(ws, fila_conteo, col_saldo),
            )

    _limpiar_totales_legacy_arriba(ws, col_saldo)

    if col_prev:
        _copiar_ancho_columna(ws, col_prev, col_saldo)

    titulo_mes = titulo_saldo_corte(fecha)
    _celda_para_escribir(ws, fila_hdr, col_saldo).value = titulo_mes
    _aplicar_encabezados_saldo_mes_alternos(ws, fecha)
    _celda_para_escribir(ws, fila_hdr, col_saldo).value = titulo_mes
    return col_saldo, col_prev


def _saldo_es_cero(valor) -> bool:
    if valor is None or valor == "":
        return True
    try:
        return abs(float(valor)) < 1e-9
    except (TypeError, ValueError):
        return True


def _suma_saldos_columna(ws, col_saldo: int) -> float:
    fila_ini, fila_fin = _rango_filas_datos_seguimiento(ws)
    suma = 0.0
    for fila in range(fila_ini, fila_fin + 1):
        raw = ws.cell(fila, col_saldo).value
        if raw is None or raw == "":
            continue
        try:
            suma += float(raw)
        except (TypeError, ValueError):
            continue
    return suma


def _conteo_con_saldo_distinto_cero(ws, col_saldo: int) -> int:
    fila_ini, fila_fin = _rango_filas_datos_seguimiento(ws)
    total = 0
    for fila in range(fila_ini, fila_fin + 1):
        if not _saldo_es_cero(ws.cell(fila, col_saldo).value):
            total += 1
    return total


def _actualizar_resumen_liquidados(
    ws,
    col_saldo: int,
    col_prev_saldo: int | None,
) -> tuple[int, int]:
    """Suma al pie; conteo con tope = celda de conteo de la columna inmediatamente anterior."""
    suma = _suma_saldos_columna(ws, col_saldo)
    conteo_real = _conteo_con_saldo_distinto_cero(ws, col_saldo)
    conteo_prev = _leer_conteo_reportado_pie_liquidados(ws, col_prev_saldo)
    if conteo_prev is None and col_prev_saldo:
        conteo_prev = float(_conteo_con_saldo_distinto_cero(ws, col_prev_saldo))
    conteo_mostrar, _ = _aplicar_tope_mes_anterior(conteo_real, conteo_prev)
    conteo_mostrar = int(conteo_mostrar)

    fila_suma = _fila_suma_liquidados(ws)
    fila_conteo = _fila_conteo_liquidados(ws)
    celda_suma = _celda_para_escribir(ws, fila_suma, col_saldo)
    celda_conteo = _celda_para_escribir(ws, fila_conteo, col_saldo)

    if not _celda_tiene_formula(celda_suma):
        celda_suma.value = suma
        _aplicar_estilo_total_saldo(
            celda_suma, ws, col_saldo, col_prev_saldo, fila_suma
        )
    if not _celda_tiene_formula(celda_conteo):
        _escribir_total_estado(celda_conteo, conteo_mostrar)
        if col_prev_saldo:
            ref = _celda_para_escribir(ws, fila_conteo, col_prev_saldo)
            if ref.has_style:
                _copiar_estilo_celda_sin_relleno(ref, celda_conteo)

    return conteo_real, conteo_mostrar


def actualizar_hoja_liquidados_con_saldo(
    ws,
    mapa_k4: dict[str, float],
    fecha: datetime | date,
) -> list[str]:
    """
    Escribe SALDO MES con saldo Matriz por k4 (con apropiación).
    Totales al pie: suma en la fila tras el último contrato; conteo debajo (saldo ≠ $0).
    """
    if not _hoja_tiene_filas_contratista(ws):
        return []

    advertencias: list[str] = []
    col_saldo, col_prev = _asegurar_columna_saldo_mes(ws, fecha)

    col_nombre = _indice_columna_en_hoja(ws, "NOMBRE CONTRATISTA")
    col_cto = _indice_columna_en_hoja(
        ws, "No. de Cto", "Número Contrato", "Numero Contrato"
    )
    col_anio = _indice_columna_en_hoja(
        ws,
        "AÑO SUSCRIPCIÓN",
        "ANO SUSCRIPCION",
        "Año Suscripción",
        "Ano Suscripcion",
    )
    col_aprop = _indice_columna_en_hoja(
        ws, "APROPIACION DISPONIBLE", "Apropiación", "Apropiacion"
    )
    if not all([col_nombre, col_cto, col_anio, col_aprop]):
        raise ValueError(
            "Liquidados con saldo: faltan columnas NOMBRE CONTRATISTA, "
            "No. de Cto, AÑO SUSCRIPCIÓN o APROPIACION DISPONIBLE."
        )

    fila_ini = _fila_inicio_datos_hoja(ws)

    for fila in range(fila_ini, ws.max_row + 1):
        if not _fila_tiene_contratista(ws, fila, col_nombre):
            continue

        nombre = ws.cell(fila, col_nombre).value
        contrato = ws.cell(fila, col_cto).value
        anio = ws.cell(fila, col_anio).value
        aprop = ws.cell(fila, col_aprop).value
        k4 = clave_cuatro(nombre, contrato, anio, aprop)
        saldo = mapa_k4.get(k4)

        celda_saldo = _celda_para_escribir(ws, fila, col_saldo)
        if col_prev:
            _copiar_estilo_celda_sin_relleno(ws.cell(fila, col_prev), celda_saldo)
        celda_saldo.value = None if saldo is None else saldo
        celda_saldo.fill = PatternFill(fill_type=None)

    conteo_real, conteo_mostrar = _actualizar_resumen_liquidados(
        ws, col_saldo, col_prev
    )
    if col_prev and conteo_real > conteo_mostrar:
        advertencias.append(
            f"Liquidados con saldo: el conteo real ({conteo_real}) supera al mes anterior "
            f"({conteo_mostrar}); en el total se dejó el valor de la columna anterior."
        )

    _autoajustar_columna_suspendidos(ws, col_saldo, titulo_saldo_suspendidos(fecha))

    return advertencias


__all__ = [
    "actualizar_hoja_liquidados_con_saldo",
    "preparar_mapa_k4_saldo_matriz",
    "resolver_hoja_liquidados_con_saldo",
    "_fila_suma_liquidados",
    "_fila_conteo_liquidados",
]
