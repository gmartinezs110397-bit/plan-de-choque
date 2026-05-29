"""Actualización de la pestaña Próximos a perder (misma estructura mensual que Suspendidos)."""

from __future__ import annotations

from datetime import date, datetime
from typing import Any

from openpyxl.styles import PatternFill

from constantes import HOJAS_PROXIMOS_A_PERDER
from cxp_cruce import (
    FILA_CONTEO_CONTRATOS,
    FILA_SUMA_CONTRATOS,
    _celda_tiene_formula,
    _copiar_estilo_celda,
    _fila_inicio_datos_contratos,
    _fila_tiene_contratista,
    _indice_columna_en_hoja,
    _normalizar,
    clave_tres,
)
from hoja_suspendidos import (
    FILL_AMARILLO,
    FILL_VERDE_NEON,
    _aplicar_tope_mes_anterior,
    _asegurar_columnas_mes,
    _autoajustar_columna_suspendidos,
    _leer_total_reportado_mes_anterior,
    _suma_saldos_columna,
    preparar_mapa_k3_saldo_estado,
    titulo_estado_suspendidos,
    titulo_saldo_suspendidos,
)


def resolver_hoja_proximos_a_perder(nombres_hojas: list[str]) -> str | None:
    for candidato in HOJAS_PROXIMOS_A_PERDER:
        if candidato in nombres_hojas:
            return candidato
    for nombre in nombres_hojas:
        norm = _normalizar(nombre)
        if norm in ("proximos a perder", "proximo a perder"):
            return nombre
    return None


def _es_liquidado(valor) -> bool:
    return _normalizar(str(valor or "")) == "liquidado"


def _saldo_es_cero(valor) -> bool:
    if valor is None or valor == "":
        return True
    try:
        return abs(float(valor)) < 1e-9
    except (TypeError, ValueError):
        return True


def _resolver_relleno_estado_proximos(
    prev_estado,
    prev_celda,
    curr_estado,
    saldo,
) -> PatternFill | None:
    """
    Verde: pasa a LIQUIDADO.
    Amarillo: no liquidado y saldo cero.
    Sin color: el resto (cuenta para el total de contratos).
    """
    del prev_celda
    prev_txt = _normalizar(str(prev_estado or ""))
    curr_txt = _normalizar(str(curr_estado or ""))
    if curr_txt == "liquidado" and prev_txt != "liquidado":
        return FILL_VERDE_NEON
    if curr_txt != "liquidado" and _saldo_es_cero(saldo):
        return FILL_AMARILLO
    return None


def _cuenta_para_total_proximos(estado, saldo) -> bool:
    """Sin color en estado: no liquidado y saldo distinto de cero."""
    if _es_liquidado(estado):
        return False
    if _saldo_es_cero(saldo):
        return False
    return True


def _conteo_sin_color_proximos(ws, col_estado: int, col_saldo: int) -> int:
    col_nombre = _indice_columna_en_hoja(ws, "NOMBRE CONTRATISTA")
    fila_ini = _fila_inicio_datos_contratos()
    total = 0
    for fila in range(fila_ini, ws.max_row + 1):
        if col_nombre and not _fila_tiene_contratista(ws, fila, col_nombre):
            continue
        estado = ws.cell(fila, col_estado).value
        saldo = ws.cell(fila, col_saldo).value
        if _cuenta_para_total_proximos(estado, saldo):
            total += 1
    return total


def _conteo_proximos_mes_anterior(
    ws,
    col_prev_estado: int | None,
    col_prev_saldo: int | None,
) -> float | None:
    """Total reportado en fila 1 del mes anterior (o recálculo si la celda está vacía)."""
    if not col_prev_estado:
        return None
    reportado = _leer_total_reportado_mes_anterior(
        ws, FILA_CONTEO_CONTRATOS, col_prev_estado
    )
    if reportado is not None:
        return reportado
    if col_prev_saldo:
        return float(_conteo_sin_color_proximos(ws, col_prev_estado, col_prev_saldo))
    return None


def _actualizar_resumen_proximos(
    ws,
    col_saldo: int,
    col_estado: int,
    col_prev_saldo: int | None,
    col_prev_estado: int | None,
) -> tuple[int, int]:
    """Fila 1: contratos sin color (tope vs mes anterior). Fila 2: suma de todos los saldos."""
    conteo_real = _conteo_sin_color_proximos(ws, col_estado, col_saldo)
    conteo_prev = _conteo_proximos_mes_anterior(
        ws, col_prev_estado, col_prev_saldo
    )
    conteo_mostrar, _ = _aplicar_tope_mes_anterior(conteo_real, conteo_prev)
    conteo_mostrar = int(conteo_mostrar)

    suma = _suma_saldos_columna(ws, col_saldo)

    celda_conteo = ws.cell(FILA_CONTEO_CONTRATOS, col_estado)
    celda_suma = ws.cell(FILA_SUMA_CONTRATOS, col_saldo)

    if col_prev_estado:
        _copiar_estilo_celda(
            ws.cell(FILA_CONTEO_CONTRATOS, col_prev_estado),
            celda_conteo,
        )
    if col_prev_saldo:
        _copiar_estilo_celda(
            ws.cell(FILA_SUMA_CONTRATOS, col_prev_saldo),
            celda_suma,
        )

    if not _celda_tiene_formula(celda_conteo):
        celda_conteo.value = conteo_mostrar
        celda_conteo.fill = PatternFill(fill_type=None)

    if not _celda_tiene_formula(celda_suma):
        celda_suma.value = suma

    return conteo_real, conteo_mostrar


def actualizar_hoja_proximos_a_perder(
    ws,
    mapa_k3: dict[str, dict[str, Any]],
    fecha: datetime | date,
) -> list[str]:
    """
    Escribe SALDO MES y ESTADO ACTUAL MES desde Matriz.
    Formato y encabezados de mes copiados del mes anterior (igual que Suspendidos).
    Verde: pasa a LIQUIDADO. Amarillo: no liquidado y saldo cero.
    Conteo: filas sin color (no liquidado y saldo ≠ 0); no puede superar el mes anterior.
    Suma: todos los saldos.
    """
    advertencias: list[str] = []
    col_saldo, col_estado, col_prev_saldo, col_prev_estado, _ = _asegurar_columnas_mes(
        ws, fecha
    )

    col_nombre = _indice_columna_en_hoja(ws, "NOMBRE CONTRATISTA")
    col_cto = _indice_columna_en_hoja(
        ws, "No. de Cto", "Número Contrato", "Numero Contrato"
    )
    col_anio = _indice_columna_en_hoja(
        ws,
        "AÑO SUSCRIPCIÓN",
        "ANO SUSCRIPCION",
        "Año Suscripción",
        "Ano Suscripcion",
    )
    if not all([col_nombre, col_cto, col_anio]):
        raise ValueError(
            "Próximos a perder: faltan columnas NOMBRE CONTRATISTA, No. de Cto o AÑO SUSCRIPCIÓN."
        )

    fila_ini = _fila_inicio_datos_contratos()

    for fila in range(fila_ini, ws.max_row + 1):
        if not _fila_tiene_contratista(ws, fila, col_nombre):
            continue

        nombre = ws.cell(fila, col_nombre).value
        contrato = ws.cell(fila, col_cto).value
        anio = ws.cell(fila, col_anio).value
        k3 = clave_tres(nombre, contrato, anio)
        datos = mapa_k3.get(k3, {"saldo": 0.0, "estado": ""})

        celda_saldo = ws.cell(fila, col_saldo)
        celda_estado = ws.cell(fila, col_estado)

        if col_prev_saldo:
            _copiar_estilo_celda(ws.cell(fila, col_prev_saldo), celda_saldo)
        celda_saldo.value = datos["saldo"]

        estado = datos.get("estado", "")
        prev_celda = None
        prev_estado = None
        if col_prev_estado:
            prev_celda = ws.cell(fila, col_prev_estado)
            prev_estado = prev_celda.value
            _copiar_estilo_celda(prev_celda, celda_estado)

        celda_estado.value = estado or None

        saldo = datos["saldo"]
        relleno = _resolver_relleno_estado_proximos(
            prev_estado, prev_celda, estado, saldo
        )
        if relleno:
            celda_estado.fill = relleno

    conteo_real, conteo_mostrar = _actualizar_resumen_proximos(
        ws, col_saldo, col_estado, col_prev_saldo, col_prev_estado
    )

    if col_prev_estado and conteo_real > conteo_mostrar:
        advertencias.append(
            f"Próximos a perder: el conteo real ({conteo_real}) supera al mes anterior "
            f"({conteo_mostrar}); en el total se dejó el valor del mes anterior."
        )

    _autoajustar_columna_suspendidos(ws, col_saldo, titulo_saldo_suspendidos(fecha))
    _autoajustar_columna_suspendidos(ws, col_estado, titulo_estado_suspendidos(fecha))

    return advertencias


__all__ = [
    "actualizar_hoja_proximos_a_perder",
    "preparar_mapa_k3_saldo_estado",
    "resolver_hoja_proximos_a_perder",
]
