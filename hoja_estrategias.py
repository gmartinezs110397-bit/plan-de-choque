"""Pestaña Estrategias: resumen de conteos y montos de las hojas de seguimiento."""

from __future__ import annotations

import re
from datetime import date, datetime

from openpyxl.styles import PatternFill

from constantes import HOJAS_ESTRATEGIAS
from cxp_cruce import (
    FILA_CONTEO_CONTRATOS,
    FILA_SUMA_CONTRATOS,
    _celda_tiene_formula,
    _indice_columna_corte_en_hoja,
    _normalizar,
    dia_fin_mes_corte,
    mes_nombre_corte,
    resolver_hoja_cruce_cxp,
)
from hoja_liquidados_con_saldo import (
    _fila_conteo_liquidados,
    _fila_suma_liquidados,
    resolver_hoja_liquidados_con_saldo,
)
from hoja_proximos_a_perder import resolver_hoja_proximos_a_perder
from hoja_suspendidos import (
    _fila_totales_seguimiento,
    _indice_columna_titulos,
    _titulos_estado_equivalentes,
    _titulos_saldo_equivalentes,
    resolver_hoja_suspendidos,
)
from hoja_tramites_sectores import resolver_hoja_tramites_sectores

COL_NOMBRE_PESTANA = 2  # B
COL_CONTRATOS = 5  # E
COL_MONTO = 6  # F
FILA_TITULOS = 3
FILA_INICIO_DATOS = 4

# Misma paleta que la plantilla de Estrategias
FILL_TITULOS_ESTRATEGIAS = PatternFill(fill_type="solid", fgColor="FFC000")
FILL_DATOS_ESTRATEGIAS = PatternFill(fill_type="solid", fgColor="00FA00")


def resolver_hoja_estrategias(nombres_hojas: list[str]) -> str | None:
    for candidato in HOJAS_ESTRATEGIAS:
        if candidato in nombres_hojas:
            return candidato
    for nombre in nombres_hojas:
        if _normalizar(nombre) == "estrategias":
            return nombre
    return None


def titulo_contratos_estrategias(fecha: datetime | date) -> str:
    """Siempre último día del mes de ejecución (ej. 29-may → 31 de mayo)."""
    dia = dia_fin_mes_corte(fecha)
    mes = mes_nombre_corte(fecha)
    return f"No. de contratos {dia} de {mes}"


def titulo_monto_total_estrategias(fecha: datetime | date) -> str:
    dia = dia_fin_mes_corte(fecha)
    mes = mes_nombre_corte(fecha)
    return f"Monto Total {dia} de {mes}"


def _valor_numerico(valor) -> float | int | None:
    if valor is None or valor == "":
        return None
    try:
        n = float(valor)
        if abs(n - round(n)) < 1e-9:
            return int(round(n))
        return n
    except (TypeError, ValueError):
        return None


def _escribir_valor(celda, valor) -> None:
    if _celda_tiene_formula(celda):
        return
    celda.value = valor


def _aplicar_relleno_datos_estrategias(ws, filas: list[int], fila_total: int | None) -> None:
    """Verde en columnas E/F de cada fila de detalle y en la fila Total (como la plantilla)."""
    objetivos = list(filas)
    if fila_total is not None:
        objetivos.append(fila_total)
    for fila in objetivos:
        for col in (COL_CONTRATOS, COL_MONTO):
            celda = ws.cell(fila, col)
            celda.fill = FILL_DATOS_ESTRATEGIAS


def _leer_totales_hoja_par(ws, fecha: datetime | date) -> tuple[float | int | None, float | int | None]:
    """Conteo y suma en la última fila de cada columna del mes (Suspendidos, Próximos, Trámites)."""
    col_saldo = _indice_columna_titulos(ws, _titulos_saldo_equivalentes(fecha))
    col_estado = _indice_columna_titulos(ws, _titulos_estado_equivalentes(fecha))
    if not col_saldo or not col_estado:
        return None, None
    fila_tot = _fila_totales_seguimiento(ws)
    conteo = _valor_numerico(ws.cell(fila_tot, col_estado).value)
    suma = _valor_numerico(ws.cell(fila_tot, col_saldo).value)
    return conteo, suma


def _leer_totales_hoja_cps(
    ws,
    fecha: datetime | date,
) -> tuple[float | int | None, float | int | None]:
    """Cps/Caja por depurar: fila 1 = No. contratos, fila 2 = saldo total del mes."""
    col_corte, _ = _indice_columna_corte_en_hoja(ws, fecha)
    if not col_corte:
        return None, None
    conteo = _valor_numerico(ws.cell(FILA_CONTEO_CONTRATOS, col_corte).value)
    suma = _valor_numerico(ws.cell(FILA_SUMA_CONTRATOS, col_corte).value)
    return conteo, suma


def _leer_totales_hoja_liquidados(
    ws,
    fecha: datetime | date,
) -> tuple[float | int | None, float | int | None]:
    """Suma y conteo al pie de la columna saldo del mes (conteo debajo de la suma)."""
    col_saldo = _indice_columna_titulos(ws, _titulos_saldo_equivalentes(fecha))
    if not col_saldo:
        return None, None
    suma = _valor_numerico(ws.cell(_fila_suma_liquidados(ws), col_saldo).value)
    conteo = _valor_numerico(ws.cell(_fila_conteo_liquidados(ws), col_saldo).value)
    return conteo, suma


def _construir_totales_desde_libro(
    wb,
    nombres_hojas: list[str],
    fecha: datetime | date,
) -> dict[str, tuple[float | int | None, float | int | None]]:
    """Clave lógica -> (conteo contratos, monto total). Solo pestañas con contratos reales."""
    from cxp_cruce import _hoja_tiene_filas_contratista

    totales: dict[str, tuple[float | int | None, float | int | None]] = {}

    nombre = resolver_hoja_suspendidos(nombres_hojas)
    if nombre and _hoja_tiene_filas_contratista(wb[nombre]):
        totales["suspendidos"] = _leer_totales_hoja_par(wb[nombre], fecha)

    nombre = resolver_hoja_proximos_a_perder(nombres_hojas)
    if nombre and _hoja_tiene_filas_contratista(wb[nombre]):
        totales["proximos"] = _leer_totales_hoja_par(wb[nombre], fecha)

    nombre = resolver_hoja_tramites_sectores(nombres_hojas)
    if nombre and _hoja_tiene_filas_contratista(wb[nombre]):
        totales["tramites"] = _leer_totales_hoja_par(wb[nombre], fecha)

    nombre = resolver_hoja_liquidados_con_saldo(nombres_hojas)
    if nombre and _hoja_tiene_filas_contratista(wb[nombre]):
        totales["liquidados"] = _leer_totales_hoja_liquidados(wb[nombre], fecha)

    nombre = resolver_hoja_cruce_cxp(nombres_hojas)
    if nombre:
        totales["cps"] = _leer_totales_hoja_cps(wb[nombre], fecha)

    return totales


def _identificar_fuente_por_texto(texto: str) -> str | None:
    norm = _normalizar(texto)
    if not norm:
        return None
    if re.search(r"\btotal\b", norm) and "suspend" not in norm:
        return "_total"
    if "suspend" in norm:
        return "suspendidos"
    if "proxim" in norm and "perd" in norm:
        return "proximos"
    if "tramit" in norm and "sector" in norm:
        return "tramites"
    if "liquid" in norm and "saldo" in norm:
        return "liquidados"
    if "depur" in norm or ("cps" in norm and "caja" not in norm):
        return "cps"
    return None


def actualizar_hoja_estrategias(
    ws,
    fecha: datetime | date,
    totales_fuentes: dict[str, tuple[float | int | None, float | int | None]],
) -> list[str]:
    """
    Actualiza títulos en E3/F3 y valores en E/F según nombre en columna B.
    No agrega columnas; solo reemplaza valores (respeta fórmulas existentes).
    """
    advertencias: list[str] = []

    # Títulos fila 3 (amarillo); datos filas 4+ (verde)
    celda_te = ws.cell(FILA_TITULOS, COL_CONTRATOS, value=titulo_contratos_estrategias(fecha))
    celda_tm = ws.cell(FILA_TITULOS, COL_MONTO, value=titulo_monto_total_estrategias(fecha))
    celda_te.fill = FILL_TITULOS_ESTRATEGIAS
    celda_tm.fill = FILL_TITULOS_ESTRATEGIAS

    filas_datos: list[int] = []
    fila_total: int | None = None
    suma_conteos = 0.0
    suma_montos = 0.0
    hubo_conteo = False
    hubo_monto = False

    for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
        raw_b = ws.cell(fila, COL_NOMBRE_PESTANA).value
        if raw_b is None or not str(raw_b).strip():
            continue

        fuente = _identificar_fuente_por_texto(str(raw_b))
        if fuente == "_total":
            fila_total = fila
            continue

        if fuente is None or fuente not in totales_fuentes:
            continue

        conteo, monto = totales_fuentes[fuente]
        if conteo is None and monto is None:
            continue

        if conteo is not None:
            _escribir_valor(ws.cell(fila, COL_CONTRATOS), conteo)
            suma_conteos += float(conteo)
            hubo_conteo = True
        if monto is not None:
            _escribir_valor(ws.cell(fila, COL_MONTO), monto)
            suma_montos += float(monto)
            hubo_monto = True

        filas_datos.append(fila)

    if fila_total is None and filas_datos:
        fila_total = max(filas_datos) + 1

    if fila_total is not None:
        if hubo_conteo:
            valor_c = int(suma_conteos) if abs(suma_conteos - round(suma_conteos)) < 1e-9 else suma_conteos
            _escribir_valor(ws.cell(fila_total, COL_CONTRATOS), valor_c)
        if hubo_monto:
            _escribir_valor(ws.cell(fila_total, COL_MONTO), suma_montos)

    _aplicar_relleno_datos_estrategias(ws, filas_datos, fila_total)

    if not filas_datos:
        advertencias.append(
            "Estrategias: no se identificó ninguna fila en columna B "
            "(Suspendidos, Próximos a perder, Trámites sectores, Cps por depurar, "
            "Liquidados con saldo)."
        )

    return advertencias


def actualizar_estrategias_en_libro(
    wb,
    fecha: datetime | date,
) -> list[str]:
    nombres = list(wb.sheetnames)
    nombre_est = resolver_hoja_estrategias(nombres)
    if not nombre_est:
        return []
    totales = _construir_totales_desde_libro(wb, nombres, fecha)
    return actualizar_hoja_estrategias(wb[nombre_est], fecha, totales)


__all__ = [
    "actualizar_estrategias_en_libro",
    "actualizar_hoja_estrategias",
    "resolver_hoja_estrategias",
    "titulo_contratos_estrategias",
    "titulo_monto_total_estrategias",
]
