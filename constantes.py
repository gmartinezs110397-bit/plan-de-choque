"""Constantes compartidas (evita errores de importación por caché desactualizado)."""

from openpyxl.styles import PatternFill

COL_DESEMPATE_MANUAL = "Desempate manual"

# Verde «LIQUIDADO» — Estrategias, Suspendidos, Próximos, Trámites y UI
VERDE_LIQUIDADO = "00FF00"
VERDE_LIQUIDADO_CSS = "#00FF00"
VERDE_LIQUIDADO_CSS_HOVER = "#00DD00"
FILL_VERDE_LIQUIDADO = PatternFill(fill_type="solid", fgColor=VERDE_LIQUIDADO)
# Tonos antiguos al leer celdas ya coloreadas en archivos previos
VERDES_RELLENO_COMPAT = (
    VERDE_LIQUIDADO,
    "CCFF00",
    "00FA00",
    "39FF14",
    "92D050",
    "C6EFCE",
    "00B050",
)

# Amarillo de títulos («Saldo a 31 de mayo», encabezados de mes en seguimiento y Cps)
AMARILLO_TITULO = "FFD966"
AMARILLO_TITULO_CSS = "#FFD966"
# Amarillo solo en pestaña Estrategias (fila 3)
AMARILLO_ESTRATEGIAS = "FFC000"
FILL_AMARILLO_ESTRATEGIAS = PatternFill(fill_type="solid", fgColor=AMARILLO_ESTRATEGIAS)
# Amarillo en celdas de datos (saldo cero, transiciones de estado) — no es el de títulos
AMARILLO_DATOS = "FFFF00"
AMARILLOS_TITULO_COMPAT = (
    AMARILLO_TITULO,
    "FFF2CC",
    "FFC000",
    "FFEB9C",
    "FFE699",
    "FFF9C4",
    "F9D472",
    "FFFF00",
)

# Hoja donde se cruza Matriz → Contratos (el nombre varía según plantilla)
HOJAS_CRUCE_CXP = (
    "Cps por depurar",
    "Caja por depurar",
)

HOJAS_SUSPENDIDOS = (
    "Suspendidos",
    "SUSPENDIDOS",
)

HOJAS_PROXIMOS_A_PERDER = (
    "Próximos a perder",
    "Proximos a perder",
    "PROXIMOS A PERDER",
)

HOJAS_TRAMITES_SECTORES = (
    "Trámites sectores",
    "Tramites sectores",
    "TRAMITES SECTORES",
)

HOJAS_LIQUIDADOS_CON_SALDO = (
    "Liquidados con saldo",
    "Liquidados Con Saldo",
    "LIQUIDADOS CON SALDO",
)

HOJAS_ESTRATEGIAS = (
    "Estrategias",
    "ESTRATEGIAS",
)

# Otras pestañas del Excel Contratos plan de choque — reglas por definir
HOJAS_CONTRATOS_OTRAS: dict[str, str] = {}
