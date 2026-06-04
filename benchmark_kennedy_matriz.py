"""Mide tiempos Matriz Kennedy — sin importar app.py (evita Streamlit)."""
from __future__ import annotations

import sys
import time
import unicodedata
from io import BytesIO
from pathlib import Path

import msoffcrypto
import msoffcrypto.exceptions as ms_exceptions
import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

PWD = "1100"
SHEET = "MATRIZ OXP"
HEADER_PD = 6
FILA_HDR = HEADER_PD + 1
MATRIZ = Path(r"C:\Users\f1rac\Downloads\Matriz OxP FDLK 2026 MAYO - KENNEDY (1).xlsx")
CONTRATOS = Path(
    r"C:\Users\f1rac\Downloads\Copia de 08. Contratos plan de choque Kennedy 2026 — Mayo Final PRUEBA.xlsx"
)


def normalizar(texto: str) -> str:
    texto = texto.lower()
    texto = unicodedata.normalize("NFD", texto)
    return "".join(c for c in texto if unicodedata.category(c) != "Mn")


def abrir_matriz(raw: bytes, password: str) -> BytesIO:
    bio = BytesIO(raw)
    office = msoffcrypto.OfficeFile(bio)
    if office.is_encrypted():
        dec = BytesIO()
        bio.seek(0)
        office.load_key(password=password)
        office.decrypt(dec)
        dec.seek(0)
        return dec
    return BytesIO(raw)


def cuenta_numericos(valores: list) -> int:
    n = 0
    for v in valores:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        try:
            float(v)
            n += 1
        except (TypeError, ValueError):
            continue
    return n


def indices_calc(ws, fila_hdr: int):
    col_a = col_g = col_l = None
    max_c = ws.max_column or 0
    for c in range(1, max_c + 1):
        titulo = ws.cell(fila_hdr, c).value
        if titulo is None:
            continue
        n = normalizar(str(titulo))
        if n == "apropiacion":
            col_a = c
        elif n == "giros":
            col_g = c
        elif "liberacion" in n or "fenecimiento" in n:
            col_l = c
    if col_a and col_g and col_l:
        return col_a, col_g, col_l
    return None


def calcular_saldo(ws, fila_hdr: int) -> list:
    cols = indices_calc(ws, fila_hdr)
    if not cols:
        return []
    col_a, col_g, col_l = cols
    max_r = ws.max_row or fila_hdr
    filas = []
    for r in range(fila_hdr + 1, max_r + 1):
        try:
            a = float(ws.cell(r, col_a).value or 0)
            g = float(ws.cell(r, col_g).value or 0)
            lib = float(ws.cell(r, col_l).value or 0)
            filas.append(a - g - lib)
        except (TypeError, ValueError):
            filas.append(None)
    return filas


def columna_saldo_final(libro: BytesIO) -> tuple[list, dict[str, float]]:
    tiempos: dict[str, float] = {}
    libro.seek(0)
    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(libro, read_only=True, data_only=True)
    tiempos["load_workbook"] = time.perf_counter() - t0
    ws = wb[SHEET]
    meta = {"max_row": ws.max_row, "max_column": ws.max_column}

    t0 = time.perf_counter()
    col_idx = None
    for c in range(1, (ws.max_column or 0) + 1):
        titulo = ws.cell(FILA_HDR, c).value
        if titulo is not None and normalizar(str(titulo)) == "saldo final":
            col_idx = c
            break
    tiempos["buscar_columna"] = time.perf_counter() - t0
    meta["col_saldo_final"] = col_idx

    valores: list = []
    if col_idx is not None:
        t0 = time.perf_counter()
        max_r = ws.max_row or FILA_HDR
        valores = [
            row[0]
            for row in ws.iter_rows(
                min_row=FILA_HDR + 1,
                max_row=max_r,
                min_col=col_idx,
                max_col=col_idx,
                values_only=True,
            )
        ]
        tiempos["iter_rows_saldo"] = time.perf_counter() - t0
        meta["filas_iter"] = len(valores)
        meta["numericos_cache"] = cuenta_numericos(valores)

    umbral = max(3, int(len(valores) * 0.01)) if valores else 3
    if cuenta_numericos(valores) < umbral:
        t0 = time.perf_counter()
        calculados = calcular_saldo(ws, FILA_HDR)
        tiempos["fallback_calcular"] = time.perf_counter() - t0
        meta["numericos_calc"] = cuenta_numericos(calculados)
        meta["fallback_usado"] = cuenta_numericos(calculados) > cuenta_numericos(valores)
        if meta["fallback_usado"]:
            valores = calculados
    else:
        meta["fallback_usado"] = False

    wb.close()
    return valores, {**meta, **tiempos}


def fmt(sec: float) -> str:
    return f"{sec:.2f} s"


def main() -> None:
    if not MATRIZ.is_file():
        print("No existe:", MATRIZ)
        sys.exit(1)

    raw = MATRIZ.read_bytes()
    print(f"Matriz: {MATRIZ.name} ({len(raw):,} bytes)\n")

    t0 = time.perf_counter()
    try:
        libro = abrir_matriz(raw, PWD)
    except (ms_exceptions.InvalidKeyError, ms_exceptions.DecryptionError):
        print("Contraseña incorrecta (probar otra que 1100)")
        sys.exit(1)
    print(f"1. Descifrar/abrir: {fmt(time.perf_counter() - t0)}")

    t0 = time.perf_counter()
    df = pd.read_excel(libro, sheet_name=SHEET, engine="openpyxl", header=HEADER_PD)
    print(f"2. pd.read_excel: {fmt(time.perf_counter() - t0)}  -> {len(df)} filas x {len(df.columns)} cols")

    libro.seek(0)
    t0 = time.perf_counter()
    vals, meta = columna_saldo_final(libro)
    print(f"3. Saldo Final (fase UI): {fmt(time.perf_counter() - t0)}")
    print(f"   max_row={meta['max_row']}, max_column={meta['max_column']}, col V={meta.get('col_saldo_final')}")
    for k in (
        "load_workbook",
        "buscar_columna",
        "iter_rows_saldo",
        "fallback_calcular",
    ):
        if k in meta:
            print(f"   · {k}: {fmt(meta[k])}")
    print(
        f"   cache numéricos={meta.get('numericos_cache')}, "
        f"calc numéricos={meta.get('numericos_calc', '—')}, "
        f"fallback={meta.get('fallback_usado')}"
    )

    t_all = time.perf_counter()
    libro2 = abrir_matriz(raw, PWD)
    pd.read_excel(libro2, sheet_name=SHEET, engine="openpyxl", header=HEADER_PD)
    libro2.seek(0)
    columna_saldo_final(libro2)
    print(f"\nTOTAL Matriz (abrir+read_excel+Saldo Final): {fmt(time.perf_counter() - t_all)}")

    if CONTRATOS.is_file():
        from datetime import datetime

        from cxp_cruce import procesar_localidad_cxp

        print(f"\nContratos: {CONTRATOS.name} ({CONTRATOS.stat().st_size:,} bytes)")
        fases: list[str] = []

        def avance(msg: str) -> None:
            fases.append(msg)
            print(f"   CXP [{len(fases)}] {msg}")

        t0 = time.perf_counter()
        libro_m = abrir_matriz(raw, PWD)
        df_m = pd.read_excel(
            libro_m, sheet_name=SHEET, engine="openpyxl", header=HEADER_PD
        )
        libro_m.seek(0)
        v, _ = columna_saldo_final(libro_m)
        if "Saldo Final" in df_m.columns:
            n = len(df_m)
            df_m["Saldo Final"] = pd.to_numeric(pd.Series(v[:n]), errors="coerce")

        procesar_localidad_cxp(
            CONTRATOS.read_bytes(),
            df_m,
            "Kennedy",
            datetime(2026, 5, 29),
            CONTRATOS.name,
            MATRIZ.name,
            avance=avance,
        )
        print(f"4. procesar_localidad_cxp: {fmt(time.perf_counter() - t0)}")
        print(f"   Fases: {len(fases)}")
    else:
        print("\nContratos no encontrado en Downloads")


if __name__ == "__main__":
    main()
