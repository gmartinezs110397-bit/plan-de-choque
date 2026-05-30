"""Cruce Matriz → Contratos (CXP / Saldo Final) y actualización de columna del mes.

Incluye desempate manual: aplicar_desempate_en_contratos (aprox. línea 304).
"""

from __future__ import annotations

import calendar
import os
import re
import tempfile
import unicodedata
import zipfile
from xml.etree import ElementTree as ET
from copy import copy
from datetime import date, datetime
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from constantes import COL_DESEMPATE_MANUAL, HOJAS_CRUCE_CXP

MESES_ES = (
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
)

HEADER_MATRIZ = 6
HEADER_CONTRATOS = 2
FILA_CONTEO_CONTRATOS = 1
FILA_SUMA_CONTRATOS = 2

METODOS_LABEL = {
    "k4_exacto": "Match exacto (4 campos)",
    "match_saldo_contrato": "Fallback por Saldo Final",
    "todos_cero_matriz": "Fallback: todos cero en matriz",
    "k3_unico": "Match por contrato (sin apropiación)",
    "verificar": "Sin resolver",
    "sin_matriz": "Sin fila en matriz",
    "sin_saldo_matriz": "Saldo vacío en matriz",
    "desempate_manual": "Desempate manual",
}

CODIGOS_SIN_RESOLVER = ("verificar", "sin_matriz")
LABELS_SIN_RESOLVER = {METODOS_LABEL[c] for c in CODIGOS_SIN_RESOLVER}
LABEL_A_CODIGO = {v: k for k, v in METODOS_LABEL.items()}

TIPO_FILA_CONTRATOS = "Contratos"
TIPO_FILA_CANDIDATO_MATRIZ = "Candidato Matriz"
COL_OPCION_MATRIZ = "Opción Matriz"


def _normalizar(texto: str) -> str:
    texto = str(texto).lower()
    texto = unicodedata.normalize("NFD", texto)
    return "".join(c for c in texto if unicodedata.category(c) != "Mn")


def _norm_num(valor) -> str:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    if isinstance(valor, float) and valor == int(valor):
        return str(int(valor))
    return str(valor).strip()


def _saldo_numerico_matriz(valor) -> float | None:
    """None si Saldo Final en Matriz está vacío; 0.0 solo si es cero explícito."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    try:
        n = float(valor)
    except (TypeError, ValueError):
        return None
    if pd.isna(n):
        return None
    return n


def _suma_saldos_grupo_matriz(serie: pd.Series) -> float | None:
    """Suma de saldos del grupo; None si todas las filas vienen sin saldo en Matriz."""
    if serie.isna().all():
        return None
    total = serie.sum(skipna=True)
    if pd.isna(total):
        return None
    return float(total)


def clave_cuatro(nombre, contrato, anio, apropiacion) -> str:
    n = re.sub(r"\s+", "", str(nombre).upper())
    return n + _norm_num(contrato) + _norm_num(anio) + _norm_num(apropiacion)


def clave_tres(nombre, contrato, anio) -> str:
    n = re.sub(r"\s+", "", str(nombre).upper())
    return n + _norm_num(contrato) + _norm_num(anio)


def _palabras_localidad(localidad: str) -> list[str]:
    palabras = re.findall(r"[a-z0-9]+", _normalizar(localidad))
    ignorar = {"de", "la", "los", "las", "el", "del", "y"}
    significativas = [p for p in palabras if p not in ignorar and len(p) >= 2]
    return significativas if significativas else palabras


def fila_coincide_localidad(valor_celda, localidad: str) -> bool:
    texto = _normalizar(str(valor_celda))
    return any(p in texto for p in _palabras_localidad(localidad))


def _columna(df: pd.DataFrame, *candidatos: str) -> str | None:
    mapa = {_normalizar(c): c for c in df.columns}
    for cand in candidatos:
        key = _normalizar(cand)
        if key in mapa:
            return mapa[key]
    return None


def preparar_indice_matriz(df_matriz: pd.DataFrame, localidad: str) -> tuple[pd.DataFrame, dict, dict]:
    col_loc = _columna(df_matriz, "Localidad", "LOCALIDAD") or df_matriz.columns[0]
    df_loc = df_matriz[df_matriz[col_loc].apply(lambda v: fila_coincide_localidad(v, localidad))].copy()

    col_nombre = _columna(df_matriz, "NOMBRE CONTRATISTA")
    col_cto = _columna(df_matriz, "Número Contrato", "Numero Contrato")
    col_anio = _columna(df_matriz, "Año Suscripción", "Ano Suscripcion")
    col_aprop = _columna(df_matriz, "Apropiación", "Apropiacion")
    col_saldo = _columna(df_matriz, "Saldo Final")

    if not all([col_nombre, col_cto, col_anio, col_aprop, col_saldo]):
        raise ValueError("La Matriz no tiene las columnas esperadas para el cruce.")

    df_loc["_k4"] = df_loc.apply(
        lambda r: clave_cuatro(r[col_nombre], r[col_cto], r[col_anio], r[col_aprop]),
        axis=1,
    )
    df_loc["_k3"] = df_loc.apply(
        lambda r: clave_tres(r[col_nombre], r[col_cto], r[col_anio]),
        axis=1,
    )
    df_loc["_saldo"] = pd.to_numeric(df_loc[col_saldo], errors="coerce")

    mapa_k4 = df_loc.drop_duplicates("_k4", keep="first").set_index("_k4")["_saldo"].to_dict()
    grupos_k3 = {k: g for k, g in df_loc.groupby("_k3", sort=False)}
    return df_loc, mapa_k4, grupos_k3


def buscar_saldo_matriz(
    mapa_k4: dict,
    grupos_k3: dict,
    nombre,
    contrato,
    anio,
    apropiacion,
    saldo_final_contrato,
) -> tuple[float | None, str, str]:
    """
    Devuelve (saldo_matriz, metodo_codigo, detalle_texto).
    """
    k4 = clave_cuatro(nombre, contrato, anio, apropiacion)
    if k4 in mapa_k4:
        saldo_k4 = _saldo_numerico_matriz(mapa_k4[k4])
        if saldo_k4 is not None:
            return saldo_k4, "k4_exacto", ""
        return (
            None,
            "sin_saldo_matriz",
            "Hay fila en la Matriz pero Saldo Final vacío.",
        )

    k3 = clave_tres(nombre, contrato, anio)
    if k3 not in grupos_k3:
        return None, "sin_matriz", "No hay fila en la Matriz para este contrato."

    g = grupos_k3[k3]
    if len(g) == 1:
        saldo_unico = _saldo_numerico_matriz(g.iloc[0]["_saldo"])
        if saldo_unico is None:
            return (
                None,
                "sin_saldo_matriz",
                "Hay fila en la Matriz pero Saldo Final vacío.",
            )
        return saldo_unico, "k3_unico", "Apropiación en Matriz distinta; una sola fila."

    sf = g["_saldo"]
    if sf.notna().sum() == 0:
        return (
            None,
            "sin_saldo_matriz",
            "Hay fila en la Matriz pero Saldo Final vacío.",
        )
    if (sf.fillna(0) > 0).sum() == 0:
        return 0.0, "todos_cero_matriz", "Varias filas en Matriz, todas con saldo 0."

    sk = float(pd.to_numeric(saldo_final_contrato, errors="coerce") or 0)
    coinciden = g[sf == sk]
    if len(coinciden) == 1:
        saldo_match = _saldo_numerico_matriz(coinciden.iloc[0]["_saldo"])
        if saldo_match is None:
            return (
                None,
                "sin_saldo_matriz",
                "Hay fila en la Matriz pero Saldo Final vacío.",
            )
        detalle = (
            f"Apropiación en Contratos ({_norm_num(apropiacion)}) distinta a la Matriz; "
            f"se usó SALDO FINAL del contrato ({sk:,.0f})."
        )
        return saldo_match, "match_saldo_contrato", detalle
    if len(coinciden) > 1:
        saldo_match = _saldo_numerico_matriz(coinciden.iloc[0]["_saldo"])
        if saldo_match is None:
            return (
                None,
                "sin_saldo_matriz",
                "Hay fila en la Matriz pero Saldo Final vacío.",
            )
        return saldo_match, "match_saldo_contrato", "Varias filas con el mismo saldo."

    return (
        None,
        "verificar",
        "Hay saldos > 0 en Matriz pero ninguno coincide con SALDO FINAL del contrato.",
    )


def clave_fila_contrato(nombre, contrato, anio, apropiacion) -> str:
    return clave_cuatro(nombre, contrato, anio, apropiacion)


def clave_desde_detalle(fila: dict) -> str:
    return clave_fila_contrato(
        fila["NOMBRE CONTRATISTA"],
        fila["No. de Cto"],
        fila["AÑO SUSCRIPCIÓN"],
        fila["APROPIACION DISPONIBLE"],
    )


def _es_fila_contratos_revision(fila: dict) -> bool:
    return fila.get("Tipo fila", TIPO_FILA_CONTRATOS) == TIPO_FILA_CONTRATOS


def _listar_opciones_matriz(grupo: pd.DataFrame) -> list[dict]:
    """Filas en Matriz con mismo nombre + contrato + año (sin apropiación)."""
    col_aprop = _columna(grupo, "APROPIACION DISPONIBLE", "Apropiación", "Apropiacion")
    if not col_aprop:
        return []
    opciones = []
    for n, (_, row) in enumerate(grupo.iterrows(), 1):
        opciones.append({
            "opcion": n,
            "apropiacion": row[col_aprop],
            "saldo": float(row["_saldo"]),
        })
    return opciones


def construir_dataframe_revision(
    detalle_filas: list[dict],
    localidad: str | None = None,
) -> pd.DataFrame:
    """
    Lista a revisar: fila del Contrato + candidatos en Matriz (misma búsqueda k3).
    «Desempate manual» solo en la fila Contratos (saldo elegido para el mes).
    """
    pendientes = [
        f
        for f in detalle_filas
        if f.get("Método") in LABELS_SIN_RESOLVER
        and _es_fila_contratos_revision(f)
        and (localidad is None or f.get("Localidad") == localidad)
    ]
    if not pendientes:
        return pd.DataFrame()

    filas_export: list[dict] = []
    for f in pendientes:
        titulo_mes_col = next(
            (c for c in f if str(c).startswith("Saldo Matriz (")),
            f"Saldo Matriz",
        )
        fila_contratos = {
            "Tipo fila": TIPO_FILA_CONTRATOS,
            COL_OPCION_MATRIZ: "—",
            "Localidad": f.get("Localidad"),
            "NOMBRE CONTRATISTA": f.get("NOMBRE CONTRATISTA"),
            "No. de Cto": f.get("No. de Cto"),
            "AÑO SUSCRIPCIÓN": f.get("AÑO SUSCRIPCIÓN"),
            "APROPIACION DISPONIBLE": f.get("APROPIACION DISPONIBLE"),
            "SALDO FINAL (Contratos)": f.get("SALDO FINAL (Contratos)"),
            "Método": f.get("Método"),
            "Detalle": f.get("Detalle"),
            titulo_mes_col: f.get(titulo_mes_col),
            COL_DESEMPATE_MANUAL: "",
        }
        filas_export.append(fila_contratos)

        candidatos = f.get("candidatos_matriz") or []
        for cand in candidatos:
            filas_export.append({
                "Tipo fila": TIPO_FILA_CANDIDATO_MATRIZ,
                COL_OPCION_MATRIZ: cand.get("opcion"),
                "Localidad": f.get("Localidad"),
                "NOMBRE CONTRATISTA": f.get("NOMBRE CONTRATISTA"),
                "No. de Cto": f.get("No. de Cto"),
                "AÑO SUSCRIPCIÓN": f.get("AÑO SUSCRIPCIÓN"),
                "APROPIACION DISPONIBLE": cand.get("apropiacion"),
                "SALDO FINAL (Contratos)": f.get("SALDO FINAL (Contratos)"),
                "Método": "—",
                "Detalle": "Opción en Matriz (referencia; no complete Desempate aquí)",
                titulo_mes_col: cand.get("saldo"),
                COL_DESEMPATE_MANUAL: "",
            })

    preferidas = [
        "Tipo fila",
        COL_OPCION_MATRIZ,
        "Localidad",
        "NOMBRE CONTRATISTA",
        "No. de Cto",
        "AÑO SUSCRIPCIÓN",
        "APROPIACION DISPONIBLE",
        "SALDO FINAL (Contratos)",
        "Método",
        "Detalle",
    ]
    df = pd.DataFrame(filas_export)
    cols_matriz = [c for c in df.columns if str(c).startswith("Saldo Matriz")]
    cols = [c for c in preferidas if c in df.columns] + cols_matriz + [COL_DESEMPATE_MANUAL]
    return df[[c for c in cols if c in df.columns]]


def parsear_mapa_desempate(df_revision: pd.DataFrame) -> dict[str, float]:
    """Lee la columna Desempate manual (solo filas con valor numérico)."""
    col_d = _columna(df_revision, COL_DESEMPATE_MANUAL, "Desempate Manual")
    if not col_d:
        raise ValueError(f"Falta la columna «{COL_DESEMPATE_MANUAL}» en el archivo subido.")

    col_nombre = _columna(df_revision, "NOMBRE CONTRATISTA")
    col_cto = _columna(df_revision, "No. de Cto", "Número Contrato")
    col_anio = _columna(df_revision, "AÑO SUSCRIPCIÓN", "Año Suscripción")
    col_aprop = _columna(df_revision, "APROPIACION DISPONIBLE", "Apropiación")
    if not all([col_nombre, col_cto, col_anio, col_aprop]):
        raise ValueError(
            "El archivo debe conservar las columnas del listado: "
            "NOMBRE CONTRATISTA, No. de Cto, AÑO SUSCRIPCIÓN, APROPIACION DISPONIBLE."
        )

    col_tipo = _columna(df_revision, "Tipo fila", "Tipo Fila")
    mapa: dict[str, float] = {}
    for idx, row in df_revision.iterrows():
        if col_tipo and str(row.get(col_tipo, "")).strip() == TIPO_FILA_CANDIDATO_MATRIZ:
            continue
        raw = row[col_d]
        if raw is None or (isinstance(raw, float) and pd.isna(raw)) or str(raw).strip() == "":
            continue
        saldo = pd.to_numeric(raw, errors="coerce")
        if pd.isna(saldo):
            raise ValueError(
                f"Fila {int(idx) + 2}: «{COL_DESEMPATE_MANUAL}» debe ser un número "
                f"(valor recibido: {raw!r})."
            )
        k = clave_fila_contrato(
            row[col_nombre], row[col_cto], row[col_anio], row[col_aprop]
        )
        mapa[k] = float(saldo)
    return mapa


def claves_pendientes_localidad(detalle_filas: list[dict], localidad: str) -> set[str]:
    return {
        clave_desde_detalle(f)
        for f in detalle_filas
        if f.get("Localidad") == localidad
        and f.get("Método") in LABELS_SIN_RESOLVER
        and _es_fila_contratos_revision(f)
    }


def validar_desempate_completo(
    pendientes: set[str],
    mapa: dict[str, float],
    detalle_localidad: list[dict] | None = None,
) -> list[str]:
    por_clave = {
        clave_desde_detalle(f): f
        for f in (detalle_localidad or [])
        if f.get("Método") in LABELS_SIN_RESOLVER and _es_fila_contratos_revision(f)
    }
    errores = []
    for k in sorted(pendientes):
        if k not in mapa:
            fila = por_clave.get(k, {})
            nombre = fila.get("NOMBRE CONTRATISTA", "Contrato")
            cto = fila.get("No. de Cto", "")
            errores.append(
                f"**{nombre}** (contrato {cto}): complete «{COL_DESEMPATE_MANUAL}»."
            )
    return errores


def _etiqueta_saldo_matriz(titulo_mes: str) -> str:
    return f"Saldo Matriz ({titulo_mes})"


def recalcular_estadisticas_localidad(
    detalle_filas: list[dict],
    titulo_mes: str,
) -> dict[str, Any]:
    col_saldo = _etiqueta_saldo_matriz(titulo_mes)
    conteo: dict[str, int] = {k: 0 for k in METODOS_LABEL}
    saldos_mes: list[float] = []

    for fila in detalle_filas:
        codigo = LABEL_A_CODIGO.get(fila.get("Método", ""), "verificar")
        conteo[codigo] = conteo.get(codigo, 0) + 1
        if codigo not in CODIGOS_SIN_RESOLVER:
            saldo = fila.get(col_saldo)
            if saldo is not None and not (isinstance(saldo, float) and pd.isna(saldo)):
                saldos_mes.append(float(saldo))

    total = sum(conteo.values())
    sin_resolver = sum(conteo.get(c, 0) for c in CODIGOS_SIN_RESOLVER)
    resumen_metodos = [
        {"Método": METODOS_LABEL[codigo], "Contratos": cantidad}
        for codigo, cantidad in sorted(conteo.items(), key=lambda x: -x[1])
        if cantidad > 0
    ]
    return {
        "total_contratos": total,
        "contratos_ok": total - sin_resolver,
        "sin_resolver": sin_resolver,
        "cxp_total": sum(saldos_mes),
        "conteo": conteo,
        "resumen_metodos": resumen_metodos,
    }


def aplicar_desempate_en_contratos(
    contratos_bytes: bytes,
    fecha_analisis: datetime,
    mapa_desempate: dict[str, float],
    detalle_filas: list[dict],
    localidad: str,
) -> tuple[bytes, list[dict]]:
    """
    Escribe saldos manuales en la columna de corte y actualiza el detalle.
    Solo aplica a contratos que siguen «sin resolver».
    """
    titulo_mes = titulo_saldo_corte(fecha_analisis)
    col_saldo_label = _etiqueta_saldo_matriz(titulo_mes)
    pendientes = claves_pendientes_localidad(detalle_filas, localidad)

    libro = pd.ExcelFile(BytesIO(contratos_bytes))
    nombre_hoja = resolver_hoja_cruce_cxp(list(libro.sheet_names))
    df_c = pd.read_excel(
        BytesIO(contratos_bytes), sheet_name=nombre_hoja, header=HEADER_CONTRATOS
    )
    col_nombre = _columna(df_c, "NOMBRE CONTRATISTA")
    col_cto = _columna(df_c, "No. de Cto", "Número Contrato")
    col_anio = _columna(df_c, "AÑO SUSCRIPCIÓN", "ANO SUSCRIPCION", "Año Suscripción")
    col_aprop = _columna(df_c, "APROPIACION DISPONIBLE", "Apropiación", "Apropiacion")
    if not all([col_nombre, col_cto, col_anio, col_aprop]):
        raise ValueError(
            f"Contratos: faltan columnas para aplicar el desempate en «{nombre_hoja}»."
        )

    col_mes_existente = _indice_columna_corte(list(df_c.columns), fecha_analisis)
    col_mes = col_mes_existente or titulo_mes
    crear_columna = col_mes_existente is None

    detalle_nuevo = []
    for fila in detalle_filas:
        if fila.get("Localidad") != localidad:
            detalle_nuevo.append(dict(fila))
            continue
        copia = dict(fila)
        if copia.get("Método") in LABELS_SIN_RESOLVER:
            k = clave_desde_detalle(copia)
            if k in mapa_desempate:
                saldo = mapa_desempate[k]
                copia[col_saldo_label] = saldo
                copia["Método"] = METODOS_LABEL["desempate_manual"]
                copia["Detalle"] = (
                    f"Saldo asignado en «{COL_DESEMPATE_MANUAL}» ({saldo:,.0f})."
                )
        detalle_nuevo.append(copia)

    valores_excel: dict[int, float | None] = {}
    for i, (_, row) in enumerate(df_c.iterrows()):
        nombre = row[col_nombre]
        if pd.isna(nombre) or not str(nombre).strip():
            continue
        fila_excel = _fila_inicio_datos_contratos() + i
        k = clave_fila_contrato(
            nombre, row[col_cto], row[col_anio], row[col_aprop]
        )
        if k in pendientes and k in mapa_desempate:
            valores_excel[fila_excel] = mapa_desempate[k]

    bytes_nuevos, _ = exportar_contratos_preservando_formato(
        contratos_bytes,
        fecha_analisis,
        valores_excel,
        crear_columna=crear_columna,
        titulo_columna=titulo_mes,
    )
    return bytes_nuevos, detalle_nuevo


def _fecha_datetime(fecha: datetime | date) -> datetime:
    if isinstance(fecha, datetime):
        return fecha
    return datetime(fecha.year, fecha.month, fecha.day)


def dia_fin_mes_corte(fecha: datetime | date) -> int:
    """
    Último día del mes de la fecha de ejecución del consolidado.
    Ej.: si se ejecuta el 29-mayo-2026 → 31 (no usa el día 29 del calendario).
    """
    f = _fecha_datetime(fecha)
    return calendar.monthrange(f.year, f.month)[1]


def mes_nombre_corte(fecha: datetime | date) -> str:
    return MESES_ES[_fecha_datetime(fecha).month - 1]


def titulo_saldo_corte(fecha: datetime | date) -> str:
    """
    Título de columna según fecha de ejecución.
    Ej.: Saldo a 31 de mayo, Saldo a 30 de abril (último día del mes).
    """
    dia = dia_fin_mes_corte(fecha)
    mes = mes_nombre_corte(fecha)
    return f"Saldo a {dia} de {mes}"


def titulo_columna_mes(fecha: datetime | date) -> str:
    """Alias del título de corte (compatibilidad)."""
    return titulo_saldo_corte(fecha)


def resolver_hoja_cruce_cxp(nombres_hojas: list[str]) -> str:
    """Primera hoja de cruce CXP que exista en el libro (Cps / Caja por depurar)."""
    for candidato in HOJAS_CRUCE_CXP:
        if candidato in nombres_hojas:
            return candidato
    encontradas = ", ".join(nombres_hojas) if nombres_hojas else "(ninguna)"
    buscadas = ", ".join(f"«{h}»" for h in HOJAS_CRUCE_CXP)
    raise ValueError(
        f"Contratos: falta la hoja de cruce. Se buscó {buscadas}. "
        f"Hojas en el archivo: {encontradas}."
    )


def _fila_encabezado_contratos() -> int:
    """Fila 1-based del encabezado en «Cps por depurar» (fila 3 en Excel)."""
    return HEADER_CONTRATOS + 1


def _fila_inicio_datos_contratos() -> int:
    """Primera fila de datos de contratos (fila 4 en Excel)."""
    return HEADER_CONTRATOS + 2


def _copiar_estilo_celda(origen, destino, *, copiar_relleno: bool = True) -> None:
    if origen.has_style:
        destino.font = copy(origen.font)
        if copiar_relleno:
            destino.fill = copy(origen.fill)
        destino.border = copy(origen.border)
        destino.alignment = copy(origen.alignment)
        destino.number_format = origen.number_format


def _copiar_estilo_celda_sin_relleno(origen, destino) -> None:
    """Fuente, bordes, alineación y número — sin color de fondo (evita arrastrar verde/amarillo)."""
    _copiar_estilo_celda(origen, destino, copiar_relleno=False)


_ALIGNMENT_CENTRO_TOTAL = Alignment(horizontal="center", vertical="center")


def _centrar_celdas_total(*celdas) -> None:
    """Centra totales en filas de resumen (no usar en Estrategias)."""
    for celda in celdas:
        celda.alignment = _ALIGNMENT_CENTRO_TOTAL


def _celda_para_escribir(ws, fila: int, col: int):
    """Devuelve la celda superior-izquierda si (fila, col) está dentro de un rango combinado."""
    from openpyxl.cell.cell import MergedCell

    celda = ws.cell(fila, col)
    if not isinstance(celda, MergedCell):
        return celda
    coord = celda.coordinate
    for rango in ws.merged_cells.ranges:
        if coord in rango:
            return ws.cell(rango.min_row, rango.min_col)
    return celda


def _texto_ancho_celda(celda) -> str:
    """Texto visible aproximado para calcular auto-ancho (no usa la cadena de la fórmula)."""
    if celda.data_type == "f":
        return ""
    valor = celda.value
    if valor is None:
        return ""
    if isinstance(valor, str) and valor.strip().startswith("="):
        return ""
    if isinstance(valor, float) and not isinstance(valor, bool):
        if abs(valor - round(valor)) < 1e-9:
            valor = int(round(valor))
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        nf = str(celda.number_format or "")
        texto = f"{valor:,}"
        if "." in nf and "," not in nf.replace("#", "").replace("0", ""):
            pass
        if "." in nf or "0.000" in nf.lower():
            texto = texto.replace(",", ".")
        if "$" in nf:
            texto = f"$ {texto}"
        return texto
    return str(valor).strip()


def _autoajustar_ancho_columna(ws, col: int) -> None:
    """
    Ajusta el ancho de la columna al contenido (equivalente a doble clic en el borde).
    Revisa filas 1-2, encabezado y filas de contratos.
    """
    col_nombre = _indice_columna_en_hoja(ws, "NOMBRE CONTRATISTA")
    fila_ini = _fila_inicio_datos_contratos()
    max_len = 0

    for fila in range(1, ws.max_row + 1):
        if fila >= fila_ini and col_nombre and not _fila_tiene_contratista(ws, fila, col_nombre):
            continue
        celda = ws.cell(fila, col)
        texto = _texto_ancho_celda(celda)
        if texto:
            max_len = max(max_len, len(texto))

    # Conversión aproximada a unidades de ancho de Excel (+ margen como AutoFit)
    ancho = min(max(max_len * 1.15 + 3, 11), 60)
    ws.column_dimensions[get_column_letter(col)].width = ancho


def _ajustar_ancho_columna_corte(
    ws,
    col: int,
    col_referencia: int,
    titulo: str = "",
) -> None:
    """
    Auto-ancho por valores visibles y título del mes.
    No copia el ancho de SALDO FINAL (suele estar inflado por fórmulas largas).
    """
    _autoajustar_ancho_columna(ws, col)
    letra = get_column_letter(col)
    if titulo:
        min_titulo = min(max(len(titulo) * 1.15 + 3, 11), 60)
        ws.column_dimensions[letra].width = max(
            ws.column_dimensions[letra].width or 0,
            min_titulo,
        )


def _titulos_columnas_hoja(ws) -> list[str]:
    fila = _fila_encabezado_contratos()
    titulos = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(fila, col).value
        if val is not None and str(val).strip():
            titulos.append(str(val).strip())
        else:
            titulos.append("")
    return titulos


def _filas_candidatas_encabezado() -> tuple[int, ...]:
    """Fila 3 (plantilla estándar) y fila 1 (algunos archivos editados a mano)."""
    return (_fila_encabezado_contratos(), 1)


def _fila_encabezado_hoja_datos(ws) -> int:
    """Primera fila que contiene el encabezado NOMBRE CONTRATISTA."""
    for fila in _filas_candidatas_encabezado():
        for col in range(1, ws.max_column + 1):
            val = ws.cell(fila, col).value
            if val is not None and _normalizar(str(val)) == "nombre contratista":
                return fila
    return _fila_encabezado_contratos()


def _fila_inicio_datos_hoja(ws) -> int:
    """Primera fila de contratista tras el encabezado (salta filas vacías o «NO TIENE»)."""
    fila_hdr = _fila_encabezado_hoja_datos(ws)
    col_nombre = _indice_columna_en_fila(ws, fila_hdr, "NOMBRE CONTRATISTA")
    if not col_nombre:
        return _fila_inicio_datos_contratos()
    for fila in range(fila_hdr + 1, ws.max_row + 1):
        val = ws.cell(fila, col_nombre).value
        if val is None or not str(val).strip():
            continue
        norm = _normalizar(str(val))
        if norm in ("nombre contratista", "no tiene"):
            continue
        return fila
    return fila_hdr + 1


def _indice_columna_en_fila(ws, fila: int, *candidatos: str) -> int | None:
    for col in range(1, ws.max_column + 1):
        val = ws.cell(fila, col).value
        if val is None or not str(val).strip():
            continue
        norm = _normalizar(str(val))
        for cand in candidatos:
            if _normalizar(cand) == norm:
                return col
    return None


def _indice_columna_en_hoja(ws, *candidatos: str) -> int | None:
    for fila in _filas_candidatas_encabezado():
        col = _indice_columna_en_fila(ws, fila, *candidatos)
        if col is not None:
            return col
    return None


def _indice_columna_corte_en_hoja(ws, fecha: datetime | date) -> tuple[int | None, str]:
    titulos = _titulos_columnas_hoja(ws)
    nombre = _indice_columna_corte(titulos, fecha)
    if not nombre:
        return None, titulo_saldo_corte(fecha)
    fila = _fila_encabezado_contratos()
    for col in range(1, ws.max_column + 1):
        if _normalizar(str(ws.cell(fila, col).value)) == _normalizar(nombre):
            return col, nombre
    return None, nombre


def _columna_estilo_saldo_final(ws) -> int:
    """Columna SALDO FINAL (junto a MONTO LIBERACIONES…) para copiar formato."""
    col_sf = _indice_columna_en_hoja(ws, "SALDO FINAL", "Saldo Final")
    if col_sf:
        return col_sf
    col_monto = _indice_columna_en_hoja(
        ws,
        "MONTO LIBERACIONES O FENECICMIENTOS",
        "MONTO LIBERACIONES O FENECICMIENTOS",
        "MONTO LIBERACIONES",
    )
    if col_monto:
        return col_monto + 1
    # Evitar recursión con _ultima_columna_con_datos (columna base fija si no hay encabezado en fila 3).
    return 10


def _columna_tiene_contenido(ws, col: int) -> bool:
    """True si la columna tiene encabezado, resumen (1-2) o dato de contrato."""
    fila_hdr = _fila_encabezado_hoja_datos(ws)
    hdr = ws.cell(fila_hdr, col).value
    if hdr is not None and str(hdr).strip():
        return True
    for fila in (FILA_CONTEO_CONTRATOS, FILA_SUMA_CONTRATOS):
        if ws.cell(fila, col).value not in (None, ""):
            return True

    col_nombre = _indice_columna_en_hoja(ws, "NOMBRE CONTRATISTA")
    fila_ini = _fila_inicio_datos_hoja(ws)
    for fila in range(fila_ini, ws.max_row + 1):
        if col_nombre and not _fila_tiene_contratista(ws, fila, col_nombre):
            continue
        if ws.cell(fila, col).value not in (None, ""):
            return True
    return False


def _ultima_columna_con_datos(ws) -> int:
    """
    Última columna con información real (no el área vacía tras el filtro de Excel).
    Busca desde la columna 1 hasta SALDO FINAL + margen de columnas mensuales.
    """
    col_sf = _columna_estilo_saldo_final(ws)
    ultima = col_sf
    limite = min(ws.max_column, col_sf + 48)

    for col in range(1, limite + 1):
        if _columna_tiene_contenido(ws, col):
            ultima = max(ultima, col)

    return ultima


def _agregar_columna_corte_en_hoja(ws, titulo: str) -> int:
    """Inserta columna tras la última con datos; estilo copiado de SALDO FINAL."""
    fila_hdr = _fila_encabezado_contratos()
    col_estilo = _columna_estilo_saldo_final(ws)
    nueva_col = _ultima_columna_con_datos(ws) + 1

    _copiar_estilo_celda(ws.cell(fila_hdr, col_estilo), ws.cell(fila_hdr, nueva_col))
    ws.cell(fila_hdr, nueva_col, value=titulo)

    for fila in (FILA_CONTEO_CONTRATOS, FILA_SUMA_CONTRATOS):
        _copiar_estilo_celda(ws.cell(fila, col_estilo), ws.cell(fila, nueva_col))

    return nueva_col


def _fila_tiene_contratista(ws, fila: int, col_nombre: int) -> bool:
    val = ws.cell(fila, col_nombre).value
    if val is None or not str(val).strip():
        return False
    norm = _normalizar(str(val))
    return norm not in ("no tiene", "nombre contratista")


def _hoja_tiene_filas_contratista(ws) -> bool:
    """True si la pestaña tiene al menos un contratista real (no vacía ni «NO TIENE»)."""
    col_nombre = _indice_columna_en_hoja(ws, "NOMBRE CONTRATISTA")
    if not col_nombre:
        return False
    fila_ini = _fila_inicio_datos_hoja(ws)
    for fila in range(fila_ini, ws.max_row + 1):
        if _fila_tiene_contratista(ws, fila, col_nombre):
            return True
    return False


def _celda_tiene_formula(celda) -> bool:
    if celda.data_type == "f":
        return True
    val = celda.value
    return isinstance(val, str) and val.startswith("=")


def _actualizar_resumen_filas_1_2(ws, col_corte: int) -> None:
    """
    Fila 1: cantidad de contratos con saldo en la columna de corte.
    Fila 2: suma de esos saldos (si no hay fórmula ya definida).
    """
    col_nombre = _indice_columna_en_hoja(ws, "NOMBRE CONTRATISTA")
    fila_ini = _fila_inicio_datos_contratos()
    conteo = 0
    suma = 0.0
    for fila in range(fila_ini, ws.max_row + 1):
        if col_nombre and not _fila_tiene_contratista(ws, fila, col_nombre):
            continue
        raw = ws.cell(fila, col_corte).value
        if raw is None or raw == "":
            continue
        try:
            saldo = float(raw)
        except (TypeError, ValueError):
            continue
        if saldo != 0:
            conteo += 1
        suma += saldo

    col_estilo = _columna_estilo_saldo_final(ws)
    celda_conteo = ws.cell(FILA_CONTEO_CONTRATOS, col_corte)
    celda_suma = ws.cell(FILA_SUMA_CONTRATOS, col_corte)
    if not _celda_tiene_formula(celda_conteo):
        celda_conteo.value = conteo
        _copiar_estilo_celda(ws.cell(FILA_CONTEO_CONTRATOS, col_estilo), celda_conteo)
    if not _celda_tiene_formula(celda_suma):
        celda_suma.value = suma
        _copiar_estilo_celda(ws.cell(FILA_SUMA_CONTRATOS, col_estilo), celda_suma)
    _centrar_celdas_total(celda_conteo, celda_suma)


def quitar_autofiltros_xlsx(
    contenido: bytes,
    hojas: list[str] | None = None,
    *,
    finalizar: bool = True,
) -> bytes:
    """
    Quita AutoFilter de las hojas indicadas (o de todo el libro).
    Devuelve los bytes originales si no hay filtros o no se pudo abrir el archivo.
    """
    try:
        wb = load_workbook(BytesIO(contenido))
    except Exception:
        return contenido

    modifico = False
    objetivos = hojas if hojas is not None else list(wb.sheetnames)
    for nombre in objetivos:
        if nombre not in wb.sheetnames:
            continue
        ws = wb[nombre]
        af = ws.auto_filter
        if af is not None and getattr(af, "ref", None):
            ws.auto_filter.ref = None
            modifico = True

    if not modifico:
        wb.close()
        return contenido

    _preparar_workbook_antes_guardar(wb)
    out = BytesIO()
    wb.save(out)
    wb.close()
    resultado = out.getvalue()
    if finalizar:
        return _finalizar_xlsx_contratos(resultado)
    return resultado


def _preparar_workbook_antes_guardar(wb) -> None:
    """Evita que openpyxl deje enlaces externos rotos al guardar."""
    if hasattr(wb, "_external_links"):
        wb._external_links = []
    if getattr(wb, "calculation", None) is not None:
        wb.calculation.fullCalcOnLoad = None


def _filtrar_xml_relaciones_externas(datos: bytes) -> bytes:
    root = ET.fromstring(datos)
    for rel in list(root):
        if "externalLink" in (rel.get("Type") or ""):
            root.remove(rel)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _filtrar_content_types_externos(datos: bytes) -> bytes:
    root = ET.fromstring(datos)
    for child in list(root):
        if "/externalLinks/" in (child.get("PartName") or ""):
            root.remove(child)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _quitar_enlaces_externos_xlsx(contenido: bytes) -> bytes:
    """
    El Contratos original suele traer fórmulas ligadas a otros Excel.
    Tras guardar con openpyxl, Excel Mac pide reparar externalLink1.xml.
    """
    entrada = BytesIO(contenido)
    salida = BytesIO()
    try:
        with zipfile.ZipFile(entrada, "r") as zin, zipfile.ZipFile(
            salida, "w", zipfile.ZIP_DEFLATED
        ) as zout:
            for info in zin.infolist():
                nombre = info.filename.replace("\\", "/")
                if nombre.startswith("xl/externalLinks/"):
                    continue
                datos = zin.read(info.filename)
                if nombre == "xl/_rels/workbook.xml.rels":
                    datos = _filtrar_xml_relaciones_externas(datos)
                elif nombre == "[Content_Types].xml":
                    datos = _filtrar_content_types_externos(datos)
                zout.writestr(info, datos)
        return salida.getvalue()
    except Exception:
        return contenido


def _finalizar_xlsx_contratos(contenido: bytes) -> bytes:
    return compatibilizar_xlsx_excel_mac(_quitar_enlaces_externos_xlsx(contenido))


def compatibilizar_xlsx_excel_mac(contenido: bytes) -> bytes:
    """
    openpyxl genera OOXML que Excel en Mac marca como dañado (inlineStr, calc, etc.).
    xlsx-fixer reescribe el archivo para Excel Mac/Windows.
    """
    try:
        from xlsx_fixer import fix
    except ImportError:
        return contenido

    ruta: str | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            tf.write(contenido)
            ruta = tf.name
        fix(ruta)
        with open(ruta, "rb") as archivo:
            return archivo.read()
    except Exception:
        return contenido
    finally:
        if ruta:
            try:
                os.unlink(ruta)
            except OSError:
                pass


def exportar_contratos_preservando_formato(
    contratos_bytes: bytes,
    fecha_analisis: datetime,
    valores_por_fila: dict[int, float | None],
    crear_columna: bool,
    titulo_columna: str,
    mapa_k3_suspendidos: dict | None = None,
    mapa_k4_liquidados: dict[str, float] | None = None,
) -> tuple[bytes, list[str], list[str]]:
    """
    Guarda el libro original intacto (filas 1-2, formatos, otras hojas).
    Actualiza Cps/Caja por depurar y hojas de seguimiento mensual si existen.
    valores_por_fila: fila Excel 1-based -> saldo.
    Devuelve (bytes, advertencias_hojas, observaciones).
    """
    from hoja_proximos_a_perder import (
        actualizar_hoja_proximos_a_perder,
        resolver_hoja_proximos_a_perder,
    )
    from hoja_suspendidos import (
        actualizar_hoja_suspendidos,
        resolver_hoja_suspendidos,
    )
    from hoja_estrategias import actualizar_estrategias_en_libro
    from hoja_liquidados_con_saldo import (
        actualizar_hoja_liquidados_con_saldo,
        resolver_hoja_liquidados_con_saldo,
    )
    from hoja_tramites_sectores import (
        actualizar_hoja_tramites_sectores,
        resolver_hoja_tramites_sectores,
    )

    advertencias: list[str] = []
    observaciones: list[str] = []
    wb = load_workbook(BytesIO(contratos_bytes))
    nombres_hojas = list(wb.sheetnames)
    nombre_hoja = resolver_hoja_cruce_cxp(nombres_hojas)
    ws = wb[nombre_hoja]

    titulo_corte = titulo_columna or titulo_saldo_corte(fecha_analisis)
    col_corte, titulo_usado = _indice_columna_corte_en_hoja(ws, fecha_analisis)
    col_estilo = _columna_estilo_saldo_final(ws)
    if col_corte is None:
        col_corte = _agregar_columna_corte_en_hoja(ws, titulo_corte)
    else:
        if crear_columna and titulo_columna:
            ws.cell(_fila_encabezado_contratos(), col_corte, value=titulo_columna)

    col_nombre = _indice_columna_en_hoja(ws, "NOMBRE CONTRATISTA")

    for fila, valor in valores_por_fila.items():
        if col_nombre and not _fila_tiene_contratista(ws, fila, col_nombre):
            continue
        celda = ws.cell(fila, col_corte, value=valor)
        _copiar_estilo_celda(ws.cell(fila, col_estilo), celda)

    _actualizar_resumen_filas_1_2(ws, col_corte)
    _ajustar_ancho_columna_corte(ws, col_corte, col_estilo, titulo_usado or titulo_corte)

    if mapa_k3_suspendidos is not None:
        nombre_susp = resolver_hoja_suspendidos(nombres_hojas)
        if nombre_susp:
            try:
                ws_susp = wb[nombre_susp]
                if _hoja_tiene_filas_contratista(ws_susp):
                    advertencias.extend(
                        actualizar_hoja_suspendidos(
                            ws_susp,
                            mapa_k3_suspendidos,
                            fecha_analisis,
                        )
                    )
            except ValueError as e:
                observaciones.append(f"Suspendidos: {e}")
            except Exception as e:
                observaciones.append(
                    f"Suspendidos: error no previsto ({type(e).__name__}: {e})"
                )
        else:
            observaciones.append(
                "No se encontró pestaña Suspendidos en el archivo de Contratos."
            )

        nombre_prox = resolver_hoja_proximos_a_perder(nombres_hojas)
        if nombre_prox:
            try:
                ws_prox = wb[nombre_prox]
                if _hoja_tiene_filas_contratista(ws_prox):
                    advertencias.extend(
                        actualizar_hoja_proximos_a_perder(
                            ws_prox,
                            mapa_k3_suspendidos,
                            fecha_analisis,
                        )
                    )
            except ValueError as e:
                observaciones.append(f"Próximos a perder: {e}")
            except Exception as e:
                observaciones.append(
                    f"Próximos a perder: error no previsto ({type(e).__name__}: {e})"
                )
        else:
            observaciones.append(
                "No se encontró pestaña Próximos a perder en el archivo de Contratos."
            )

        nombre_tram = resolver_hoja_tramites_sectores(nombres_hojas)
        if nombre_tram:
            try:
                ws_tram = wb[nombre_tram]
                if _hoja_tiene_filas_contratista(ws_tram):
                    advertencias.extend(
                        actualizar_hoja_tramites_sectores(
                            ws_tram,
                            mapa_k3_suspendidos,
                            fecha_analisis,
                        )
                    )
            except ValueError as e:
                observaciones.append(f"Trámites sectores: {e}")
            except Exception as e:
                observaciones.append(
                    f"Trámites sectores: error no previsto ({type(e).__name__}: {e})"
                )
        else:
            observaciones.append(
                "No se encontró pestaña Trámites sectores en el archivo de Contratos."
            )

    if mapa_k4_liquidados is not None:
        nombre_liq = resolver_hoja_liquidados_con_saldo(nombres_hojas)
        if nombre_liq:
            try:
                ws_liq = wb[nombre_liq]
                if _hoja_tiene_filas_contratista(ws_liq):
                    advertencias.extend(
                        actualizar_hoja_liquidados_con_saldo(
                            ws_liq,
                            mapa_k4_liquidados,
                            fecha_analisis,
                        )
                    )
            except ValueError as e:
                observaciones.append(f"Liquidados con saldo: {e}")
            except Exception as e:
                observaciones.append(
                    f"Liquidados con saldo: error no previsto "
                    f"({type(e).__name__}: {e})"
                )
        else:
            observaciones.append(
                "No se encontró pestaña Liquidados con saldo en el archivo de Contratos."
            )

    try:
        advertencias.extend(actualizar_estrategias_en_libro(wb, fecha_analisis))
    except ValueError as e:
        observaciones.append(f"Estrategias: {e}")
    except Exception as e:
        observaciones.append(
            f"Estrategias: error no previsto ({type(e).__name__}: {e})"
        )

    _preparar_workbook_antes_guardar(wb)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return _finalizar_xlsx_contratos(out.getvalue()), advertencias, observaciones


def _indice_columna_corte(columnas: list, fecha: datetime | date) -> str | None:
    """Busca columna existente: Saldo a 31 de mayo, SALDO A 31 DE MAYO, etc."""
    f = _fecha_datetime(fecha)
    objetivo = _normalizar(titulo_saldo_corte(f))
    for col in columnas:
        if _normalizar(str(col)) == objetivo:
            return col

    dia = str(calendar.monthrange(f.year, f.month)[1])
    mes = _normalizar(MESES_ES[f.month - 1])
    for col in columnas:
        n = _normalizar(str(col))
        if "saldo" in n and dia in n and mes in n:
            return col

    mes_cap = MESES_ES[f.month - 1][:1].upper() + MESES_ES[f.month - 1][1:]
    for col in columnas:
        if _normalizar(str(col)) == _normalizar(mes_cap):
            return col
    return None


def procesar_localidad_cxp(
    contratos_bytes: bytes,
    df_matriz: pd.DataFrame,
    localidad: str,
    fecha_analisis: datetime,
    nombre_contratos: str = "",
    nombre_matriz: str = "",
) -> dict[str, Any]:
    titulo_mes = titulo_saldo_corte(fecha_analisis)
    _, mapa_k4, grupos_k3 = preparar_indice_matriz(df_matriz, localidad)

    libro = pd.ExcelFile(BytesIO(contratos_bytes))
    nombre_hoja = resolver_hoja_cruce_cxp(list(libro.sheet_names))

    df_c = pd.read_excel(
        BytesIO(contratos_bytes), sheet_name=nombre_hoja, header=HEADER_CONTRATOS
    )

    col_nombre = _columna(df_c, "NOMBRE CONTRATISTA")
    col_cto = _columna(df_c, "No. de Cto", "Número Contrato", "Numero Contrato")
    col_anio = _columna(df_c, "AÑO SUSCRIPCIÓN", "ANO SUSCRIPCION", "Año Suscripción")
    col_aprop = _columna(df_c, "APROPIACION DISPONIBLE", "Apropiación", "Apropiacion")
    col_saldo_k = _columna(df_c, "SALDO FINAL", "Saldo Final")

    if not all([col_nombre, col_cto, col_anio, col_aprop, col_saldo_k]):
        raise ValueError(
            f"Contratos: faltan columnas para el cruce en «{nombre_hoja}»."
        )

    col_mes_existente = _indice_columna_corte(list(df_c.columns), fecha_analisis)
    if col_mes_existente:
        col_mes = col_mes_existente
        accion_columna = "actualizada"
        crear_columna = False
    else:
        col_mes = titulo_mes
        accion_columna = "creada"
        crear_columna = True

    conteo: dict[str, int] = {k: 0 for k in METODOS_LABEL}
    detalle_filas: list[dict] = []
    saldos_mes: list[float] = []
    valores_excel: dict[int, float | None] = {}

    for i, (_, row) in enumerate(df_c.iterrows()):
        nombre = row[col_nombre]
        if pd.isna(nombre) or not str(nombre).strip():
            continue

        saldo, metodo, texto_det = buscar_saldo_matriz(
            mapa_k4,
            grupos_k3,
            nombre,
            row[col_cto],
            row[col_anio],
            row[col_aprop],
            row[col_saldo_k],
        )
        conteo[metodo] = conteo.get(metodo, 0) + 1
        fila_excel = _fila_inicio_datos_contratos() + i

        if saldo is None:
            valores_excel[fila_excel] = None
            saldo_guardar = None
        else:
            valores_excel[fila_excel] = float(saldo)
            saldo_guardar = saldo
            if metodo != "verificar" and metodo != "sin_matriz":
                saldos_mes.append(float(saldo))

        candidatos_matriz: list[dict] = []
        if metodo == "verificar":
            k3 = clave_tres(nombre, row[col_cto], row[col_anio])
            if k3 in grupos_k3:
                candidatos_matriz = _listar_opciones_matriz(grupos_k3[k3])

        detalle_filas.append({
            "Tipo fila": TIPO_FILA_CONTRATOS,
            "Localidad": localidad,
            "NOMBRE CONTRATISTA": nombre,
            "No. de Cto": row[col_cto],
            "AÑO SUSCRIPCIÓN": row[col_anio],
            "APROPIACION DISPONIBLE": row[col_aprop],
            "SALDO FINAL (Contratos)": row[col_saldo_k],
            f"Saldo Matriz ({titulo_mes})": saldo_guardar,
            "Método": METODOS_LABEL.get(metodo, metodo),
            "Detalle": texto_det,
            "candidatos_matriz": candidatos_matriz,
        })

    from hoja_liquidados_con_saldo import preparar_mapa_k4_saldo_matriz
    from hoja_suspendidos import preparar_mapa_k3_saldo_estado

    mapa_k3_suspendidos = preparar_mapa_k3_saldo_estado(df_matriz, localidad)
    mapa_k4_liquidados = preparar_mapa_k4_saldo_matriz(df_matriz, localidad)

    bytes_export, advertencias_hojas, observaciones = (
        exportar_contratos_preservando_formato(
            contratos_bytes,
            fecha_analisis,
            valores_excel,
            crear_columna=crear_columna,
            titulo_columna=titulo_mes,
            mapa_k3_suspendidos=mapa_k3_suspendidos,
            mapa_k4_liquidados=mapa_k4_liquidados,
        )
    )

    out = BytesIO()
    out.write(bytes_export)
    out.seek(0)

    total_contratos = sum(conteo.values())
    resumen_metodos = [
        {
            "Método": METODOS_LABEL.get(codigo, codigo),
            "Contratos": cantidad,
        }
        for codigo, cantidad in sorted(conteo.items(), key=lambda x: -x[1])
        if cantidad > 0
    ]

    sin_resolver = conteo.get("verificar", 0) + conteo.get("sin_matriz", 0)
    ok = total_contratos - sin_resolver

    return {
        "localidad": localidad,
        "nombre_contratos": nombre_contratos,
        "nombre_matriz": nombre_matriz,
        "hoja_cruce": nombre_hoja,
        "columna_mes": col_mes,
        "accion_columna": accion_columna,
        "bytes_contratos": out.getvalue(),
        "total_contratos": total_contratos,
        "contratos_ok": ok,
        "sin_resolver": sin_resolver,
        "cxp_total": sum(saldos_mes),
        "conteo": conteo,
        "resumen_metodos": resumen_metodos,
        "detalle": detalle_filas,
        "advertencias_hojas": advertencias_hojas,
        "advertencias_suspendidos": advertencias_hojas,
        "observaciones": observaciones,
    }


__all__ = [
    "COL_DESEMPATE_MANUAL",
    "METODOS_LABEL",
    "aplicar_desempate_en_contratos",
    "buscar_saldo_matriz",
    "claves_pendientes_localidad",
    "construir_dataframe_revision",
    "parsear_mapa_desempate",
    "procesar_localidad_cxp",
    "quitar_autofiltros_xlsx",
    "recalcular_estadisticas_localidad",
    "resolver_hoja_cruce_cxp",
    "titulo_saldo_corte",
    "validar_desempate_completo",
]
