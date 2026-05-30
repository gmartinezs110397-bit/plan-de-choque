"""Actualización de la pestaña Suspendidos (seguimiento mensual desde Matriz)."""

from __future__ import annotations

import re
from copy import copy
from datetime import date, datetime
from typing import Any

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from constantes import HOJAS_SUSPENDIDOS
from cxp_cruce import (
    FILA_CONTEO_CONTRATOS,
    FILA_SUMA_CONTRATOS,
    MESES_ES,
    _celda_tiene_formula,
    _columna,
    _copiar_estilo_celda,
    _centrar_celdas_total,
    _celda_para_escribir,
    _copiar_estilo_celda_sin_relleno,
    _fecha_datetime,
    _fila_encabezado_contratos,
    _fila_encabezado_hoja_datos,
    _fila_inicio_datos_contratos,
    _fila_inicio_datos_hoja,
    _fila_tiene_contratista,
    _hoja_tiene_filas_contratista,
    _indice_columna_en_hoja,
    _normalizar,
    _ultima_columna_con_datos,
    clave_tres,
    preparar_indice_matriz,
    titulo_saldo_corte,
    titulo_saldo_corte_para_mes,
    titulos_columna_saldo_mes,
)

FILL_VERDE_NEON = PatternFill(fill_type="solid", fgColor="CCFF00")
FILL_AMARILLO = PatternFill(fill_type="solid", fgColor="FFFF00")
FILL_ENCABEZADO_AZUL = PatternFill(fill_type="solid", fgColor="BDD7EE")
FILL_ENCABEZADO_AMARILLO = PatternFill(fill_type="solid", fgColor="FFF2CC")

_MESES_POR_NOMBRE = {_normalizar(m): i + 1 for i, m in enumerate(MESES_ES)}

_RGB_AZUL_ENCABEZADO = (
    "BDD7EE",
    "9BC2E6",
    "8DB4E2",
    "4472C4",
    "5B9BD5",
    "DDEBF7",
    "DAEEF3",
    "C5D9F1",
)
_RGB_AMARILLO_ENCABEZADO = (
    "FFF2CC",
    "FFFF00",
    "FFEB9C",
    "FFC000",
    "FFE699",
    "FFF9C4",
    "FCE4D6",
)


def resolver_hoja_suspendidos(nombres_hojas: list[str]) -> str | None:
    for candidato in HOJAS_SUSPENDIDOS:
        if candidato in nombres_hojas:
            return candidato
    for nombre in nombres_hojas:
        if _normalizar(nombre) == "suspendidos":
            return nombre
    return None


def titulo_saldo_suspendidos(fecha: datetime | date) -> str:
    """Mismo título que Cps: Saldo a 31 de mayo, Saldo a 30 de abril, etc."""
    return titulo_saldo_corte(fecha)


def titulo_estado_suspendidos(fecha: datetime | date) -> str:
    mes = MESES_ES[_fecha_datetime(fecha).month - 1].upper()
    return f"ESTADO ACTUAL {mes}"


def _titulos_saldo_equivalentes(fecha: datetime | date) -> tuple[str, ...]:
    return titulos_columna_saldo_mes(fecha)


def _titulos_estado_equivalentes(fecha: datetime | date) -> tuple[str, ...]:
    mes = MESES_ES[_fecha_datetime(fecha).month - 1].upper()
    return (
        titulo_estado_suspendidos(fecha),
        f"ESTADO ACTUAL ({mes})",
        f"ESTADO ACTUAL A {mes}",
    )


def preparar_mapa_k3_saldo_estado(
    df_matriz,
    localidad: str,
) -> dict[str, dict[str, Any]]:
    """Por clave nombre+contrato+año: suma de saldos en Matriz y estado representativo."""
    df_loc, _, grupos_k3 = preparar_indice_matriz(df_matriz, localidad)
    col_estado = _columna(
        df_loc,
        "ESTADO ACTUAL",
        "Estado Actual",
        "Estado",
        "ESTADO",
    )
    mapa: dict[str, dict[str, Any]] = {}
    from cxp_cruce import _suma_saldos_grupo_matriz

    for k3, grupo in grupos_k3.items():
        saldo = _suma_saldos_grupo_matriz(grupo["_saldo"])
        estado = ""
        if col_estado:
            g = grupo.copy()
            g["_abs"] = g["_saldo"].abs()
            if g["_saldo"].notna().any():
                fila_ref = g.loc[g["_abs"].idxmax()]
            else:
                fila_ref = grupo.iloc[0]
            raw = fila_ref[col_estado]
            if raw is not None and str(raw).strip():
                estado = str(raw).strip()
        mapa[k3] = {"saldo": saldo, "estado": estado}
    return mapa


def _mes_desde_titulo(titulo: str) -> int | None:
    norm = _normalizar(str(titulo or ""))
    if not norm:
        return None

    m = re.search(r"\(([^)]+)\)", norm)
    if m:
        mes = _MESES_POR_NOMBRE.get(_normalizar(m.group(1)))
        if mes:
            return mes

    for prefijo in ("estado actual ", "estado actual a ", "saldo "):
        if norm.startswith(prefijo):
            resto = norm[len(prefijo) :].strip()
            mes = _MESES_POR_NOMBRE.get(resto)
            if mes:
                return mes

    if norm in _MESES_POR_NOMBRE:
        return _MESES_POR_NOMBRE[norm]

    if norm.startswith("saldo a "):
        for mes_nombre, num in _MESES_POR_NOMBRE.items():
            if mes_nombre in norm:
                return num
    return None


def _es_columna_saldo_mes(titulo: str) -> bool:
    norm = _normalizar(str(titulo or ""))
    if not norm:
        return False
    if norm.startswith("estado actual"):
        return False
    if "responsable" in norm or "nombre" in norm:
        return False
    mes = _mes_desde_titulo(titulo)
    return mes is not None and (
        norm == mes
        or norm.startswith("saldo ")
        or norm.startswith("saldo a ")
        or norm in _MESES_POR_NOMBRE
    )


def _es_columna_estado_mes(titulo: str) -> bool:
    norm = _normalizar(str(titulo or ""))
    return norm.startswith("estado actual") and _mes_desde_titulo(titulo) is not None


def _listar_pares_mes_seguimiento(ws) -> list[tuple[int, int, int]]:
    """(col_saldo, col_estado, número de mes) ordenados por mes."""
    fila_hdr = _fila_encabezado_hoja_datos(ws)
    saldos: dict[int, int] = {}
    estados: dict[int, int] = {}

    for col in range(1, ws.max_column + 1):
        val = ws.cell(fila_hdr, col).value
        if val is None or not str(val).strip():
            continue
        titulo = str(val).strip()
        mes = _mes_desde_titulo(titulo)
        if mes is None:
            continue
        if _es_columna_estado_mes(titulo):
            estados[mes] = col
        elif _es_columna_saldo_mes(titulo):
            saldos[mes] = col

    pares: list[tuple[int, int, int]] = []
    for mes in sorted(set(saldos) | set(estados)):
        col_s = saldos.get(mes)
        col_e = estados.get(mes)
        if col_s and col_e:
            pares.append((col_s, col_e, mes))
    return pares


def _listar_columnas_saldo_mes(ws) -> list[tuple[int, int]]:
    """(columna, número de mes) ordenados por mes — p. ej. Liquidados con saldo."""
    fila_hdr = _fila_encabezado_hoja_datos(ws)
    columnas: dict[int, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(fila_hdr, col).value
        if val is None or not str(val).strip():
            continue
        titulo = str(val).strip()
        if not _es_columna_saldo_mes(titulo):
            continue
        mes = _mes_desde_titulo(titulo)
        if mes is not None:
            columnas[mes] = col
    return [(columnas[m], m) for m in sorted(columnas)]


def _titulo_saldo_mes_numero(mes_num: int, anio: int | None = None) -> str:
    year = anio if anio is not None else datetime.now().year
    return titulo_saldo_corte_para_mes(year, mes_num)


def _columna_saldo_mes_en_hoja(ws, fecha: datetime | date) -> int | None:
    """
    Columna del mes en curso (reconoce «Saldo a 31 de mayo», «SALDO MAYO», etc.).
    """
    mes_actual = _fecha_datetime(fecha).month
    for col, mes in _listar_columnas_saldo_mes(ws):
        if mes == mes_actual:
            return col
    return _indice_columna_titulos(ws, _titulos_saldo_equivalentes(fecha))


def _titulo_estado_mes_numero(mes_num: int) -> str:
    return f"ESTADO ACTUAL {MESES_ES[mes_num - 1].upper()}"


def _reforzar_titulos_pares_mes_seguimiento(ws, anio: int) -> None:
    """Asegura títulos SALDO/ESTADO del mes tras copiar estilos de encabezado."""
    fila_hdr = _fila_encabezado_hoja_datos(ws)
    for col_saldo, col_estado, mes_num in _listar_pares_mes_seguimiento(ws):
        _celda_para_escribir(ws, fila_hdr, col_saldo).value = _titulo_saldo_mes_numero(
            mes_num, anio
        )
        _celda_para_escribir(ws, fila_hdr, col_estado).value = _titulo_estado_mes_numero(
            mes_num
        )


def _indice_columna_titulos(ws, titulos: tuple[str, ...]) -> int | None:
    fila_hdr = _fila_encabezado_hoja_datos(ws)
    objetivos = {_normalizar(t) for t in titulos}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(fila_hdr, col).value
        if val is not None and _normalizar(str(val)) in objetivos:
            return col
    return None


def _par_mes_anterior(
    ws,
    fecha: datetime | date,
    col_saldo: int,
    col_estado: int,
) -> tuple[int | None, int | None]:
    mes_actual = _fecha_datetime(fecha).month
    pares = _listar_pares_mes_seguimiento(ws)
    previas = [p for p in pares if p[2] < mes_actual]
    if previas:
        return previas[-1][0], previas[-1][1]

    if pares:
        idx_actual = next(
            (i for i, p in enumerate(pares) if p[0] == col_saldo and p[1] == col_estado),
            None,
        )
        if idx_actual is not None and idx_actual > 0:
            return pares[idx_actual - 1][0], pares[idx_actual - 1][1]
        if pares[-1][0] != col_saldo:
            return pares[-1][0], pares[-1][1]
        if len(pares) >= 2:
            return pares[-2][0], pares[-2][1]
    return None, None


def _rgb_celda(celda) -> str:
    color = celda.fill.fgColor if celda.fill else None
    rgb = getattr(color, "rgb", None) if color else None
    if not rgb:
        return ""
    s = str(rgb).upper()
    if len(s) == 8:
        return s[2:]
    return s


def _es_fill_azul_encabezado(celda) -> bool:
    rgb = _rgb_celda(celda)
    if not rgb:
        return False
    return any(a in rgb for a in _RGB_AZUL_ENCABEZADO)


def _es_fill_amarillo_encabezado(celda) -> bool:
    rgb = _rgb_celda(celda)
    if not rgb:
        return False
    return any(a in rgb for a in _RGB_AMARILLO_ENCABEZADO)


def _fill_encabezado_alterno(celda_referencia) -> PatternFill:
    if _es_fill_azul_encabezado(celda_referencia):
        return FILL_ENCABEZADO_AMARILLO
    if _es_fill_amarillo_encabezado(celda_referencia):
        return FILL_ENCABEZADO_AZUL
    rgb = _rgb_celda(celda_referencia)
    if rgb:
        return FILL_ENCABEZADO_AMARILLO
    return FILL_ENCABEZADO_AZUL


def _fill_encabezado_por_mes(mes_num: int) -> PatternFill:
    """
    Color de título según el mes (misma regla en todas las pestañas).
    Mes par (p. ej. abril=4) → azul; mes impar (p. ej. mayo=5) → amarillo.
    """
    return FILL_ENCABEZADO_AMARILLO if (mes_num % 2) else FILL_ENCABEZADO_AZUL


def _copiar_estilo_columna(
    ws,
    col_origen: int,
    col_destino: int,
    fila_desde: int = 1,
    fila_hasta: int | None = None,
    *,
    sin_relleno_desde_fila: int | None = None,
) -> None:
    if fila_hasta is None:
        fila_hasta = ws.max_row
    fila_datos = sin_relleno_desde_fila or _fila_inicio_datos_hoja(ws)
    for fila in range(fila_desde, fila_hasta + 1):
        origen = ws.cell(fila, col_origen)
        destino = ws.cell(fila, col_destino)
        if fila >= fila_datos:
            _copiar_estilo_celda_sin_relleno(origen, destino)
        else:
            _copiar_estilo_celda(origen, destino)


def _copiar_ancho_columna(ws, col_origen: int, col_destino: int) -> None:
    letra_origen = get_column_letter(col_origen)
    letra_destino = get_column_letter(col_destino)
    ancho = ws.column_dimensions[letra_origen].width
    if ancho:
        ws.column_dimensions[letra_destino].width = ancho


def _copiar_formato_mes_desde_anterior(
    ws,
    col_saldo: int,
    col_estado: int,
    col_prev_saldo: int | None,
    col_prev_estado: int | None,
) -> None:
    """Fuente, bordes, alineación, número y ancho desde el par de columnas del mes anterior."""
    if not col_prev_saldo or not col_prev_estado:
        return
    _copiar_estilo_columna(ws, col_prev_saldo, col_saldo, sin_relleno_desde_fila=_fila_inicio_datos_hoja(ws))
    _copiar_estilo_columna(ws, col_prev_estado, col_estado, sin_relleno_desde_fila=_fila_inicio_datos_hoja(ws))
    _copiar_ancho_columna(ws, col_prev_saldo, col_saldo)
    _copiar_ancho_columna(ws, col_prev_estado, col_estado)


def _filas_encabezado_seguimiento(ws) -> tuple[int, int]:
    """Fila(s) del encabezado (p. ej. 1:2 si las columnas están combinadas en vertical)."""
    fila_hdr = _fila_encabezado_hoja_datos(ws)
    fila_fin = fila_hdr
    for rango in ws.merged_cells.ranges:
        if rango.min_row == fila_hdr and rango.max_row > fila_hdr:
            fila_fin = max(fila_fin, rango.max_row)
            break
    return fila_hdr, fila_fin


def _aplicar_merge_encabezado_columna(
    ws,
    col: int,
    fila_ini: int,
    fila_fin: int,
) -> None:
    if fila_fin <= fila_ini:
        return
    letra = get_column_letter(col)
    objetivo = f"{letra}{fila_ini}:{letra}{fila_fin}"
    for rango in list(ws.merged_cells.ranges):
        if (
            rango.min_col <= col <= rango.max_col
            and rango.min_row <= fila_fin
            and rango.max_row >= fila_ini
        ):
            ws.unmerge_cells(str(rango))
    ws.merge_cells(objetivo)


def _aplicar_merge_encabezados_mes(ws, col_saldo: int, col_estado: int) -> None:
    fila_ini, fila_fin = _filas_encabezado_seguimiento(ws)
    _aplicar_merge_encabezado_columna(ws, col_saldo, fila_ini, fila_fin)
    _aplicar_merge_encabezado_columna(ws, col_estado, fila_ini, fila_fin)


def _normalizar_titulos_mes_cortos(ws, fecha: datetime | date) -> None:
    """Renombra encabezados sueltos «ABRIL»/«MAYO» a «SALDO {MES}»."""
    fila_hdr = _fila_encabezado_hoja_datos(ws)
    for col in range(1, ws.max_column + 1):
        val = ws.cell(fila_hdr, col).value
        if val is None or not str(val).strip():
            continue
        titulo = str(val).strip()
        if _es_columna_estado_mes(titulo):
            continue
        mes = _mes_desde_titulo(titulo)
        if mes is None:
            continue
        if _normalizar(titulo) != _normalizar(MESES_ES[mes - 1]):
            continue
        _celda_para_escribir(ws, fila_hdr, col).value = _titulo_saldo_mes_numero(
            mes, _fecha_datetime(fecha).year
        )


def _aplicar_titulos_encabezado_mes(
    ws,
    col_saldo: int,
    col_estado: int,
    fecha: datetime | date,
    col_prev_saldo: int | None,
    col_prev_estado: int | None,
) -> None:
    del col_prev_saldo, col_prev_estado
    fila_hdr = _fila_encabezado_hoja_datos(ws)
    titulo_saldo = titulo_saldo_suspendidos(fecha)
    titulo_estado = titulo_estado_suspendidos(fecha)

    _celda_para_escribir(ws, fila_hdr, col_saldo).value = titulo_saldo
    _celda_para_escribir(ws, fila_hdr, col_estado).value = titulo_estado
    _aplicar_merge_encabezados_mes(ws, col_saldo, col_estado)


def _aplicar_encabezados_meses_alternos(ws) -> None:
    """
    Encabezados de mes con alternancia azul/amarillo entre meses.
    SALDO y ESTADO del mismo mes comparten el mismo color de título.
    """
    fila_hdr = _fila_encabezado_hoja_datos(ws)
    pares = _listar_pares_mes_seguimiento(ws)
    if not pares:
        return

    for col_saldo, col_estado, mes_num in pares:
        celda_saldo = _celda_para_escribir(ws, fila_hdr, col_saldo)
        celda_estado = _celda_para_escribir(ws, fila_hdr, col_estado)
        _aplicar_merge_encabezados_mes(ws, col_saldo, col_estado)

        fill_titulo = _fill_encabezado_por_mes(mes_num)

        _copiar_estilo_celda(celda_saldo, celda_estado)
        celda_saldo.fill = fill_titulo
        celda_estado.fill = fill_titulo


def _aplicar_encabezados_saldo_mes_alternos(ws, fecha: datetime | date) -> None:
    """
    Alternancia azul/amarillo por mes en columnas solo SALDO (Liquidados con saldo).
    """
    anio = _fecha_datetime(fecha).year
    fila_hdr = _fila_encabezado_hoja_datos(ws)
    fila_ini, fila_fin = _filas_encabezado_seguimiento(ws)
    cols = _listar_columnas_saldo_mes(ws)
    if not cols:
        return

    for col, mes_num in cols:
        celda = _celda_para_escribir(ws, fila_hdr, col)
        _aplicar_merge_encabezado_columna(ws, col, fila_ini, fila_fin)
        celda.fill = _fill_encabezado_por_mes(mes_num)
        celda.value = _titulo_saldo_mes_numero(mes_num, anio)


def _es_suspendido(valor) -> bool:
    return "suspendido" in _normalizar(str(valor or ""))


def _celda_tiene_relleno_verde(celda) -> bool:
    fill = celda.fill
    if not fill or fill.fill_type != "solid":
        return False
    color = fill.fgColor
    rgb = getattr(color, "rgb", None) if color else None
    if not rgb:
        return False
    s = str(rgb).upper()
    verdes = ("CCFF00", "39FF14", "00FF00", "92D050", "C6EFCE", "00B050")
    return any(v in s for v in verdes)


def _celda_tiene_relleno_amarillo(celda) -> bool:
    fill = celda.fill
    if not fill or fill.fill_type != "solid":
        return False
    rgb = getattr(fill.fgColor, "rgb", None) if fill.fgColor else None
    if not rgb:
        return False
    s = str(rgb).upper()
    return "FFFF00" in s or "FFEB9C" in s or "FFC000" in s


def _asegurar_columnas_mes(
    ws,
    fecha: datetime | date,
) -> tuple[int, int, int | None, int | None, bool]:
    """
    Devuelve (col_saldo, col_estado, col_prev_saldo, col_prev_estado, columnas_nuevas).
    """
    _normalizar_titulos_mes_cortos(ws, fecha)
    col_saldo = _columna_saldo_mes_en_hoja(ws, fecha)
    col_estado = _indice_columna_titulos(ws, _titulos_estado_equivalentes(fecha))
    columnas_nuevas = col_saldo is None or col_estado is None

    col_prev_saldo, col_prev_estado = _par_mes_anterior(
        ws, fecha, col_saldo or 0, col_estado or 0
    )

    siguiente = _ultima_columna_con_datos(ws) + 1
    if col_saldo is None:
        col_saldo = siguiente
        siguiente += 1
    if col_estado is None:
        col_estado = siguiente

    fila_hdr = _fila_encabezado_hoja_datos(ws)

    if columnas_nuevas and col_prev_saldo and col_prev_estado:
        _copiar_estilo_columna(ws, col_prev_saldo, col_saldo)
        _copiar_estilo_columna(ws, col_prev_estado, col_estado)
    elif columnas_nuevas:
        col_estilo = _indice_columna_en_hoja(ws, "ESTADO ACTUAL") or col_prev_estado
        if col_estilo:
            _copiar_estilo_columna(ws, col_estilo, col_estado)
        col_sf = _indice_columna_en_hoja(ws, "SALDO FINAL", "Saldo Final")
        if col_sf:
            _copiar_estilo_columna(ws, col_sf, col_saldo)

    if columnas_nuevas or (
        _normalizar(str(ws.cell(fila_hdr, col_saldo).value or ""))
        != _normalizar(titulo_saldo_suspendidos(fecha))
    ):
        _aplicar_titulos_encabezado_mes(
            ws, col_saldo, col_estado, fecha, col_prev_saldo, col_prev_estado
        )

    if columnas_nuevas and col_prev_saldo and col_prev_estado:
        fila_tot = _fila_totales_seguimiento(ws)
        _copiar_estilo_celda(ws.cell(fila_tot, col_prev_saldo), ws.cell(fila_tot, col_saldo))
        _copiar_estilo_celda(ws.cell(fila_tot, col_prev_estado), ws.cell(fila_tot, col_estado))

    if not col_prev_saldo or not col_prev_estado:
        col_prev_saldo, col_prev_estado = _par_mes_anterior(
            ws, fecha, col_saldo, col_estado
        )

    _copiar_formato_mes_desde_anterior(
        ws, col_saldo, col_estado, col_prev_saldo, col_prev_estado
    )
    _aplicar_encabezados_meses_alternos(ws)
    _reforzar_titulos_pares_mes_seguimiento(ws, _fecha_datetime(fecha).year)

    return col_saldo, col_estado, col_prev_saldo, col_prev_estado, columnas_nuevas


def _resolver_relleno_estado(
    prev_estado,
    prev_celda,
    curr_estado,
) -> PatternFill | None:
    prev_txt = str(prev_estado or "").strip()
    curr_txt = str(curr_estado or "").strip()
    prev_susp = _es_suspendido(prev_txt)
    curr_susp = _es_suspendido(curr_txt)
    prev_verde = _celda_tiene_relleno_verde(prev_celda) if prev_celda else False
    prev_amarillo = _celda_tiene_relleno_amarillo(prev_celda) if prev_celda else False

    if _normalizar(prev_txt) == _normalizar(curr_txt):
        if prev_verde and not curr_susp:
            return FILL_VERDE_NEON
        if prev_amarillo and curr_susp:
            return FILL_AMARILLO
        return None

    if not curr_susp:
        return FILL_VERDE_NEON

    if prev_verde or (prev_txt and not prev_susp):
        return FILL_AMARILLO

    if not prev_susp and prev_txt:
        return FILL_AMARILLO

    return None


def _ultima_fila_contratista(ws) -> int | None:
    col_nombre = _indice_columna_en_hoja(ws, "NOMBRE CONTRATISTA")
    if not col_nombre:
        return None
    fila_ini = _fila_inicio_datos_hoja(ws)
    ultima: int | None = None
    for fila in range(fila_ini, ws.max_row + 1):
        if _fila_tiene_contratista(ws, fila, col_nombre):
            ultima = fila
    return ultima


def _fila_totales_seguimiento(ws) -> int:
    """Fila al pie de la lista de contratos (conteo en col. estado, suma en col. saldo)."""
    ultima = _ultima_fila_contratista(ws)
    if ultima is not None:
        return ultima + 1
    return _fila_inicio_datos_hoja(ws)


def _rango_filas_datos_seguimiento(ws) -> tuple[int, int]:
    """(primera fila contrato, última fila contrato) sin incluir la fila de totales."""
    fila_ini = _fila_inicio_datos_hoja(ws)
    ultima = _ultima_fila_contratista(ws)
    if ultima is None:
        return fila_ini, fila_ini - 1
    return fila_ini, ultima


def _conteo_suspendidos_en_columna(ws, col_estado: int) -> int:
    fila_ini, fila_fin = _rango_filas_datos_seguimiento(ws)
    total = 0
    for fila in range(fila_ini, fila_fin + 1):
        if _es_suspendido(ws.cell(fila, col_estado).value):
            total += 1
    return total


def _conteo_suspendidos_mes_anterior(
    ws,
    col_prev_estado: int | None,
) -> int | None:
    if not col_prev_estado:
        return None
    fila_tot = _fila_totales_seguimiento(ws)
    celda_prev = ws.cell(fila_tot, col_prev_estado)
    if celda_prev.value not in (None, ""):
        try:
            return int(float(celda_prev.value))
        except (TypeError, ValueError):
            pass
    return _conteo_suspendidos_en_columna(ws, col_prev_estado)


def _leer_total_reportado_mes_anterior(
    ws,
    col_prev: int | None,
    *,
    fila_total: int | None = None,
) -> float | None:
    """Valor reportado al pie de la columna del mes anterior."""
    if not col_prev:
        return None
    fila = fila_total if fila_total is not None else _fila_totales_seguimiento(ws)
    celda = ws.cell(fila, col_prev)
    if celda.value in (None, ""):
        return None
    try:
        return float(celda.value)
    except (TypeError, ValueError):
        return None


def _aplicar_tope_mes_anterior(
    valor_real: float | int,
    valor_anterior: float | None,
) -> tuple[float | int, float | int]:
    """
    No se puede reportar un número mayor al del mes anterior.
    Devuelve (valor_a_mostrar, valor_real).
    """
    if valor_anterior is None:
        return valor_real, valor_real
    if float(valor_real) > float(valor_anterior):
        return valor_anterior, valor_real
    return valor_real, valor_real


def _suma_saldos_columna(ws, col_saldo: int) -> float:
    fila_ini, fila_fin = _rango_filas_datos_seguimiento(ws)
    suma = 0.0
    for fila in range(fila_ini, fila_fin + 1):
        raw = ws.cell(fila, col_saldo).value
        if raw is None or raw == "":
            continue
        try:
            suma += float(raw)
        except (TypeError, ValueError):
            continue
    return suma


def _actualizar_resumen_suspendidos(
    ws,
    col_saldo: int,
    col_estado: int,
    col_prev_saldo: int | None,
    col_prev_estado: int | None,
) -> tuple[int, int]:
    """Última fila: conteo en col. estado (tope vs mes anterior) y suma en col. saldo."""
    conteo_real = _conteo_suspendidos_en_columna(ws, col_estado)
    conteo_prev = _conteo_suspendidos_mes_anterior(ws, col_prev_estado)
    if conteo_prev is None and col_prev_estado:
        reportado = _leer_total_reportado_mes_anterior(ws, col_prev_estado)
        if reportado is not None:
            conteo_prev = int(reportado)
    conteo_mostrar, _ = _aplicar_tope_mes_anterior(
        conteo_real,
        float(conteo_prev) if conteo_prev is not None else None,
    )
    conteo_mostrar = int(conteo_mostrar)

    suma = _suma_saldos_columna(ws, col_saldo)

    fila_tot = _fila_totales_seguimiento(ws)
    celda_conteo = _celda_para_escribir(ws, fila_tot, col_estado)
    celda_suma = _celda_para_escribir(ws, fila_tot, col_saldo)

    if col_prev_estado:
        _copiar_estilo_celda(
            ws.cell(fila_tot, col_prev_estado),
            celda_conteo,
        )
    if col_prev_saldo:
        _copiar_estilo_celda(
            ws.cell(fila_tot, col_prev_saldo),
            celda_suma,
        )

    if not _celda_tiene_formula(celda_conteo):
        celda_conteo.value = conteo_mostrar
        celda_conteo.fill = PatternFill(fill_type=None)

    if not _celda_tiene_formula(celda_suma):
        celda_suma.value = suma

    _centrar_celdas_total(celda_conteo, celda_suma)

    return conteo_real, conteo_mostrar


def _autoajustar_columna_suspendidos(ws, col: int, titulo: str = "") -> None:
    from cxp_cruce import _autoajustar_ancho_columna

    _autoajustar_ancho_columna(ws, col)
    if titulo:
        letra = get_column_letter(col)
        min_titulo = min(max(len(titulo) * 1.15 + 3, 11), 60)
        ws.column_dimensions[letra].width = max(
            ws.column_dimensions[letra].width or 0,
            min_titulo,
        )


def actualizar_hoja_suspendidos(
    ws,
    mapa_k3: dict[str, dict[str, Any]],
    fecha: datetime | date,
) -> list[str]:
    """
    Escribe SALDO MES y ESTADO ACTUAL MES desde Matriz.
    Colores solo en celdas de estado del mes actual.
    """
    if not _hoja_tiene_filas_contratista(ws):
        return []

    advertencias: list[str] = []
    col_saldo, col_estado, col_prev_saldo, col_prev_estado, _ = _asegurar_columnas_mes(
        ws, fecha
    )

    col_nombre = _indice_columna_en_hoja(ws, "NOMBRE CONTRATISTA")
    col_cto = _indice_columna_en_hoja(
        ws, "No. de Cto", "Número Contrato", "Numero Contrato"
    )
    col_anio = _indice_columna_en_hoja(
        ws,
        "AÑO SUSCRIPCIÓN",
        "ANO SUSCRIPCION",
        "Año Suscripción",
        "Ano Suscripcion",
    )
    if not all([col_nombre, col_cto, col_anio]):
        raise ValueError(
            "Suspendidos: faltan columnas NOMBRE CONTRATISTA, No. de Cto o AÑO SUSCRIPCIÓN."
        )

    fila_ini = _fila_inicio_datos_hoja(ws)

    for fila in range(fila_ini, ws.max_row + 1):
        if not _fila_tiene_contratista(ws, fila, col_nombre):
            continue

        nombre = ws.cell(fila, col_nombre).value
        contrato = ws.cell(fila, col_cto).value
        anio = ws.cell(fila, col_anio).value
        k3 = clave_tres(nombre, contrato, anio)
        datos = mapa_k3.get(k3)

        celda_saldo = _celda_para_escribir(ws, fila, col_saldo)
        celda_estado = _celda_para_escribir(ws, fila, col_estado)

        if col_prev_saldo:
            _copiar_estilo_celda(ws.cell(fila, col_prev_saldo), celda_saldo)
        if datos is None:
            celda_saldo.value = None
            estado = ""
        else:
            saldo_celda = datos.get("saldo")
            celda_saldo.value = (
                None
                if saldo_celda is None
                else saldo_celda
            )
            estado = datos.get("estado", "")
        prev_celda = None
        prev_estado = None
        if col_prev_estado:
            prev_celda = ws.cell(fila, col_prev_estado)
            prev_estado = prev_celda.value
            _copiar_estilo_celda(prev_celda, celda_estado)

        celda_estado.value = estado or None

        relleno = _resolver_relleno_estado(prev_estado, prev_celda, estado)
        if relleno:
            celda_estado.fill = relleno

    conteo_real, conteo_mostrar = _actualizar_resumen_suspendidos(
        ws, col_saldo, col_estado, col_prev_saldo, col_prev_estado
    )

    if col_prev_estado and conteo_real > conteo_mostrar:
        advertencias.append(
            f"Suspendidos: el conteo real ({conteo_real}) supera al mes anterior "
            f"({conteo_mostrar}); en el total se dejó el valor del mes anterior."
        )

    _autoajustar_columna_suspendidos(ws, col_saldo, titulo_saldo_suspendidos(fecha))
    _autoajustar_columna_suspendidos(ws, col_estado, titulo_estado_suspendidos(fecha))

    return advertencias


__all__ = [
    "actualizar_hoja_suspendidos",
    "preparar_mapa_k3_saldo_estado",
    "resolver_hoja_suspendidos",
    "titulo_estado_suspendidos",
    "titulo_saldo_suspendidos",
    "_fila_totales_seguimiento",
    "_rango_filas_datos_seguimiento",
    "_leer_total_reportado_mes_anterior",
]
