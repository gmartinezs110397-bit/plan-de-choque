"""Compara consolidación automática vs archivo manual Usaquén Mayo 2026."""

from __future__ import annotations

import re
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

_APP = Path(__file__).resolve().parent
if str(_APP) not in sys.path:
    sys.path.insert(0, str(_APP))

from cxp_cruce import (
    FILA_CONTEO_CONTRATOS,
    FILA_SUMA_CONTRATOS,
    MESES_ES,
    _fila_encabezado_contratos,
    _indice_columna_en_hoja,
    _normalizar,
    procesar_localidad_cxp,
    resolver_hoja_cruce_cxp,
    titulo_saldo_corte,
)
from hoja_estrategias import (
    COL_CONTRATOS,
    COL_MONTO,
    FILA_TITULOS as FILA_TIT_ESTR,
    actualizar_estrategias_en_libro,
    resolver_hoja_estrategias,
)
from hoja_liquidados_con_saldo import (
    _fila_conteo_liquidados,
    _fila_suma_liquidados,
    resolver_hoja_liquidados_con_saldo,
)
from hoja_proximos_a_perder import resolver_hoja_proximos_a_perder
from hoja_suspendidos import (
    _es_columna_estado_mes,
    _es_columna_saldo_mes,
    _listar_pares_mes_seguimiento,
    _mes_desde_titulo,
    resolver_hoja_suspendidos,
    titulo_estado_suspendidos,
    titulo_saldo_suspendidos,
)
from hoja_tramites_sectores import resolver_hoja_tramites_sectores

DOWNLOADS = Path(r"C:\Users\f1rac\Downloads")
LOCALIDAD = "Usaquén"
FECHA = datetime(2026, 5, 29)
PWD_MATRIZ = "1100"


def _buscar_archivo(patron: str) -> Path:
    for p in DOWNLOADS.glob(patron):
        return p
    raise FileNotFoundError(f"No se encontró {patron} en Downloads")


def _rgb(celda) -> str:
    fill = celda.fill
    if not fill or fill.fill_type != "solid":
        return ""
    color = fill.fgColor
    rgb = getattr(color, "rgb", None) if color else None
    if not rgb:
        return ""
    s = str(rgb).upper()
    return s[2:] if len(s) == 8 else s


def _es_verde(rgb: str) -> bool:
    if not rgb:
        return False
    verdes = ("CCFF00", "39FF14", "00FF00", "92D050", "C6EFCE", "00B050")
    return any(v in rgb for v in verdes)


def _es_amarillo(rgb: str) -> bool:
    if not rgb:
        return False
    return "FFFF00" in rgb or "FFEB9C" in rgb or "FFC000" in rgb or "FFF2CC" in rgb


def analizar_formato_hoja(ws, nombre_hoja: str) -> dict:
    fila_hdr = _fila_encabezado_contratos()
    cols_mes = []
    for col in range(1, min(ws.max_column + 1, 80)):
        val = ws.cell(fila_hdr, col).value
        if val is None:
            continue
        titulo = str(val).strip()
        if _es_columna_saldo_mes(titulo) or _es_columna_estado_mes(titulo):
            hdr_rgb = _rgb(ws.cell(fila_hdr, col))
            cols_mes.append(
                {
                    "col": col,
                    "titulo": titulo,
                    "hdr_rgb": hdr_rgb,
                    "hdr_azul": hdr_rgb in ("BDD7EE", "9BC2E6", "8DB4E2", "4472C4", "DDEBF7"),
                    "hdr_amarillo": _es_amarillo(hdr_rgb),
                }
            )

    verde_estado = amarillo_estado = sin_color = 0
    col_estado_mayo = None
    for c in cols_mes:
        if "ESTADO" in _normalizar(c["titulo"]) and "MAYO" in _normalizar(c["titulo"]):
            col_estado_mayo = c["col"]
            break
    if col_estado_mayo:
        col_nombre = _indice_columna_en_hoja(ws, "NOMBRE CONTRATISTA")
        fila_ini = 4
        for fila in range(fila_ini, ws.max_row + 1):
            if col_nombre:
                v = ws.cell(fila, col_nombre).value
                if v is None or not str(v).strip():
                    continue
            rgb = _rgb(ws.cell(fila, col_estado_mayo))
            if _es_verde(rgb):
                verde_estado += 1
            elif _es_amarillo(rgb):
                amarillo_estado += 1
            else:
                sin_color += 1

    return {
        "hoja": nombre_hoja,
        "columnas_mes": cols_mes[-6:],
        "estado_mayo_col": col_estado_mayo,
        "celdas_estado_mayo": {
            "verde": verde_estado,
            "amarillo": amarillo_estado,
            "sin_color": sin_color,
        },
        "fila1_e": {c["col"]: ws.cell(1, c["col"]).value for c in cols_mes if c["col"] <= 30},
    }


def leer_totales_hoja_manual(ws, fecha) -> tuple[int | float | None, float | None]:
    from hoja_estrategias import _leer_totales_hoja_liquidados, _leer_totales_hoja_par

    nombre = ws.title
    if resolver_hoja_liquidados_con_saldo([nombre]):
        return _leer_totales_hoja_liquidados(ws, fecha)
    if (
        resolver_hoja_suspendidos([nombre])
        or resolver_hoja_proximos_a_perder([nombre])
        or resolver_hoja_tramites_sectores([nombre])
    ):
        return _leer_totales_hoja_par(ws, fecha)
    return None, None


def leer_manual_completo(wb_manual) -> dict:
    out = {}
    for nombre in wb_manual.sheetnames:
        ws = wb_manual[nombre]
        out[nombre] = {
            "formato": analizar_formato_hoja(ws, nombre),
            "totales": leer_totales_hoja_manual(ws, FECHA),
        }
    if resolver_hoja_estrategias(wb_manual.sheetnames):
        ws = wb_manual[resolver_hoja_estrategias(wb_manual.sheetnames)]
        filas = []
        for fila in range(4, ws.max_row + 1):
            b = ws.cell(fila, 2).value
            if b is None or not str(b).strip():
                continue
            filas.append(
                {
                    "b": str(b).strip(),
                    "e": ws.cell(fila, COL_CONTRATOS).value,
                    "f": ws.cell(fila, COL_MONTO).value,
                }
            )
        out["__estrategias__"] = filas
    return out


def leer_cruce_saldos_manual(wb, col_titulo: str) -> pd.DataFrame:
    nombre = resolver_hoja_cruce_cxp(list(wb.sheetnames))
    ws = wb[nombre]
    fila_hdr = _fila_encabezado_contratos()
    col_mes = None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(fila_hdr, col).value
        if v and _normalizar(str(v)) == _normalizar(col_titulo):
            col_mes = col
            break
    if not col_mes:
        return pd.DataFrame()

    col_nombre = _indice_columna_en_hoja(ws, "NOMBRE CONTRATISTA")
    col_cto = _indice_columna_en_hoja(ws, "No. de Cto", "Número Contrato")
    col_anio = _indice_columna_en_hoja(ws, "AÑO SUSCRIPCIÓN", "ANO SUSCRIPCION")
    filas = []
    for fila in range(4, ws.max_row + 1):
        nom = ws.cell(fila, col_nombre).value if col_nombre else None
        if nom is None or not str(nom).strip():
            continue
        filas.append(
            {
                "fila": fila,
                "nombre": str(nom).strip(),
                "contrato": ws.cell(fila, col_cto).value if col_cto else "",
                "anio": ws.cell(fila, col_anio).value if col_anio else "",
                "saldo_manual": ws.cell(fila, col_mes).value,
            }
        )
    return pd.DataFrame(filas)


def main():
    path_contratos = _buscar_archivo("*Usaqu*Mayo*Final*.xlsx")
    path_matriz = _buscar_archivo("MATRIZ OxP-FDLUSA ACTUALIZADA 25-05-2026*.xlsx")

    print("Contratos manual:", path_contratos)
    print("Matriz:", path_matriz)

    with open(path_matriz, "rb") as f:
        matriz_bytes = f.read()
    with open(path_contratos, "rb") as f:
        contratos_manual_bytes = f.read()

    from app import leer_hoja_matriz

    df_matriz = leer_hoja_matriz(matriz_bytes, PWD_MATRIZ, path_matriz.name, header=6)

    resultado = procesar_localidad_cxp(
        contratos_manual_bytes,
        df_matriz,
        LOCALIDAD,
        FECHA,
        path_contratos.name,
        path_matriz.name,
    )

    bytes_auto = resultado["bytes_contratos"]
    wb_auto = load_workbook(BytesIO(bytes_auto), data_only=True)
    wb_manual = load_workbook(path_contratos, data_only=True)

    titulo_mes = titulo_saldo_corte(FECHA)
    print("\n=== Título mes corte ===", titulo_mes)

    manual_cruce = leer_cruce_saldos_manual(wb_manual, titulo_mes)
    auto_cruce = leer_cruce_saldos_manual(wb_auto, titulo_mes)

    merged = manual_cruce.merge(
        auto_cruce,
        on=["fila", "nombre", "contrato", "anio"],
        how="outer",
        suffixes=("_manual", "_auto"),
    )

    def _num(x):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return None
        try:
            return float(x)
        except (TypeError, ValueError):
            return None

    difs = []
    for _, r in merged.iterrows():
        m = _num(r.get("saldo_manual"))
        a = _num(r.get("saldo_auto"))
        if m is None and a is None:
            continue
        if m is None or a is None or abs(m - a) > 0.01:
            difs.append(
                {
                    "fila": r["fila"],
                    "nombre": r["nombre"],
                    "contrato": r["contrato"],
                    "manual": m,
                    "auto": a,
                    "diff": (a - m) if m is not None and a is not None else None,
                }
            )

    print(f"\n=== Cps/Caja — filas comparadas: {len(merged)}")
    print(f"Diferencias saldo (>0.01 o vacío): {len(difs)}")
    if difs:
        df_d = pd.DataFrame(difs[:25])
        print(df_d.to_string(index=False))
        if len(difs) > 25:
            print(f"... y {len(difs) - 25} más")

    print("\n=== FORMATO (manual) — resumen por hoja ===")
    manual_info = leer_manual_completo(wb_manual)
    for key, info in manual_info.items():
        if key == "__estrategias__":
            continue
        fmt = info.get("formato", {})
        tot = info.get("totales", (None, None))
        print(f"\n{fmt.get('hoja', key)}")
        if fmt.get("columnas_mes"):
            for c in fmt["columnas_mes"][-4:]:
                print(f"  col {c['col']}: {c['titulo'][:40]} hdr={c['hdr_rgb']}")
        em = fmt.get("celdas_estado_mayo")
        if em:
            print(f"  ESTADO MAYO colores: {em}")
        if tot[0] is not None or tot[1] is not None:
            print(f"  Totales leídos conteo={tot[0]} suma={tot[1]}")

    print("\n=== TOTALES manual vs auto (mismo mes) ===")
    hojas_map = [
        ("Suspendidos", resolver_hoja_suspendidos),
        ("Próximos a perder", resolver_hoja_proximos_a_perder),
        ("Trámites sectores", resolver_hoja_tramites_sectores),
        ("Liquidados con saldo", resolver_hoja_liquidados_con_saldo),
    ]
    for etiqueta, resolver in hojas_map:
        nm_m = resolver(wb_manual.sheetnames)
        nm_a = resolver(wb_auto.sheetnames)
        if not nm_m:
            print(f"{etiqueta}: falta en manual")
            continue
        t_m = leer_totales_hoja_manual(wb_manual[nm_m], FECHA)
        t_a = leer_totales_hoja_manual(wb_auto[nm_a], FECHA) if nm_a else (None, None)
        ok_c = t_m[0] == t_a[0] if t_m[0] is not None and t_a[0] is not None else "?"
        ok_s = abs((t_m[1] or 0) - (t_a[1] or 0)) < 0.02 if t_m[1] is not None and t_a[1] is not None else "?"
        print(f"{etiqueta}: manual conteo={t_m[0]} suma={t_m[1]} | auto conteo={t_a[0]} suma={t_a[1]} | OK conteo={ok_c} OK suma={ok_s}")

    print(f"\nCXP total proceso: {resultado['cxp_total']:.2f}")
    print(f"Contratos OK: {resultado['contratos_ok']}/{resultado['total_contratos']}")
    print(f"Sin resolver: {resultado['sin_resolver']}")
    print("Conteo métodos:", resultado["conteo"])

    out_path = _APP / "comparacion_usaquen_mayo.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="cruce_saldos", index=False)
        if difs:
            pd.DataFrame(difs).to_excel(writer, sheet_name="diferencias", index=False)
        pd.DataFrame(resultado["detalle"]).to_excel(writer, sheet_name="detalle_auto", index=False)
    print(f"\nExportado: {out_path}")


if __name__ == "__main__":
    main()
