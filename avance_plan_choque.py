from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LOCALIDADES_AVANCE = [
    (1, "USAQUEN "),
    (2, "CHAPINERO  "),
    (3, "SANTA  FE "),
    (4, "SAN  CRISTOBAL"),
    (5, "USME "),
    (6, "TUNJUELITO "),
    (7, "BOSA  "),
    (8, "KENNEDY"),
    (9, "FONTIBON"),
    (10, "ENGATIVA "),
    (11, "SUBA  "),
    (12, "BARRIOS UNIDOS "),
    (13, "TEUSAQUILLO "),
    (14, "LOS  MARTIRES "),
    (15, "ANTONIO  NARI\u00d1O "),
    (16, "PUENTE  ARANDA  "),
    (17, "LA  CANDELARIA  "),
    (18, "RAFAEL  URIBE URIBE "),
    (19, "CIUDAD  BOLIVAR"),
    (20, "SUMAPAZ"),
]

MESES_ES = {
    1: "enero",
    2: "febrero",
    3: "marzo",
    4: "abril",
    5: "mayo",
    6: "junio",
    7: "julio",
    8: "agosto",
    9: "septiembre",
    10: "octubre",
    11: "noviembre",
    12: "diciembre",
}

FILL_ROJO = PatternFill("solid", fgColor="C00000")
FILL_AMARILLO = PatternFill("solid", fgColor="FFC000")
FILL_BLANCO = PatternFill("solid", fgColor="FFFFFF")
BORDE_FINO = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="thin", color="808080"),
)
FUENTE_BLANCA = Font(bold=True, color="FFFFFF")
FUENTE_NEGRITA = Font(bold=True, color="000000")
FORMATO_MONEDA = '"$"\\ #,##0;[Red]\\-"$"\\ #,##0'
FORMATO_PORCENTAJE = "0.00%"


@dataclass(frozen=True)
class EstrategiaAvance:
    clave: str
    hoja: str
    titulo: str
    etiqueta_general: str
    texto_accion: str


@dataclass
class ValoresEstrategia:
    inicial_contratos: float = 0
    inicial_monto: float = 0
    actual_contratos: float = 0
    actual_monto: float = 0

    @property
    def avance_contratos(self) -> float:
        return self.inicial_contratos - self.actual_contratos

    @property
    def avance_porcentaje(self) -> float | None:
        if not self.inicial_contratos:
            return None
        return self.avance_contratos / self.inicial_contratos


@dataclass
class LocalidadAvance:
    numero: int
    localidad: str
    estrategias: dict[str, ValoresEstrategia]

    @property
    def inicial_contratos(self) -> float:
        return sum(v.inicial_contratos for v in self.estrategias.values())

    @property
    def inicial_monto(self) -> float:
        return sum(v.inicial_monto for v in self.estrategias.values())

    @property
    def actual_contratos(self) -> float:
        return sum(v.actual_contratos for v in self.estrategias.values())

    @property
    def actual_monto(self) -> float:
        return sum(v.actual_monto for v in self.estrategias.values())

    @property
    def avance_contratos(self) -> float:
        return self.inicial_contratos - self.actual_contratos

    @property
    def avance_porcentaje(self) -> float:
        if not self.inicial_contratos:
            return 0
        return self.avance_contratos / self.inicial_contratos


ESTRATEGIAS = [
    EstrategiaAvance(
        "suspendidos",
        "Suspendidos",
        "Contratos Suspendidos",
        "Contratos Suspendidos",
        (
            "El objetivo es la reactivaci\u00f3n de los contratos suspendidos "
            "priorizados con corte al 31 de marzo de 2026"
        ),
    ),
    EstrategiaAvance(
        "proximos",
        "Pr\u00f3ximos a perder",
        "Pr\u00f3ximos a Perder Competencia",
        "Contratos Pr\u00f3ximos a perder competencia",
        (
            "Liquidar los contratos priorizados con corte al 31 de marzo y que "
            "estar\u00edan pr\u00f3ximos a perder competencia antes de que se cumpla "
            "la fecha l\u00edmite para su liquidaci\u00f3n de conformidad con lo "
            "contemplado en el art\u00edculo 11 de la Ley 1150 de 2007"
        ),
    ),
    EstrategiaAvance(
        "tramites",
        "Tr\u00e1mites con sectores",
        "Tr\u00e1mites pendientes con sectores",
        "Tr\u00e1mites pendientes con sectores",
        (
            "Se considera cumplida la estrategia cuando el tr\u00e1mite que "
            "dificultaba la ejecuci\u00f3n o liquidaci\u00f3n sea realizado"
        ),
    ),
    EstrategiaAvance(
        "cps",
        "CPS",
        "CPS por depurar",
        "Contratos de Prestaci\u00f3n de Servicios",
        (
            "Se requiere la depuraci\u00f3n de este grupo de contratos que se "
            "encuentran con corte al 31 de marzo en estado terminado no se "
            "liquida o terminado en proceso de liquidaci\u00f3n."
        ),
    ),
    EstrategiaAvance(
        "liquidados",
        "Liquidados con saldo",
        "Liquidados con saldo",
        "Contratos liquidados con saldo",
        (
            "Se requiere que se depuren aquellos contratos que en la matriz de "
            "seguimiento a las obligaciones por pagar remitida por cada FDL, "
            "se encuentren registrados en estado liquidado y que tengan saldo."
        ),
    ),
]


def _normalizar(texto: object) -> str:
    raw = str(texto or "").strip().lower()
    raw = "".join(
        ch for ch in unicodedata.normalize("NFD", raw) if unicodedata.category(ch) != "Mn"
    )
    raw = re.sub(r"[^a-z0-9]+", " ", raw)
    return " ".join(raw.split())


def _localidades_por_normalizado() -> dict[str, tuple[int, str]]:
    mapa: dict[str, tuple[int, str]] = {}
    for numero, nombre in LOCALIDADES_AVANCE:
        mapa[_normalizar(nombre)] = (numero, nombre)
    return mapa


def _resolver_localidad(localidad: str) -> tuple[int, str]:
    mapa = _localidades_por_normalizado()
    norm = _normalizar(localidad)
    if norm in mapa:
        return mapa[norm]
    for clave, valor in mapa.items():
        if clave in norm or norm in clave:
            return valor
    return (999, str(localidad).strip().upper())


def _mes(fecha: datetime | date) -> str:
    return MESES_ES.get(fecha.month, str(fecha.month))


def _dia_fin_mes(fecha: datetime | date) -> int:
    import calendar

    return calendar.monthrange(fecha.year, fecha.month)[1]


def _encabezado_actual_contratos(fecha: datetime | date) -> str:
    return f"No. de contratos {_mes(fecha)} {_dia_fin_mes(fecha)}"


def _encabezado_actual_monto(fecha: datetime | date) -> str:
    return f"Monto Total {_mes(fecha)} {_dia_fin_mes(fecha)}"


def _valor_numero(valor: object) -> float:
    if valor in (None, ""):
        return 0
    try:
        return float(valor)
    except (TypeError, ValueError):
        return 0


def _clave_estrategia(texto: object) -> str | None:
    norm = _normalizar(texto)
    if not norm or norm == "total":
        return None
    if "suspend" in norm:
        return "suspendidos"
    if "proxim" in norm and "perder" in norm:
        return "proximos"
    if "tramit" in norm and "sector" in norm:
        return "tramites"
    if "liquid" in norm and "saldo" in norm:
        return "liquidados"
    if "cps" in norm or "depurar" in norm or "prestacion de servicios" in norm:
        return "cps"
    return None


def _resolver_hoja_estrategias(wb) -> str | None:
    for nombre in wb.sheetnames:
        if _normalizar(nombre) == "estrategias":
            return nombre
    return None


def _extraer_valores_estrategias(raw: bytes) -> dict[str, ValoresEstrategia]:
    wb = load_workbook(BytesIO(raw), data_only=False, read_only=True)
    try:
        nombre = _resolver_hoja_estrategias(wb)
        if not nombre:
            return {}
        ws = wb[nombre]
        valores: dict[str, ValoresEstrategia] = {}
        for fila in range(1, ws.max_row + 1):
            clave = _clave_estrategia(ws.cell(fila, 2).value)
            if not clave:
                continue
            valores[clave] = ValoresEstrategia(
                inicial_contratos=_valor_numero(ws.cell(fila, 3).value),
                inicial_monto=_valor_numero(ws.cell(fila, 4).value),
                actual_contratos=_valor_numero(ws.cell(fila, 5).value),
                actual_monto=_valor_numero(ws.cell(fila, 6).value),
            )
        return valores
    finally:
        wb.close()


def filas_avance_desde_contratos(
    contratos_por_localidad: Iterable[dict],
) -> list[LocalidadAvance]:
    filas: list[LocalidadAvance] = []
    for item in contratos_por_localidad:
        raw = item.get("bytes") or item.get("bytes_contratos")
        localidad_raw = str(item.get("localidad") or "").strip()
        if not raw or not localidad_raw:
            continue
        numero, localidad = _resolver_localidad(localidad_raw)
        valores = _extraer_valores_estrategias(raw)
        estrategias = {
            estrategia.clave: valores.get(estrategia.clave, ValoresEstrategia())
            for estrategia in ESTRATEGIAS
        }
        filas.append(LocalidadAvance(numero, localidad, estrategias))
    return sorted(filas, key=lambda fila: fila.numero)


def _set_value(ws, row: int, col: int, value, style: str | None = None) -> None:
    cell = ws.cell(row, col, value)
    cell.border = BORDE_FINO
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if style == "title":
        cell.fill = FILL_ROJO
        cell.font = FUENTE_BLANCA
    elif style == "header":
        cell.fill = FILL_AMARILLO
        cell.font = FUENTE_NEGRITA
    elif style == "total":
        cell.font = FUENTE_NEGRITA
    else:
        cell.fill = FILL_BLANCO


def _aplicar_formatos_tabla(ws, min_row: int, max_row: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if col in (3, 5, 12, 14):
                cell.number_format = FORMATO_MONEDA
            if col in (7, 16):
                cell.number_format = FORMATO_PORCENTAJE


def _escribir_hoja_estrategia(
    wb: Workbook,
    estrategia: EstrategiaAvance,
    filas: list[LocalidadAvance],
    fecha: datetime | date,
) -> int:
    ws = wb.create_sheet(estrategia.hoja)
    ws.merge_cells("A2:G2")
    ws.merge_cells("B3:G3")
    _set_value(ws, 2, 1, estrategia.titulo, "title")
    _set_value(ws, 3, 1, "Acciones a realizar", "header")
    _set_value(ws, 3, 2, estrategia.texto_accion, "header")

    headers = [
        "FDL",
        "No. de contratos marzo 31",
        "Monto Total marzo 31",
        _encabezado_actual_contratos(fecha),
        _encabezado_actual_monto(fecha),
        "Avance",
        "Avance",
    ]
    for idx, header in enumerate(headers, start=1):
        _set_value(ws, 4, idx, header, "header")

    row = 5
    for fila in filas:
        valores = fila.estrategias[estrategia.clave]
        _set_value(ws, row, 1, fila.localidad)
        _set_value(ws, row, 2, valores.inicial_contratos)
        _set_value(ws, row, 3, valores.inicial_monto)
        _set_value(ws, row, 4, valores.actual_contratos)
        _set_value(ws, row, 5, valores.actual_monto)
        if valores.inicial_contratos:
            _set_value(ws, row, 6, f"=B{row}-D{row}")
            _set_value(ws, row, 7, f"=F{row}/B{row}")
        else:
            _set_value(ws, row, 6, "N/A")
            _set_value(ws, row, 7, "N/A")
        row += 1

    total_row = row
    _set_value(ws, total_row, 1, "Total", "total")
    for col in range(2, 6):
        letter = get_column_letter(col)
        _set_value(ws, total_row, col, f"=SUM({letter}5:{letter}{total_row - 1})", "total")
    total_inicial = sum(
        fila.estrategias[estrategia.clave].inicial_contratos for fila in filas
    )
    if total_inicial:
        _set_value(ws, total_row, 6, f"=B{total_row}-D{total_row}", "total")
        _set_value(ws, total_row, 7, f"=F{total_row}/B{total_row}", "total")
    else:
        _set_value(ws, total_row, 6, "N/A", "total")
        _set_value(ws, total_row, 7, "N/A", "total")

    _aplicar_formatos_tabla(ws, 5, total_row, 7)
    widths = {"A": 24, "B": 15, "C": 18, "D": 15, "E": 18, "F": 13, "G": 13}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 42
    ws.row_dimensions[4].height = 38
    return total_row


def _fila_total_localidad(xloc_row: int) -> int:
    return xloc_row + 7


def _escribir_x_loc(
    wb: Workbook,
    filas: list[LocalidadAvance],
    estrategia_rows: dict[str, dict[int, int]],
    fecha: datetime | date,
) -> dict[int, int]:
    ws = wb.create_sheet("X Loc")
    block_rows: dict[int, int] = {}
    for idx, fila in enumerate(filas):
        base = 1 + (idx * 10)
        block_rows[fila.numero] = base
        ws.merge_cells(start_row=base, start_column=1, end_row=base, end_column=7)
        _set_value(ws, base, 1, f"Plan de Choque OxP {fila.localidad.title()}", "title")
        headers = [
            "Estrategia",
            "No. de contratos marzo 31",
            "Monto Total marzo 31",
            _encabezado_actual_contratos(fecha),
            _encabezado_actual_monto(fecha),
            "Avance",
            "Avance",
        ]
        for col, header in enumerate(headers, start=1):
            _set_value(ws, base + 1, col, header, "header")
        for offset, estrategia in enumerate(ESTRATEGIAS, start=2):
            row = base + offset
            source_row = estrategia_rows[estrategia.clave][fila.numero]
            sheet = estrategia.hoja.replace("'", "''")
            _set_value(ws, row, 1, estrategia.etiqueta_general)
            for col in range(2, 8):
                letter = get_column_letter(col)
                _set_value(ws, row, col, f"='{sheet}'!{letter}{source_row}")
        total_row = _fila_total_localidad(base)
        _set_value(ws, total_row, 1, "Total", "total")
        for col in range(2, 6):
            letter = get_column_letter(col)
            _set_value(ws, total_row, col, f"=SUM({letter}{base + 2}:{letter}{base + 6})", "total")
        if fila.inicial_contratos:
            _set_value(ws, total_row, 6, f"=B{total_row}-D{total_row}", "total")
            _set_value(ws, total_row, 7, f"=F{total_row}/B{total_row}", "total")
        else:
            _set_value(ws, total_row, 6, "N/A", "total")
            _set_value(ws, total_row, 7, "N/A", "total")
        _aplicar_formatos_tabla(ws, base + 2, total_row, 7)
    widths = {"A": 36, "B": 19, "C": 16, "D": 17, "E": 16, "F": 10, "G": 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    return block_rows


def _escribir_general(
    wb: Workbook,
    filas: list[LocalidadAvance],
    estrategia_total_rows: dict[str, int],
    xloc_block_rows: dict[int, int],
    fecha: datetime | date,
) -> None:
    ws = wb.create_sheet("General")
    ws.merge_cells("A2:G2")
    ws.merge_cells("I2:P2")
    _set_value(ws, 2, 1, f"Plan de Choque OxP Vigencia {fecha.year}", "title")
    _set_value(
        ws,
        2,
        9,
        f"Ranking Plan de Choque OxP Vigencia {fecha.year} Por localidad",
        "title",
    )
    headers = [
        "Estrategia",
        "No. de contratos marzo 31",
        "Monto Total marzo 31",
        _encabezado_actual_contratos(fecha),
        _encabezado_actual_monto(fecha),
        "Avance",
        "Avance",
    ]
    for col, header in enumerate(headers, start=1):
        _set_value(ws, 3, col, header, "header")
    ranking_headers = [
        "RK",
        "Localidad",
        "No. de contratos marzo 31",
        "Monto Total marzo 31",
        _encabezado_actual_contratos(fecha),
        _encabezado_actual_monto(fecha),
        "Avance",
        "Avance",
    ]
    for col, header in enumerate(ranking_headers, start=9):
        _set_value(ws, 3, col, header, "header")

    for offset, estrategia in enumerate(ESTRATEGIAS, start=4):
        total_row = estrategia_total_rows[estrategia.clave]
        sheet = estrategia.hoja.replace("'", "''")
        _set_value(ws, offset, 1, estrategia.etiqueta_general)
        for col in range(2, 6):
            letter = get_column_letter(col)
            _set_value(ws, offset, col, f"='{sheet}'!{letter}{total_row}")
        total_inicial = sum(
            fila.estrategias[estrategia.clave].inicial_contratos for fila in filas
        )
        if total_inicial:
            _set_value(ws, offset, 6, f"=B{offset}-D{offset}")
            _set_value(ws, offset, 7, f"=F{offset}/B{offset}")
        else:
            _set_value(ws, offset, 6, "N/A")
            _set_value(ws, offset, 7, "N/A")
    total_general = 4 + len(ESTRATEGIAS)
    _set_value(ws, total_general, 1, "Total", "total")
    for col in range(2, 7):
        letter = get_column_letter(col)
        _set_value(
            ws,
            total_general,
            col,
            f"=SUM({letter}4:{letter}{total_general - 1})",
            "total",
        )
    _set_value(ws, total_general, 7, f"=F{total_general}/B{total_general}", "total")

    ranking = sorted(
        filas,
        key=lambda fila: (fila.avance_porcentaje, fila.avance_contratos, -fila.numero),
        reverse=True,
    )
    for idx, fila in enumerate(ranking, start=1):
        row = 3 + idx
        total_row_xloc = _fila_total_localidad(xloc_block_rows[fila.numero])
        _set_value(ws, row, 9, idx)
        _set_value(ws, row, 10, fila.localidad)
        for col in range(11, 17):
            source_col = get_column_letter(col - 9)
            _set_value(ws, row, col, f"='X Loc'!{source_col}{total_row_xloc}")
    total_rank = 4 + len(ranking)
    _set_value(ws, total_rank, 9, "Total general", "total")
    for col in range(11, 16):
        letter = get_column_letter(col)
        _set_value(ws, total_rank, col, f"=SUM({letter}4:{letter}{total_rank - 1})", "total")
    total_inicial_localidades = sum(fila.inicial_contratos for fila in filas)
    if total_inicial_localidades:
        _set_value(ws, total_rank, 16, f"=O{total_rank}/K{total_rank}", "total")
    else:
        _set_value(ws, total_rank, 16, "N/A", "total")

    _aplicar_formatos_tabla(ws, 4, max(total_general, total_rank), 16)
    widths = {
        "A": 38,
        "B": 15,
        "C": 15,
        "D": 15,
        "E": 15,
        "F": 10,
        "G": 10,
        "I": 8,
        "J": 22,
        "K": 15,
        "L": 18,
        "M": 15,
        "N": 18,
        "O": 10,
        "P": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def crear_excel_avance_plan_de_choque(
    contratos_por_localidad: Iterable[dict],
    fecha_corte: datetime | date,
) -> bytes:
    filas = filas_avance_desde_contratos(contratos_por_localidad)
    if not filas:
        raise ValueError("No hay Contratos actualizados para construir el Avance.")

    wb = Workbook()
    wb.remove(wb.active)

    estrategia_total_rows: dict[str, int] = {}
    estrategia_rows: dict[str, dict[int, int]] = {}
    for estrategia in ESTRATEGIAS:
        total_row = _escribir_hoja_estrategia(wb, estrategia, filas, fecha_corte)
        estrategia_total_rows[estrategia.clave] = total_row
        estrategia_rows[estrategia.clave] = {
            fila.numero: idx for idx, fila in enumerate(filas, start=5)
        }

    xloc_rows = _escribir_x_loc(wb, filas, estrategia_rows, fecha_corte)
    _escribir_general(wb, filas, estrategia_total_rows, xloc_rows, fecha_corte)

    # Orden visual del ejemplo: estrategias, General y X Loc al final.
    wb._sheets = [
        wb[estrategia.hoja] for estrategia in ESTRATEGIAS
    ] + [wb["General"], wb["X Loc"]]
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
