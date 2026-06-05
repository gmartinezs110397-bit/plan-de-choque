from __future__ import annotations

import json
import pickle
import random
import re
import sys
import tempfile
import unicodedata
import uuid
import zipfile
from html import escape
from io import BytesIO
from datetime import datetime, date
from pathlib import Path

import streamlit as st
import streamlit.components.v1 as components

# Carpeta del proyecto primero (evita importar un cxp_cruce viejo en caché)
_APP_DIR = Path(__file__).resolve().parent
if str(_APP_DIR) not in sys.path:
    sys.path.insert(0, str(_APP_DIR))

# Localidades de Bogotá D.C. (20), orden alfabético
LOCALIDADES = [
    "Antonio Nariño",
    "Barrios Unidos",
    "Bosa",
    "Chapinero",
    "Ciudad Bolívar",
    "Engativá",
    "Fontibón",
    "Kennedy",
    "La Candelaria",
    "Los Mártires",
    "Puente Aranda",
    "Rafael Uribe Uribe",
    "San Cristóbal",
    "Santa Fe",
    "Suba",
    "Sumapaz",
    "Teusaquillo",
    "Tunjuelito",
    "Usaquén",
    "Usme",
]

st.set_page_config(
    page_title="Plan de Choque",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(
    """
    <style>
    html, body, [class*="css"] {
        font-family: 'Inter', system-ui, -apple-system, 'Segoe UI', Roboto, sans-serif;
    }
    #MainMenu, footer, header { visibility: hidden; }
    :root {
        --pc-pink-25: #fff9fc;
        --pc-pink-50: #fff1f8;
        --pc-pink-100: #ffe4f1;
        --pc-pink-200: #fbcfe8;
        --pc-pink-300: #f9a8d4;
        --pc-pink-500: #e0218a;
        --pc-pink-600: #c2186a;
        --pc-pink-700: #9d174d;
        --pc-ink: #5b0a37;
        --pc-muted: #8a3a63;
    }
    [data-testid="stAppViewContainer"] {
        background: #ffffff !important;
    }
    [data-testid="stHeader"] { background: transparent !important; }
    .block-container { padding-top: 1.25rem; max-width: 960px; }
    div[data-testid="stVerticalBlockBorderWrapper"] {
        background: #ffffff !important;
        border-color: var(--pc-pink-200) !important;
        box-shadow: 0 6px 18px rgba(224, 33, 138, 0.05) !important;
    }

    /* Solo ocultar el texto «Press Enter…», no el icono del ojo */
    [data-testid="stFormSubmitInstruction"],
    div[data-testid="InputInstructions"] > span {
        display: none !important;
    }
    h1.app-title,
    .app-title,
    .app-title * {
        font-size: 2rem;
        font-weight: 700;
        color: var(--pc-pink-600) !important;
        letter-spacing: -0.03em;
        margin: 0 0 0.35rem 0;
        text-align: center;
    }
    .app-subtitle,
    .app-subtitle * {
        text-align: center;
        color: var(--pc-muted) !important;
        font-size: 0.95rem;
        margin: 0 0 1.75rem 0;
    }
    .form-card-title {
        font-size: 1.05rem;
        font-weight: 600;
        color: var(--pc-ink);
        margin: 0 0 1.25rem 0;
        padding-bottom: 0.75rem;
        border-bottom: 2px solid var(--pc-pink-200);
    }
    .field-label {
        font-size: 0.8rem;
        font-weight: 600;
        color: var(--pc-pink-700);
        text-transform: uppercase;
        letter-spacing: 0.04em;
        margin-bottom: 0.35rem;
    }
    .field-num {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 1.35rem;
        height: 1.35rem;
        background: var(--pc-pink-500);
        color: white;
        border-radius: 50%;
        font-size: 0.72rem;
        font-weight: 700;
        margin-right: 0.5rem;
    }
    .file-ok { color: var(--pc-pink-600); font-size: 0.85rem; font-weight: 500; }
    .section-title {
        font-size: 1rem;
        font-weight: 600;
        color: var(--pc-ink);
        margin: 2rem 0 0.75rem;
        padding-bottom: 0.4rem;
        border-bottom: 2px solid var(--pc-pink-200);
    }
    .metric-card {
        background: #fff;
        border: 1px solid var(--pc-pink-200);
        border-radius: 8px;
        padding: 0.85rem 0.75rem;
        min-height: 5.25rem;
        display: flex;
        flex-direction: column;
        justify-content: center;
        overflow: hidden;
        box-shadow: 0 8px 24px rgba(224, 33, 138, 0.07);
    }
    .metric-label {
        font-size: 0.68rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.04em;
        color: var(--pc-muted);
        line-height: 1.25;
        margin-bottom: 0.35rem;
    }
    .metric-value {
        font-size: clamp(1rem, 2.4vw, 1.35rem);
        font-weight: 700;
        color: var(--pc-ink);
        line-height: 1.15;
        white-space: nowrap;
    }
    .metric-value-sm { font-size: clamp(1.05rem, 2.6vw, 1.45rem); }

    /* Select localidad — borde y foco fucsia */
    [class*="st-key-select_localidad"] [data-baseweb="select"] > div,
    [data-testid="stSelectbox"] [data-baseweb="select"] > div {
        border-color: var(--pc-pink-200) !important;
        border-radius: 8px !important;
    }
    [class*="st-key-select_localidad"] [data-baseweb="select"]:focus-within > div,
    [class*="st-key-select_localidad"] [data-baseweb="select"]:hover > div,
    [data-testid="stSelectbox"] [data-baseweb="select"]:focus-within > div,
    [data-testid="stSelectbox"] [data-baseweb="select"]:hover > div {
        border-color: var(--pc-pink-500) !important;
        box-shadow: 0 0 0 1px var(--pc-pink-500) !important;
    }
    div[data-baseweb="popover"] li[role="option"]:hover,
    div[data-baseweb="menu"] li[role="option"]:hover {
        background-color: var(--pc-pink-100) !important;
    }
    div[data-baseweb="popover"] li[role="option"][aria-selected="true"],
    div[data-baseweb="menu"] li[role="option"][aria-selected="true"] {
        background-color: var(--pc-pink-50) !important;
        color: var(--pc-pink-700) !important;
    }

    @keyframes pulse { 0%,100%{opacity:1} 50%{opacity:.5} }
    .loading-text { animation: pulse 1.5s ease-in-out infinite; color: var(--pc-pink-500); }
    [data-testid="stProgress"] > div > div:has([role="progressbar"]),
    [data-testid="stProgress"] > div > div:has(progress) {
        background-color: var(--pc-pink-100) !important;
    }
    [data-testid="stProgress"] > div > div > div,
    [data-testid="stProgress"] div[role="progressbar"],
    [data-testid="stProgress"] progress::-webkit-progress-value {
        background-color: var(--pc-pink-500) !important;
    }
    [data-testid="stProgress"] progress::-moz-progress-bar {
        background-color: var(--pc-pink-500) !important;
    }
    [data-testid="stProgress"] p,
    [data-testid="stProgress"] span {
        color: var(--pc-muted) !important;
    }
    [data-testid="stProgress"] [data-testid="stMarkdownContainer"],
    [data-testid="stProgress"] [data-testid="stMarkdownContainer"] *,
    [data-testid="stProgress"] > div > div:has([data-testid="stMarkdownContainer"]) {
        background: transparent !important;
        background-color: transparent !important;
    }

    /* Acciones principales — fucsia */
    button[kind="primary"],
    button[kind="primaryFormSubmit"],
    [data-testid="stBaseButton-primary"],
    .st-key-btn_ejecutar_consolidacion button,
    .st-key-btn_iniciar_nuevo_reporte button {
        background: var(--pc-pink-500) !important;
        background-color: var(--pc-pink-500) !important;
        color: #ffffff !important;
        border: 1px solid var(--pc-pink-500) !important;
        font-weight: 600 !important;
        border-radius: 8px !important;
        box-shadow: 0 4px 12px rgba(224, 33, 138, 0.32) !important;
    }
    button[kind="primary"]:hover,
    button[kind="primaryFormSubmit"]:hover,
    [data-testid="stBaseButton-primary"]:hover,
    .st-key-btn_ejecutar_consolidacion button:hover,
    .st-key-btn_iniciar_nuevo_reporte button:hover {
        background: var(--pc-pink-600) !important;
        background-color: var(--pc-pink-600) !important;
        color: #ffffff !important;
        border-color: var(--pc-pink-600) !important;
    }
    button[kind="primary"]:active,
    button[kind="primaryFormSubmit"]:active,
    [data-testid="stBaseButton-primary"]:active,
    .st-key-btn_ejecutar_consolidacion button:active,
    .st-key-btn_iniciar_nuevo_reporte button:active {
        background: var(--pc-pink-700) !important;
        background-color: var(--pc-pink-700) !important;
    }
    button[kind="primary"] p,
    button[kind="primary"] span,
    button[kind="primaryFormSubmit"] p,
    button[kind="primaryFormSubmit"] span,
    [data-testid="stBaseButton-primary"] p,
    [data-testid="stBaseButton-primary"] span,
    .st-key-btn_ejecutar_consolidacion button p,
    .st-key-btn_ejecutar_consolidacion button span,
    .st-key-btn_iniciar_nuevo_reporte button p,
    .st-key-btn_iniciar_nuevo_reporte button span {
        color: #ffffff !important;
    }
    /* Quitar de cola — icono basura rojo centrado */
    div[class*="st-key-quitar_cola_"] button {
        position: relative !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        background: #ffffff !important;
        background-color: #ffffff !important;
        border: 1px solid #fecaca !important;
        box-shadow: none !important;
        color: #b91c1c !important;
        font-weight: 600 !important;
        border-radius: 8px !important;
        min-height: 2.35rem !important;
        min-width: 2.35rem !important;
    }
    div[class*="st-key-quitar_cola_"] button:hover {
        background: #fef2f2 !important;
        background-color: #fef2f2 !important;
        border-color: #fecaca !important;
    }
    div[class*="st-key-quitar_cola_"] button > div {
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        width: 100% !important;
        margin: 0 !important;
        padding: 0 !important;
        min-height: 0 !important;
    }
    div[class*="st-key-quitar_cola_"] button p,
    div[class*="st-key-quitar_cola_"] button [data-testid="stMarkdownContainer"] {
        display: block !important;
        width: auto !important;
        height: auto !important;
        margin: 0 !important;
        padding: 0 !important;
        overflow: visible !important;
        line-height: normal !important;
        white-space: nowrap !important;
    }
    div[class*="st-key-quitar_cola_"] button::before {
        content: none;
        position: absolute;
        left: 50%;
        top: 50%;
        transform: translate(-50%, -50%);
        display: block;
        width: 1.35rem;
        height: 1.35rem;
        margin: 0;
        background-repeat: no-repeat;
        background-position: center;
        background-size: contain;
        background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='%23dc2626'%3E%3Cpath d='M9 3h6a1 1 0 0 1 1 1v1h4a1 1 0 1 1 0 2h-1v13a3 3 0 0 1-3 3H8a3 3 0 0 1-3-3V7H4a1 1 0 1 1 0-2h4V4a1 1 0 0 1 1-1zm1 2h4V4h-4v1zm-2 3v12a1 1 0 0 0 1 1h8a1 1 0 0 0 1-1V8H8zm3 3a1 1 0 1 1 2 0v7a1 1 0 1 1-2 0v-7zm4 0a1 1 0 1 1 2 0v7a1 1 0 1 1-2 0v-7z'/%3E%3C/svg%3E");
        pointer-events: none;
    }
    /* Descargas — fucsia UI, no afecta colores del Excel */
    .st-key-btn_descargar_excel button,
    .st-key-dl_contratos_todas button {
        background: var(--pc-pink-600) !important;
        background-color: var(--pc-pink-600) !important;
        color: #ffffff !important;
        border: 1px solid var(--pc-pink-600) !important;
        font-weight: 600 !important;
        border-radius: 8px !important;
    }
    .st-key-btn_descargar_excel button:hover,
    .st-key-dl_contratos_todas button:hover {
        background: var(--pc-pink-700) !important;
        background-color: var(--pc-pink-700) !important;
        border-color: var(--pc-pink-700) !important;
        color: #ffffff !important;
    }
    .st-key-btn_descargar_excel button p,
    .st-key-btn_descargar_excel button span,
    .st-key-dl_contratos_todas button p,
    .st-key-dl_contratos_todas button span {
        color: #ffffff !important;
    }
    .st-key-btn_descargar_excel button:disabled,
    .st-key-dl_contratos_todas button:disabled {
        background: var(--pc-pink-200) !important;
        background-color: var(--pc-pink-200) !important;
        border-color: var(--pc-pink-200) !important;
        color: var(--pc-pink-700) !important;
        opacity: 0.65 !important;
        box-shadow: none !important;
        cursor: not-allowed !important;
    }

    /* Uploaders: texto y caja mas limpios que el control nativo de Streamlit */
    [data-testid="stFileUploaderDropzone"] {
        background: var(--pc-pink-25) !important;
        border: 1px solid var(--pc-pink-200) !important;
        border-radius: 8px !important;
        min-height: 4.3rem !important;
        padding: 0.8rem 0.9rem !important;
    }
    [data-testid="stFileUploaderDropzone"]:hover {
        border-color: var(--pc-pink-300) !important;
        background: #fff7fb !important;
    }
    [data-testid="stFileUploader"] [data-testid="stBaseButton-secondary"] {
        border-radius: 8px !important;
        border-color: var(--pc-pink-200) !important;
        min-height: 2.45rem !important;
        padding: 0.45rem 0.8rem !important;
    }
    [data-testid="stFileUploader"] [data-testid="stBaseButton-secondary"] p {
        display: none !important;
    }
    [data-testid="stFileUploader"] [data-testid="stBaseButton-secondary"]::after {
        content: "Seleccionar archivo";
        color: var(--pc-ink);
        font-weight: 500;
        font-size: 0.88rem;
        white-space: nowrap;
    }
    [data-testid="stFileUploaderDropzoneInstructions"] {
        visibility: hidden !important;
        min-width: 13rem !important;
        position: relative !important;
    }
    [data-testid="stFileUploaderDropzoneInstructions"]::after {
        content: "Excel .xlsx o .xls, max. 200 MB";
        visibility: visible !important;
        position: absolute !important;
        inset: 0 auto auto 0 !important;
        color: var(--pc-muted) !important;
        font-size: 0.82rem !important;
        line-height: 1.25 !important;
        white-space: normal !important;
    }
    .queue-index {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 1.35rem;
        height: 1.35rem;
        border-radius: 999px;
        background: var(--pc-pink-500);
        color: #ffffff;
        font-size: 0.68rem;
        font-weight: 700;
        margin-top: 0.1rem;
    }
    .queue-localidad {
        font-size: 0.9rem;
        font-weight: 700;
        color: var(--pc-ink);
        margin: 0 0 0.15rem 0;
    }
    .queue-file-label {
        font-size: 0.64rem;
        font-weight: 700;
        letter-spacing: 0.04em;
        text-transform: uppercase;
        color: var(--pc-pink-700);
        margin: 0;
    }
    .queue-file-name {
        font-size: 0.76rem;
        color: var(--pc-muted);
        overflow-wrap: anywhere;
        margin: 0;
        line-height: 1.25;
    }
    .queue-compact-row {
        border-top: 1px solid var(--pc-pink-200);
        padding: 0.5rem 0;
    }
    .queue-compact-row:first-child {
        border-top: 0;
    }
    [data-testid="stAlertContainer"]:has([data-testid="stAlertContentInfo"]) {
        background: var(--pc-pink-50) !important;
        border: 1px solid var(--pc-pink-200) !important;
        color: var(--pc-ink) !important;
    }
    [data-testid="stAlertContentInfo"] [data-testid="stMarkdownContainer"],
    [data-testid="stAlertContentInfo"] [data-testid="stMarkdownContainer"] * {
        color: var(--pc-ink) !important;
    }
    [data-testid="stAlertContainer"]:has([data-testid="stAlertContentWarning"]) {
        background: #fff8db !important;
        border: 1px solid #facc15 !important;
    }
    [data-testid="stAlertContentWarning"] [data-testid="stMarkdownContainer"],
    [data-testid="stAlertContentWarning"] [data-testid="stMarkdownContainer"] * {
        color: #7a5600 !important;
    }
    [data-testid="stAlertContainer"]:has([data-testid="stAlertContentSuccess"]) {
        background: #ecfdf3 !important;
        border: 1px solid #86efac !important;
    }
    [data-testid="stAlertContentSuccess"] [data-testid="stMarkdownContainer"],
    [data-testid="stAlertContentSuccess"] [data-testid="stMarkdownContainer"] * {
        color: #166534 !important;
    }
    .desempate-mini-alert,
    .desempate-callout {
        border: 1px solid #e2e8f0;
        border-left: 5px solid var(--pc-pink-500);
        border-radius: 8px;
        background: #ffffff;
        box-shadow: 0 8px 22px rgba(15, 23, 42, 0.06);
    }
    .desempate-mini-alert {
        padding: 0.85rem 1rem;
        margin: 0.8rem 0 0.4rem;
        color: var(--pc-ink);
        font-weight: 600;
    }
    .desempate-callout {
        padding: 1rem 1.05rem;
        margin: 0.65rem 0 1rem;
    }
    .desempate-kicker {
        color: #64748b;
        font-size: 0.72rem;
        font-weight: 800;
        letter-spacing: 0.06em;
        text-transform: uppercase;
        margin-bottom: 0.25rem;
    }
    .desempate-callout-title,
    .desempate-case-title,
    .desempate-options-title {
        color: #0f172a;
        font-weight: 800;
        letter-spacing: 0;
    }
    .desempate-callout-title {
        font-size: 1.25rem;
        margin-bottom: 0.25rem;
    }
    .desempate-callout p,
    .desempate-case-header p,
    .desempate-note {
        color: #475569;
        margin: 0;
        line-height: 1.42;
    }
    .desempate-progress-text {
        margin: 0.6rem 0 0.35rem;
        color: #475569;
        font-size: 0.92rem;
        background: transparent !important;
    }
    .desempate-progress-text strong {
        color: #0f172a;
    }
    .desempate-case-header {
        margin: 0.65rem 0 0.55rem;
        padding: 0.7rem 0.85rem;
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        background: #ffffff;
    }
    .desempate-case-title {
        font-size: 1.02rem;
        margin-bottom: 0.12rem;
    }
    .desempate-saldo {
        color: #0f172a;
        font-size: 1.05rem;
        font-weight: 800;
    }
    .desempate-detail-grid {
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: 0.45rem;
        margin: 0.5rem 0 0.6rem;
    }
    .desempate-detail-item {
        border: 1px solid #e2e8f0;
        border-radius: 6px;
        padding: 0.52rem 0.6rem;
        background: #ffffff;
        min-height: 3rem;
    }
    .desempate-detail-item span {
        display: block;
        color: #64748b;
        font-size: 0.66rem;
        font-weight: 700;
        letter-spacing: 0.04em;
        text-transform: uppercase;
        margin-bottom: 0.16rem;
    }
    .desempate-detail-item strong {
        color: #334155;
        font-size: 0.86rem;
        line-height: 1.18;
        overflow-wrap: anywhere;
    }
    .desempate-note {
        border-left: 4px solid #cbd5e1;
        border-radius: 6px;
        background: #f8fafc;
        padding: 0.65rem 0.75rem;
        margin: 0 0 0.85rem;
    }
    .desempate-options-title {
        font-size: 1rem;
        margin: 0.85rem 0 0.35rem;
    }
    [data-testid="stRadio"] label,
    [data-testid="stRadio"] label * {
        color: #334155 !important;
    }
    .result-table-wrap {
        width: 100%;
        margin: 0.45rem 0 0.9rem;
    }
    .result-table {
        width: 100%;
        table-layout: fixed;
        border-collapse: collapse;
        border: 1px solid #e2e8f0;
        border-radius: 6px;
        overflow: hidden;
        font-size: 0.82rem;
    }
    .result-table th,
    .result-table td {
        border-top: 1px solid #e2e8f0;
        border-right: 1px solid #e2e8f0;
        padding: 0.46rem 0.5rem;
        vertical-align: top;
    }
    .result-table th:last-child,
    .result-table td:last-child {
        border-right: 0;
    }
    .result-table th {
        background: #f8fafc;
        color: #334155;
        font-size: 0.74rem;
        font-weight: 700;
        letter-spacing: 0;
        line-height: 1.16;
        overflow-wrap: anywhere;
        text-transform: none;
        white-space: normal;
    }
    .result-table td {
        color: #0f172a;
        line-height: 1.2;
        overflow-wrap: anywhere;
        white-space: normal;
    }
    .result-table td.result-number {
        font-variant-numeric: tabular-nums;
        font-size: 0.9rem;
        font-weight: 600;
        line-height: 1.15;
        overflow-wrap: normal;
        text-align: center;
        white-space: nowrap;
        word-break: keep-all;
    }
    .result-table td.result-empty {
        color: #64748b;
    }
    @media (max-width: 640px) {
        .desempate-detail-grid {
            grid-template-columns: 1fr;
        }
        .result-table {
            font-size: 0.72rem;
        }
        .result-table th,
        .result-table td {
            padding: 0.4rem 0.38rem;
        }
        .result-table td.result-number {
            font-size: 0.8rem;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

SHEET_MATRIZ = "MATRIZ OXP"
MATRIZ_HEADER_FILA = 6  # fila de encabezados en pandas (Excel fila 7)
FILA_INICIO_MATRIZ = 8  # columna A desde fila 8 en hoja MATRIZ OXP
SELECCION_LOCALIDAD = "Seleccione Localidad"
KW_CONTRATOS = "plan de choque"
KW_MATRIZ = "matriz"
PALABRAS_IGNORAR = {"de", "la", "los", "las", "el", "del", "y"}
ARCHIVO_AVANCE_BASE = "Avance plan de choque"
ARCHIVO_RESUMEN_BASE = "Tabla de resumen"
MESES_ES = (
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
)


def init_session_state():
    defaults = {
        "cola_localidades": [],
        "consolidated_df": None,
        "processed": False,
        "file_stats": [],
        "last_processed_at": None,
        "upload_key": 0,
        "pendiente_consolidacion": False,
        "consolidacion_en_curso": False,
        "ejecutar_consolidacion_ahora": False,
        "cola_ejecucion": [],
        "error_ultima_ejecucion": None,
        "errores_ejecucion": [],
        "fecha_analisis": None,
        "cruce_informe": [],
        "cruce_detalle": [],
        "contratos_actualizados": {},
        "cruce_resumen_global": [],
        "titulo_saldo_corte": "",
        "desempate_wizard_idx": 0,
        "desempate_wizard_mapa": {},
        "acceso_autorizado": False,
        "reporte_ejecucion": None,
        "_pc_mostrar_formulario_correccion": False,
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val


def contrasena_acceso_esperada() -> str | None:
    """Contraseña en .streamlit/secrets.toml (local) o Secrets de Streamlit Cloud."""
    try:
        if st.secrets.get("sin_contrasena_acceso") in (True, "true", "1", "yes", "si", "sí"):
            return None
        valor = st.secrets.get("contrasena_acceso")
        if valor is None:
            valor = st.secrets.get("codigo_acceso")
        if valor is None:
            return None
        texto = str(valor).strip()
        return texto if texto else None
    except Exception:
        return None


CLAVE_INPUT_CONTRASENA = "input_contrasena_portada"
CLAVE_MENSAJE_BIENVENIDA = "_mensaje_bienvenida_portada"
CLAVE_ULTIMO_MENSAJE_BIENVENIDA = "_ultimo_mensaje_bienvenida_portada"
MENSAJES_PERSONALES_BIENVENIDA = (
    "Eres luz, mujer maravillosa ✨💖",
    "Gracias por ser una hermana excepcional 💕",
    "Te amo mucho hermanita 💖",
    "Estoy muy orgullosa de ti 🌸",
    "Tu corazón hace todo más bonito 💖",
    "Gracias por ser tú 🌷💖",
    "Qué fortuna tenerte como hermana 💕",
    "Tu alegría también es mi alegría 💖",
)
MENSAJES_MOTIVACION_BIENVENIDA = (
    "La vida te abraza bonito ✨",
    "Tu energía abre caminos bonitos 🌸",
    "La vida te guarda cosas hermosas 💕",
    "Brilla siempre a tu manera ✨",
    "Eres magia bonita en movimiento 💖",
    "Todo lo bueno también te busca 🌷",
    "Tu luz se nota en todo lo que haces ✨",
    "Mereces todo lo bonito del mundo 💕",
)
MENSAJES_DIOS_BIENVENIDA = (
    "Eres la princesa de Dios 👑💖",
    "Dios puso una luz preciosa en ti ✨",
    "Eres una hija amada de Dios 💕",
    "Dios te hizo fuerte y hermosa 🌸",
    "Dios bendice tu camino 💖",
    "Tu corazón es un regalo de Dios ✨",
    "Dios camina contigo 💕",
    "Dios sonríe con tu luz 🌷",
)
MENSAJES_BIENVENIDA_CATEGORIAS = (
    ("personal", MENSAJES_PERSONALES_BIENVENIDA),
    ("motivacion", MENSAJES_MOTIVACION_BIENVENIDA),
    ("dios", MENSAJES_DIOS_BIENVENIDA),
)
MENSAJES_BIENVENIDA = tuple(
    mensaje
    for _, mensajes_categoria in MENSAJES_BIENVENIDA_CATEGORIAS
    for mensaje in mensajes_categoria
)


def mensaje_bienvenida_aleatorio() -> str:
    if not MENSAJES_BIENVENIDA:
        return ""
    ultimo = st.session_state.get(CLAVE_ULTIMO_MENSAJE_BIENVENIDA)
    candidatos = list(range(len(MENSAJES_BIENVENIDA)))
    if len(candidatos) > 1 and ultimo in candidatos:
        candidatos.remove(ultimo)
    elegido = random.choice(candidatos)
    st.session_state[CLAVE_ULTIMO_MENSAJE_BIENVENIDA] = elegido
    return MENSAJES_BIENVENIDA[elegido]


def render_mensaje_bienvenida_pendiente() -> None:
    mensaje = st.session_state.pop(CLAVE_MENSAJE_BIENVENIDA, "")
    if not mensaje:
        return
    components.html(
        f"""
        <script>
        (function () {{
          const mensaje = {json.dumps(mensaje, ensure_ascii=False)};
          let doc = document;
          try {{ doc = window.parent.document; }} catch (err) {{}}

          const previo = doc.getElementById("pc-mensaje-bienvenida");
          if (previo) previo.remove();
          const estiloPrevio = doc.getElementById("pc-mensaje-bienvenida-style");
          if (estiloPrevio) estiloPrevio.remove();

          const estilo = doc.createElement("style");
          estilo.id = "pc-mensaje-bienvenida-style";
          estilo.textContent = `
            #pc-mensaje-bienvenida {{
              position: fixed;
              inset: 0;
              z-index: 2147483647;
              display: flex;
              align-items: center;
              justify-content: center;
              padding: 1rem;
              background: rgba(91, 10, 55, 0.22);
              backdrop-filter: blur(5px);
            }}
            #pc-mensaje-bienvenida .pc-card {{
              width: min(92vw, 430px);
              background: linear-gradient(180deg, #fff9fc 0%, #fff1f8 100%);
              border: 1px solid #f9a8d4;
              border-radius: 8px;
              box-shadow: 0 22px 60px rgba(157, 23, 77, 0.26);
              padding: 1.35rem;
              text-align: center;
              font-family: Inter, system-ui, -apple-system, Segoe UI, sans-serif;
              color: #5b0a37;
            }}
            #pc-mensaje-bienvenida .pc-message {{
              margin: 0 0 1rem;
              font-size: 1.08rem;
              line-height: 1.45;
              font-weight: 700;
            }}
            #pc-mensaje-bienvenida button {{
              border: 1px solid #e0218a;
              border-radius: 8px;
              background: #e0218a;
              color: white;
              min-height: 2.35rem;
              padding: 0.45rem 1rem;
              font-weight: 700;
              cursor: pointer;
              box-shadow: 0 6px 16px rgba(224, 33, 138, 0.28);
            }}
            #pc-mensaje-bienvenida button:hover {{
              background: #c2186a;
              border-color: #c2186a;
            }}
          `;
          doc.head.appendChild(estilo);

          const overlay = doc.createElement("div");
          overlay.id = "pc-mensaje-bienvenida";
          const card = doc.createElement("div");
          card.className = "pc-card";
          const texto = doc.createElement("p");
          texto.className = "pc-message";
          texto.textContent = mensaje;
          const boton = doc.createElement("button");
          boton.type = "button";
          boton.textContent = "💌";
          boton.setAttribute("aria-label", "Cerrar mensaje de amor");

          card.appendChild(texto);
          card.appendChild(boton);
          overlay.appendChild(card);
          doc.body.appendChild(overlay);

          let cerrado = false;
          function cerrar() {{
            if (cerrado) return;
            cerrado = true;
            doc.removeEventListener("keydown", cerrarConTecla, true);
            overlay.remove();
          }}
          function cerrarConTecla(e) {{
            if (e.key !== "Enter" && e.key !== "Escape") return;
            e.preventDefault();
            e.stopPropagation();
            cerrar();
          }}
          boton.addEventListener("click", cerrar);
          overlay.addEventListener("click", function (e) {{
            if (e.target === overlay) cerrar();
          }});
          doc.addEventListener("keydown", cerrarConTecla, true);
          setTimeout(function () {{ boton.focus(); }}, 80);
          setTimeout(cerrar, 10000);
        }})();
        </script>
        """,
        height=0,
    )


def _componente_teclado_portada_acceso(clave_widget: str) -> None:
    """Foco y captura de teclas (components.html suele funcionar mejor que st.html en Cloud)."""
    selector = f".st-key-{clave_widget} input"
    caja = ".st-key-portada_acceso_box"
    components.html(
        f"""
        <script>
        (function () {{
          const selector = "{selector}";
          const caja = "{caja}";

          function documentos() {{
            const docs = [];
            const vistos = new Set();
            function agregar(doc) {{
              if (!doc || vistos.has(doc)) return;
              vistos.add(doc);
              docs.push(doc);
            }}
            agregar(document);
            try {{ agregar(window.parent.document); }} catch (err) {{}}
            try {{
              window.parent.document.querySelectorAll("iframe").forEach(function (f) {{
                try {{ agregar(f.contentDocument); }} catch (err) {{}}
              }});
            }} catch (err) {{}}
            return docs;
          }}

          function buscarInput() {{
            for (const doc of documentos()) {{
              let el = doc.querySelector(selector);
              if (el) return el;
              const box = doc.querySelector(caja);
              if (box) {{
                el = box.querySelector('[data-testid="stTextInput"] input');
                if (el) return el;
              }}
              const form = doc.querySelector('form[data-testid="stForm"]');
              if (form) {{
                el = form.querySelector("input");
                if (el) return el;
              }}
            }}
            return null;
          }}

          function configurar(el) {{
            if (!el || el.dataset.pcAcceso === "1") return;
            el.dataset.pcAcceso = "1";
            el.setAttribute("autofocus", "");
            el.setAttribute("inputmode", "numeric");
            el.setAttribute("autocomplete", "one-time-code");
          }}

          function enfocar() {{
            const el = buscarInput();
            if (!el) return false;
            configurar(el);
            try {{
              el.focus({{ preventScroll: true }});
              el.click();
            }} catch (err) {{}}
            return true;
          }}

          function insertarTexto(el, ch) {{
            const proto = window.HTMLInputElement.prototype;
            const desc = Object.getOwnPropertyDescriptor(proto, "value");
            const next = el.value + ch;
            if (desc && desc.set) desc.set.call(el, next);
            else el.value = next;
            try {{
              el.dispatchEvent(new InputEvent("input", {{
                bubbles: true,
                inputType: "insertText",
                data: ch,
              }}));
            }} catch (err) {{
              el.dispatchEvent(new Event("input", {{ bubbles: true }}));
            }}
          }}

          function activoEsOtroInput() {{
            const el = buscarInput();
            for (const doc of documentos()) {{
              const ae = doc.activeElement;
              if (!ae) continue;
              if (ae === el) return false;
              const tag = (ae.tagName || "").toUpperCase();
              if (tag === "INPUT" || tag === "TEXTAREA" || tag === "SELECT") return true;
            }}
            return false;
          }}

          function manejarTecla(e) {{
            if (activoEsOtroInput()) return;
            const el = buscarInput();
            if (!el) return;
            if (e.ctrlKey || e.metaKey || e.altKey) return;
            if (e.key === "Tab" || e.key === "Escape" || e.key.startsWith("Arrow")) return;

            if (e.key === "Enter") {{
              for (const doc of documentos()) {{
                if (doc.activeElement === el) return;
              }}
              e.preventDefault();
              e.stopPropagation();
              enfocar();
              const form = el.closest("form");
              const btn = form && (
                form.querySelector('button[kind="primaryFormSubmit"]') ||
                form.querySelector('button[type="submit"]') ||
                form.querySelector("button")
              );
              if (btn) btn.click();
              return;
            }}

            if (e.key.length !== 1) return;
            e.preventDefault();
            e.stopPropagation();
            enfocar();
            insertarTexto(el, e.key);
          }}

          function vincular(doc) {{
            if (!doc || doc.documentElement.dataset.pcAccesoTeclas === "1") return;
            doc.documentElement.dataset.pcAccesoTeclas = "1";
            doc.addEventListener("keydown", manejarTecla, true);
          }}

          function iniciar() {{
            documentos().forEach(vincular);
            let intentos = 0;
            const timer = setInterval(function () {{
              enfocar();
              if (++intentos > 40) clearInterval(timer);
            }}, 50);
            try {{
              const obs = new MutationObserver(enfocar);
              obs.observe(window.parent.document.body, {{
                childList: true,
                subtree: true,
              }});
            }} catch (err) {{}}
          }}

          iniciar();
        }})();
        </script>
        """,
        height=0,
    )


def render_portada_acceso() -> None:
    """Pantalla de ingreso; detiene la app hasta contraseña correcta."""
    contrasena_ok = contrasena_acceso_esperada()
    st.markdown('<h1 class="app-title">Plan de Choque</h1>', unsafe_allow_html=True)
    st.markdown(
        '<p class="app-subtitle">Ingrese la contraseña para continuar</p>',
        unsafe_allow_html=True,
    )

    with st.container(border=True, key="portada_acceso_box"):
        with st.form(
            "form_contrasena_acceso",
            clear_on_submit=False,
            enter_to_submit=True,
        ):
            ingresado = st.text_input(
                "Contraseña",
                type="password",
                placeholder="Contraseña",
                key=CLAVE_INPUT_CONTRASENA,
                label_visibility="collapsed",
                autocomplete="one-time-code",
            )
            enviado = st.form_submit_button(
                "Entrar",
                type="primary",
                use_container_width=True,
            )
        _componente_teclado_portada_acceso(CLAVE_INPUT_CONTRASENA)

    if enviado:
        texto = str(st.session_state.get(CLAVE_INPUT_CONTRASENA, ingresado)).strip()
        if texto == contrasena_ok:
            st.session_state.acceso_autorizado = True
            st.session_state[CLAVE_MENSAJE_BIENVENIDA] = mensaje_bienvenida_aleatorio()
            st.rerun()
        st.error("Contraseña incorrecta.")

    st.stop()


init_session_state()


def mes_en_espanol(fecha: datetime | date) -> str:
    return MESES_ES[fecha.month - 1]


def mes_capitalizado(fecha: datetime | date) -> str:
    """Mes con primera letra en mayúscula: Mayo, Junio, …"""
    mes = mes_en_espanol(fecha)
    return mes[:1].upper() + mes[1:] if mes else mes


def formato_numero_metrica(valor: float) -> str:
    """Número compacto para tarjetas (sin salto de línea)."""
    n = int(round(valor))
    return f"{n:,}".replace(",", ".")


def formato_valor_simple(valor) -> str:
    """Texto corto para identificadores: evita mostrar 593.0 cuando es 593."""
    if valor is None:
        return "—"
    try:
        if pd.isna(valor):
            return "—"
    except TypeError:
        pass
    if isinstance(valor, float) and valor.is_integer():
        return formato_numero_metrica(valor)
    return str(valor).strip() or "—"


def formato_moneda_contable(valor) -> str:
    """Moneda colombiana para tablas de resumen."""
    try:
        n = float(valor or 0)
    except (TypeError, ValueError):
        return "$ 0"
    signo = "-" if n < 0 else ""
    entero = f"{abs(n):,.0f}".replace(",", ".")
    return f"{signo}$ {entero}"


def _nombre_columna_monetaria_ui(columna) -> bool:
    texto = unicodedata.normalize("NFD", str(columna).lower())
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    columnas_conteo = {
        "por saldo final",
        "por contrato",
        "matriz en cero",
        "seleccion aplicada",
        "total",
    }
    if texto in columnas_conteo or texto.startswith("por "):
        return False
    return any(palabra in texto for palabra in ("saldo", "saldos", "apropiacion", "cxp"))


def formato_moneda_ui(valor) -> str:
    if valor is None:
        return "—"
    try:
        if pd.isna(valor):
            return "—"
    except TypeError:
        pass
    if isinstance(valor, str):
        texto = valor.strip()
        if not texto or texto == "—":
            return "—"
        if texto.startswith("$") or texto.startswith("-$") or texto.startswith("- $"):
            return texto
    try:
        return formato_moneda_contable(valor)
    except (TypeError, ValueError):
        return str(valor).strip() or "—"


def dataframe_con_monedas_ui(df: pd.DataFrame) -> pd.DataFrame:
    """Formatea como pesos las columnas visuales de saldos y apropiación."""
    if df is None or df.empty:
        return df
    salida = df.copy()
    for columna in salida.columns:
        if _nombre_columna_monetaria_ui(columna):
            salida[columna] = salida[columna].map(formato_moneda_ui)
    return salida


def _valor_celda_resultado(valor) -> str:
    if valor is None:
        return "—"
    try:
        if pd.isna(valor):
            return "—"
    except TypeError:
        pass
    if isinstance(valor, float) and valor.is_integer():
        return formato_numero_metrica(valor)
    return str(valor).strip() or "—"


def _celda_resultado_es_numero(texto: str) -> bool:
    limpio = texto.strip()
    if limpio in ("", "—"):
        return False
    if limpio.startswith("$") or limpio.startswith("-$") or limpio.startswith("- $"):
        return True
    if "/" in limpio and re.fullmatch(r"\d+\s*/\s*\d+", limpio):
        return True
    return bool(re.fullmatch(r"-?\d+(?:[.,]\d+)*", limpio))


def render_tabla_resultados(df: pd.DataFrame) -> None:
    """Tabla de lectura sin scroll horizontal: texto envuelve, números no."""
    if df is None or df.empty:
        return
    vista = dataframe_con_monedas_ui(df)
    cols = list(vista.columns)
    thead = "".join(f"<th>{escape(str(col))}</th>" for col in cols)
    filas = []
    for _, row in vista.iterrows():
        celdas = []
        for col in cols:
            texto = _valor_celda_resultado(row[col])
            clases = []
            if texto == "—":
                clases.append("result-empty")
            if _celda_resultado_es_numero(texto):
                clases.append("result-number")
            clase_attr = f' class="{" ".join(clases)}"' if clases else ""
            celdas.append(f"<td{clase_attr}>{escape(texto)}</td>")
        filas.append(f"<tr>{''.join(celdas)}</tr>")
    st.markdown(
        '<div class="result-table-wrap"><table class="result-table">'
        f"<thead><tr>{thead}</tr></thead><tbody>{''.join(filas)}</tbody>"
        "</table></div>",
        unsafe_allow_html=True,
    )


def render_scroll_resultados_si_pendiente() -> None:
    if not st.session_state.pop(_CLAVE_SCROLL_RESULTADOS, False):
        return
    components.html(
        """
        <script>
        setTimeout(() => {
          const doc = window.parent.document;
          const el = doc.getElementById('resultado-consolidado-anchor');
          if (el) {
            el.scrollIntoView({ behavior: 'smooth', block: 'start' });
          }
        }, 120);
        </script>
        """,
        height=0,
    )


def formato_fecha_colombia(fecha: datetime | date, con_hora: bool = False) -> str:
    """Fecha legible en Colombia: día/mes/año."""
    if con_hora and isinstance(fecha, datetime):
        return fecha.strftime("%d/%m/%Y %H:%M")
    return fecha.strftime("%d/%m/%Y")


def parsear_fecha_flexible(
    valor,
    preferir_dia_primero: bool = True,
) -> datetime | None:
    """
    Interpreta fechas en texto o Excel.
    Por defecto asume formato colombiano (día/mes/año). Si día o mes > 12, infiere el orden.
    Ej.: 28/05/2026 → 28 may 2026 | 05/28/2026 (ambiguo) → con preferir_dia_primero falla 5/28;
    sin ambigüedad: 13/05/2026 siempre es 13 mayo.
    """
    if valor is None:
        return None
    if isinstance(valor, str) and not valor.strip():
        return None
    try:
        if pd.isna(valor):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime(valor.year, valor.month, valor.day)
    if isinstance(valor, pd.Timestamp):
        return valor.to_pydatetime()

    texto = str(valor).strip().split()[0]
    if not texto or texto.lower() in {"nat", "none", "nan"}:
        return None

    match = re.match(
        r"^(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})$",
        texto,
    )
    if match:
        p1, p2, anio = int(match.group(1)), int(match.group(2)), int(match.group(3))
        if anio < 100:
            anio += 2000 if anio < 50 else 1900
        if p1 > 12:
            dia, mes = p1, p2
        elif p2 > 12:
            mes, dia = p1, p2
        elif preferir_dia_primero:
            dia, mes = p1, p2
        else:
            mes, dia = p1, p2
        try:
            return datetime(anio, mes, dia)
        except ValueError:
            return None

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(texto, fmt)
        except ValueError:
            continue

    try:
        parsed = pd.to_datetime(valor, dayfirst=preferir_dia_primero, errors="coerce")
        if pd.notna(parsed):
            return parsed.to_pydatetime()
    except Exception:
        pass
    return None


def fecha_referencia_analisis() -> datetime:
    """Fecha del análisis (al consolidar); si no hay, usa la fecha actual."""
    guardada = st.session_state.get("fecha_analisis")
    if isinstance(guardada, datetime):
        return guardada
    if isinstance(guardada, str):
        parsed = parsear_fecha_flexible(guardada)
        if parsed:
            return parsed
    return datetime.now()


def sanitizar_nombre_archivo(nombre: str) -> str:
    """Quita caracteres no válidos en Windows."""
    for char in '<>:"/\\|?*':
        nombre = nombre.replace(char, "")
    return nombre.strip() or "salida.xlsx"


def localidades_en_nombre_archivo(localidades: list[str]) -> str:
    """Ej.: (Usaquén,Kennedy) — nombres originales, separados por coma."""
    limpias = []
    for loc in localidades:
        nombre = str(loc).strip()
        for char in '<>:"/\\|?*':
            nombre = nombre.replace(char, "")
        if nombre:
            limpias.append(nombre)
    if not limpias:
        return ""
    return f"({','.join(limpias)})"


def nombre_archivo_salida(
    base: str,
    fecha: datetime | date | None = None,
    localidades: list[str] | None = None,
) -> str:
    """
    Ej.: Avance plan de choque Mayo (Usaquén,Kennedy).xlsx
    Mes junto al nombre base; localidades con su nombre original entre paréntesis.
    """
    f = fecha or fecha_referencia_analisis()
    mes = mes_capitalizado(f)
    nombre = f"{base} {mes}"
    sufijo_loc = localidades_en_nombre_archivo(localidades or [])
    if sufijo_loc:
        nombre += f" {sufijo_loc}"
    nombre += ".xlsx"
    return sanitizar_nombre_archivo(nombre)


def _stem_descarga_contratos(stem: str, mes: str) -> str:
    """
    Si el archivo ya trae «… - Mayo», conserva lo anterior al último « - »
    y sustituye el tramo del mes (p. ej. «… - Junio»).
    Si no hay « - », añade « - {mes}» al final.
    """
    limpio = stem.replace("—", "-").replace("–", "-").strip()
    sep = " - "
    if sep in limpio:
        base, _viejo_mes = limpio.rsplit(sep, 1)
        return f"{base.strip()}{sep}{mes}"
    return f"{limpio} - {mes}"


def nombre_descarga_contratos_actualizado(
    localidad: str,
    nombre_original: str,
    fecha: datetime | date | None = None,
) -> str:
    """
    Nombre del Excel de Contratos actualizado por localidad.
    Ej.: «Contratos plan de choque Suba - Mayo» → «… Suba - Junio» al consolidar junio.
    """
    f = fecha or fecha_referencia_analisis()
    mes = mes_capitalizado(f)
    if nombre_original and str(nombre_original).strip():
        stem = Path(nombre_original).stem
    else:
        stem = f"Contratos plan de choque {localidad}"
    nombre = _stem_descarga_contratos(stem, mes) + ".xlsx"
    return sanitizar_nombre_archivo(nombre)


def _directorio_archivos_sesion() -> Path:
    """Guarda Excels en disco (no en session_state) para que F5 no reenvíe megabytes."""
    raiz = Path(tempfile.gettempdir()) / "plan_de_choque"
    raiz.mkdir(parents=True, exist_ok=True)
    sid = st.session_state.get("_pc_sid_archivos")
    if not sid:
        sid = uuid.uuid4().hex
        st.session_state._pc_sid_archivos = sid
    carpeta = raiz / str(sid)
    carpeta.mkdir(parents=True, exist_ok=True)
    return carpeta


def bytes_archivo_cola(entrada: dict) -> bytes:
    """Lee bytes desde disco o desde entrada antigua que aún trae bytes en sesión."""
    if entrada.get("bytes") is not None:
        return entrada["bytes"]
    ruta = entrada.get("path")
    if ruta:
        p = Path(ruta)
        if p.is_file():
            return p.read_bytes()
    return b""


@st.cache_data(show_spinner=False)
def _leer_binario_desde_ruta(ruta: str, modificado: float) -> bytes:
    """Cache en memoria del servidor: F5 no vuelve a leer el ZIP del disco."""
    del modificado
    return Path(ruta).read_bytes()


def bytes_contratos_de_salida(data: dict) -> bytes:
    if data.get("bytes_contratos") is not None:
        return data["bytes_contratos"]
    ruta = data.get("path_contratos")
    if ruta and Path(ruta).is_file():
        return Path(ruta).read_bytes()
    return b""


_CLAVE_RUTA_DETALLE = "_pc_cruce_detalle_ruta"
_CLAVE_RUTA_REPORTE = "_pc_reporte_ejecucion_ruta"
_CLAVE_SNAPSHOT = "_pc_snapshot_consolidacion_ruta"
_CLAVE_MOSTRAR_RESULTADOS = "mostrar_resultados_completos"
_CLAVE_SCROLL_RESULTADOS = "_pc_scroll_resultados"
_CLAVE_RESUMEN_LIGERO = "_pc_resumen_consolidado"
_ARCHIVO_META_DETALLE = "cruce_detalle_meta.pkl"
_PREFIJOS_WIDGET_ARCHIVO = (
    "uploader_contratos_",
    "uploader_matriz_",
    "select_localidad_",
)


def _guardar_contratos_en_disco(localidad: str, data: dict) -> dict:
    """Quita bytes_contratos de session_state (pesado en cada F5)."""
    if data.get("path_contratos") and data.get("bytes_contratos") is None:
        return data
    raw = data.get("bytes_contratos")
    if raw is None:
        return data
    carpeta = _directorio_archivos_sesion()
    base = sanitizar_nombre_archivo(
        data.get("nombre_contratos") or f"contratos_{localidad}.xlsx"
    )
    destino = carpeta / f"salida_{sanitizar_nombre_archivo(localidad)}_{base}"
    destino.write_bytes(raw)
    ligero = {k: v for k, v in data.items() if k != "bytes_contratos"}
    ligero["path_contratos"] = str(destino)
    return ligero


def _archivo_cola_en_disco(entrada: dict, localidad: str, tipo: str) -> dict:
    """Garantiza path en disco; si solo hay bytes en memoria, los persiste."""
    nombre = entrada.get("name") or f"{tipo}_{localidad}.xlsx"
    if entrada.get("path") and Path(entrada["path"]).is_file():
        return {"name": nombre, "path": entrada["path"]}
    datos = entrada.get("bytes")
    if datos:
        carpeta = _directorio_archivos_sesion()
        destino = carpeta / sanitizar_nombre_archivo(f"{tipo}_{localidad}_{nombre}")
        destino.write_bytes(datos)
        return {"name": nombre, "path": str(destino)}
    return {"name": nombre}


def _asegurar_cola_en_disco(cola: list) -> list:
    """Cola lista para procesar: cada Contratos/Matriz con archivo legible en disco."""
    refs = []
    for item in cola:
        loc = item["localidad"]
        refs.append({
            "localidad": loc,
            "contratos": _archivo_cola_en_disco(item.get("contratos") or {}, loc, "contratos"),
            "matriz": _archivo_cola_en_disco(item.get("matriz") or {}, loc, "matriz"),
        })
    return refs


def _mensaje_archivo_cola_no_disponible(loc: str, etiqueta: str, entrada: dict) -> str:
    nombre = entrada.get("name")
    if nombre:
        return (
            f"**{loc}** — {etiqueta} **{nombre}** ya no está disponible. "
            "Elimine esa entrada y cárguela de nuevo."
        )
    return f"**{loc}** — Falta el archivo de {etiqueta}."


def _validar_archivos_accesibles(cola: list) -> list[str]:
    errores = []
    for item in cola:
        loc = item["localidad"]
        for clave, etiqueta in (("contratos", "Contratos"), ("matriz", "Matriz")):
            ent = item.get(clave) or {}
            if not entrada_cola_tiene_archivo(ent):
                errores.append(_mensaje_archivo_cola_no_disponible(loc, etiqueta, ent))
                continue
            if not bytes_archivo_cola(ent):
                errores.append(
                    f"**{loc}** — No se pudo leer {etiqueta} "
                    f"(`{ent.get('name', '')}`). Vuelva a subirlo a la cola."
                )
    return errores


def _es_ruta_meta_detalle(ruta: str) -> bool:
    return Path(ruta).name == _ARCHIVO_META_DETALLE


def _borrar_cruce_detalle_en_disco() -> None:
    ruta = st.session_state.pop(_CLAVE_RUTA_DETALLE, None)
    carpeta = _directorio_archivos_sesion()
    if ruta:
        try:
            Path(ruta).unlink(missing_ok=True)
        except OSError:
            pass
    for f in carpeta.glob("detalle_*.pkl"):
        try:
            f.unlink()
        except OSError:
            pass
    legacy = carpeta / "cruce_detalle.pkl"
    if legacy.is_file():
        try:
            legacy.unlink()
        except OSError:
            pass
    st.session_state.cruce_detalle = []


def _persistir_cruce_detalle(detalle: list) -> None:
    """Detalle en disco por localidad (F5 no carga todo si no hace falta)."""
    carpeta = _directorio_archivos_sesion()
    _borrar_cruce_detalle_en_disco()
    por_loc: dict[str, list] = {}
    for fila in detalle:
        loc = str(fila.get("Localidad") or "_sin_loc")
        por_loc.setdefault(loc, []).append(fila)
    indice: list[tuple[str, str]] = []
    for loc, rows in por_loc.items():
        ruta_loc = carpeta / sanitizar_nombre_archivo(f"detalle_{loc}.pkl")
        with open(ruta_loc, "wb") as f:
            pickle.dump(rows, f, protocol=4)
        indice.append((loc, str(ruta_loc)))
    meta = carpeta / _ARCHIVO_META_DETALLE
    with open(meta, "wb") as f:
        pickle.dump(indice, f, protocol=4)
    st.session_state[_CLAVE_RUTA_DETALLE] = str(meta)
    st.session_state.cruce_detalle = []


@st.cache_data(show_spinner=False)
def _cargar_cruce_detalle_cache(ruta: str, modificado: float) -> list:
    del modificado
    with open(ruta, "rb") as f:
        return pickle.load(f)


@st.cache_data(show_spinner=False)
def _cargar_indice_detalle_cache(ruta_meta: str, modificado: float) -> list[tuple[str, str]]:
    del modificado
    with open(ruta_meta, "rb") as f:
        return pickle.load(f)


@st.cache_data(show_spinner=False)
def _cargar_detalle_completo_desde_meta(ruta_meta: str, modificado: float) -> list:
    del modificado
    out: list = []
    for _loc, ruta in _cargar_indice_detalle_cache(
        ruta_meta, Path(ruta_meta).stat().st_mtime
    ):
        p = Path(ruta)
        if p.is_file():
            out.extend(_cargar_cruce_detalle_cache(str(p), p.stat().st_mtime))
    return out


def obtener_cruce_detalle_localidad(localidad: str) -> list:
    inline = st.session_state.get("cruce_detalle") or []
    if inline:
        return [d for d in inline if d.get("Localidad") == localidad]
    ruta = st.session_state.get(_CLAVE_RUTA_DETALLE)
    if not ruta or not Path(ruta).is_file():
        return []
    p = Path(ruta)
    if _es_ruta_meta_detalle(str(p)):
        for loc, ruta_loc in _cargar_indice_detalle_cache(str(p), p.stat().st_mtime):
            if loc == localidad:
                pl = Path(ruta_loc)
                if pl.is_file():
                    return list(_cargar_cruce_detalle_cache(str(pl), pl.stat().st_mtime))
        return []
    return [
        d
        for d in _cargar_cruce_detalle_cache(str(p), p.stat().st_mtime)
        if d.get("Localidad") == localidad
    ]


def obtener_cruce_detalle() -> list:
    inline = st.session_state.get("cruce_detalle") or []
    if inline:
        return list(inline)
    ruta = st.session_state.get(_CLAVE_RUTA_DETALLE)
    if not ruta or not Path(ruta).is_file():
        return []
    p = Path(ruta)
    if _es_ruta_meta_detalle(str(p)):
        return list(_cargar_detalle_completo_desde_meta(str(p), p.stat().st_mtime))
    return list(_cargar_cruce_detalle_cache(str(p), p.stat().st_mtime))


def informe_requiere_detalle_cruce(informe: list) -> bool:
    return any(_localidad_requiere_detalle_cruce(i) for i in informe)


def establecer_cruce_detalle(detalle: list) -> None:
    if not detalle:
        _borrar_cruce_detalle_en_disco()
        return
    _persistir_cruce_detalle(detalle)


def dataframe_consolidado():
    detalle = obtener_cruce_detalle()
    if not detalle:
        return pd.DataFrame()
    return pd.DataFrame(detalle)


def _purgar_uploaders_obsoletos() -> None:
    """Cada «Añadir a cola» deja Excels en claves viejas del file_uploader (F5 más lento)."""
    uk = int(st.session_state.get("upload_key", 0))
    for key in list(st.session_state.keys()):
        for pref in _PREFIJOS_WIDGET_ARCHIVO:
            if not key.startswith(pref):
                continue
            try:
                if int(key[len(pref) :]) != uk:
                    st.session_state.pop(key, None)
            except ValueError:
                pass


def _aligerar_sesion_archivos_pesados() -> None:
    """Migra sesiones viejas que guardaban ZIP/Excel como bytes (F5 muy lento)."""
    zip_info = st.session_state.get("zip_descarga_contratos")
    if zip_info and zip_info.get("data") and not zip_info.get("path"):
        carpeta = _directorio_archivos_sesion()
        nombre = zip_info.get("nombre") or "contratos.zip"
        ruta = carpeta / sanitizar_nombre_archivo(nombre)
        ruta.write_bytes(zip_info["data"])
        st.session_state.zip_descarga_contratos = {
            "path": str(ruta),
            "nombre": nombre,
            "mime": zip_info.get("mime", "application/zip"),
        }

    act = st.session_state.get("contratos_actualizados") or {}
    if act and any((d or {}).get("bytes_contratos") for d in act.values()):
        st.session_state.contratos_actualizados = {
            loc: _guardar_contratos_en_disco(loc, dict(datos))
            for loc, datos in act.items()
        }

    work = st.session_state.get("consolidacion_work")
    if work and work.get("cola"):
        work["cola"] = _asegurar_cola_en_disco(work["cola"])
        ca = work.get("contratos_actualizados") or {}
        if ca and any((d or {}).get("bytes_contratos") for d in ca.values()):
            for loc, datos in ca.items():
                work["contratos_actualizados"][loc] = _guardar_contratos_en_disco(
                    loc, dict(datos)
                )


def _resumen_ligero_desde_informe(informe: list) -> dict:
    return {
        "n_localidades": len(informe),
        "sin_resolver": sum(i.get("sin_resolver", 0) for i in informe),
        "total_contratos": sum(i.get("total_contratos", 0) for i in informe),
        "total_ok": sum(i.get("contratos_ok", 0) for i in informe),
        "cxp_total": sum(i.get("cxp_total", 0) for i in informe),
        "titulo_mes": st.session_state.get("titulo_saldo_corte", ""),
        "procesado_en": st.session_state.get("last_processed_at", ""),
    }


def _persistir_snapshot_consolidacion() -> None:
    """Informe y stats fuera de session_state (F5 liviano)."""
    informe = st.session_state.get("cruce_informe") or []
    if not informe and not st.session_state.get("processed"):
        return
    snapshot = {
        "informe": informe,
        "file_stats": st.session_state.get("file_stats") or [],
        "cruce_resumen_global": st.session_state.get("cruce_resumen_global") or [],
        "fecha_analisis": st.session_state.get("fecha_analisis"),
        "last_processed_at": st.session_state.get("last_processed_at"),
        "titulo_saldo_corte": st.session_state.get("titulo_saldo_corte", ""),
    }
    ruta = _directorio_archivos_sesion() / "consolidacion_snapshot.pkl"
    with open(ruta, "wb") as f:
        pickle.dump(snapshot, f, protocol=4)
    st.session_state[_CLAVE_SNAPSHOT] = str(ruta)
    st.session_state[_CLAVE_RESUMEN_LIGERO] = _resumen_ligero_desde_informe(informe)
    st.session_state.cruce_informe = []
    st.session_state.file_stats = []
    st.session_state.cruce_resumen_global = []


@st.cache_data(show_spinner=False)
def _cargar_snapshot_consolidacion_cache(ruta: str, modificado: float) -> dict:
    del modificado
    with open(ruta, "rb") as f:
        return pickle.load(f)


def _cargar_snapshot_consolidacion() -> dict:
    if st.session_state.get("cruce_informe"):
        return {
            "informe": st.session_state.get("cruce_informe") or [],
            "file_stats": st.session_state.get("file_stats") or [],
            "cruce_resumen_global": st.session_state.get("cruce_resumen_global") or [],
            "fecha_analisis": st.session_state.get("fecha_analisis"),
            "last_processed_at": st.session_state.get("last_processed_at"),
            "titulo_saldo_corte": st.session_state.get("titulo_saldo_corte", ""),
        }
    ruta = st.session_state.get(_CLAVE_SNAPSHOT)
    if ruta and Path(ruta).is_file():
        p = Path(ruta)
        return _cargar_snapshot_consolidacion_cache(str(p), p.stat().st_mtime)
    return {
        "informe": [],
        "file_stats": [],
        "cruce_resumen_global": [],
        "fecha_analisis": st.session_state.get("fecha_analisis"),
        "last_processed_at": st.session_state.get("last_processed_at"),
        "titulo_saldo_corte": st.session_state.get("titulo_saldo_corte", ""),
    }


def _informe_para_ui() -> list:
    inline = st.session_state.get("cruce_informe") or []
    if inline:
        return list(inline)
    return list(_cargar_snapshot_consolidacion().get("informe") or [])


def _stats_para_ui() -> list:
    inline = st.session_state.get("file_stats") or []
    if inline:
        return list(inline)
    return list(_cargar_snapshot_consolidacion().get("file_stats") or [])


def _resumen_global_para_ui() -> list:
    inline = st.session_state.get("cruce_resumen_global") or []
    if inline:
        return list(inline)
    return list(_cargar_snapshot_consolidacion().get("cruce_resumen_global") or [])


def _borrar_snapshot_consolidacion() -> None:
    ruta = st.session_state.pop(_CLAVE_SNAPSHOT, None)
    if ruta:
        try:
            Path(ruta).unlink(missing_ok=True)
        except OSError:
            pass
    st.session_state.pop(_CLAVE_RESUMEN_LIGERO, None)


def _externalizar_reporte_en_sesion() -> None:
    payload = st.session_state.get("reporte_ejecucion")
    if not payload or not payload.get("tabla"):
        return
    ruta = _directorio_archivos_sesion() / "reporte_ejecucion.pkl"
    with open(ruta, "wb") as f:
        pickle.dump(payload, f, protocol=4)
    st.session_state[_CLAVE_RUTA_REPORTE] = str(ruta)
    st.session_state.pop("reporte_ejecucion", None)


def _compactar_sesion_vista_completa() -> None:
    """Persistencia liviana sin cerrar el panel (rerun por descarga, desempate, etc.)."""
    if st.session_state.get("cruce_detalle"):
        _persistir_cruce_detalle(list(st.session_state.cruce_detalle))
    _aligerar_sesion_archivos_pesados()


def _compactar_sesion_para_f5() -> None:
    """Reduce lo que Streamlit serializa en cada recarga (F5) — vista resumen."""
    if st.session_state.get("cruce_detalle"):
        _persistir_cruce_detalle(list(st.session_state.cruce_detalle))
    if st.session_state.get("processed") and st.session_state.get("cruce_informe"):
        _persistir_snapshot_consolidacion()
    st.session_state.pop("consolidated_df", None)
    st.session_state[_CLAVE_MOSTRAR_RESULTADOS] = False
    _externalizar_reporte_en_sesion()
    _purgar_uploaders_obsoletos()
    _aligerar_sesion_archivos_pesados()


def entrada_cola_tiene_archivo(entrada: dict | None) -> bool:
    if not entrada:
        return False
    if entrada.get("bytes"):
        return True
    ruta = entrada.get("path")
    return bool(ruta and Path(ruta).is_file())


def _regenerar_zip_descarga_contratos() -> None:
    """Genera el ZIP en disco; en sesión solo ruta + nombre (F5 liviano)."""
    contratos_act = st.session_state.get("contratos_actualizados") or {}
    if not contratos_act:
        st.session_state.pop("zip_descarga_contratos", None)
        return
    fecha_dl = st.session_state.get("fecha_analisis") or fecha_referencia_analisis()
    try:
        datos, nombre, mime = empaquetar_descarga_contratos(contratos_act, fecha_dl)
    except ValueError:
        st.session_state.pop("zip_descarga_contratos", None)
        return
    carpeta = _directorio_archivos_sesion()
    ruta = carpeta / sanitizar_nombre_archivo(nombre)
    ruta.write_bytes(datos)
    st.session_state.zip_descarga_contratos = {
        "path": str(ruta),
        "nombre": nombre,
        "mime": mime,
    }


def _vaciar_cola_tras_consolidar() -> None:
    """Libera memoria y acelera recargas; los resultados ya quedaron en la consolidación."""
    for item in st.session_state.get("cola_localidades", []):
        _borrar_archivo_cola_en_disco(item.get("contratos"))
        _borrar_archivo_cola_en_disco(item.get("matriz"))
    st.session_state.cola_localidades = []
    st.session_state.cola_ejecucion = []
    st.session_state.upload_key = int(st.session_state.get("upload_key", 0)) + 1


def empaquetar_descarga_contratos(
    contratos_actualizados: dict,
    fecha: datetime | date | None = None,
) -> tuple[bytes, str, str]:
    """
    Siempre ZIP: evita archivos .xlsx corruptos al descargar directo en Mac/Safari.
    Cada Excel dentro conserva su nombre original + «- {Mes}».
    """
    f = fecha or fecha_referencia_analisis()
    items = sorted(contratos_actualizados.items(), key=lambda x: x[0])
    if not items:
        raise ValueError("No hay contratos actualizados para descargar.")

    localidades = [loc for loc, _ in items]
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for loc, data in items:
            nombre = nombre_descarga_contratos_actualizado(
                loc, data.get("nombre_contratos", ""), f
            )
            zf.writestr(nombre, bytes_contratos_de_salida(data))
    buf.seek(0)
    mes = mes_capitalizado(f)
    zip_nombre = sanitizar_nombre_archivo(
        f"Contratos plan de choque actualizados {mes} "
        f"{localidades_en_nombre_archivo(localidades)}.zip"
    )
    return buf.getvalue(), zip_nombre, "application/zip"


# Etiquetas fijas (mismas que METODOS_LABEL en cxp_cruce); no importar pandas aquí al arranque.
METODOS_SIN_RESOLVER = frozenset({"Sin resolver", "Sin fila en matriz"})
METODOS_AMIGABLES = {
    "Match exacto (4 campos)": "Cruce directo",
    "Fallback por Saldo Final": "Coincidencia por saldo final",
    "Fallback: todos cero en matriz": "Matriz con saldo cero",
    "Match por contrato (sin apropiación)": "Cruce por contrato",
    "Sin resolver": "Sin coincidencia automática",
    "Sin fila en matriz": "No encontrado en Matriz",
    "Saldo vacío en matriz": "Saldo vacío en Matriz",
    "Desempate manual": "Confirmado por selección",
}
CODIGOS_CRUCE_COMPLEMENTARIO = (
    "match_saldo_contrato",
    "todos_cero_matriz",
    "k3_unico",
    "desempate_manual",
)
METODOS_CRUCE_COMPLEMENTARIO = frozenset(
    [
        "Fallback por Saldo Final",
        "Fallback: todos cero en matriz",
        "Match por contrato (sin apropiación)",
        "Desempate manual",
        "Cruce complementario por saldo final",
        "Coincidencia por saldo final",
        "Cruce automático",
        "Matriz con saldo cero",
        "Cruce por contrato",
        "Confirmado por selección",
    ]
)


def etiqueta_metodo_amigable(metodo: str) -> str:
    return METODOS_AMIGABLES.get(str(metodo), str(metodo))


def dataframe_metodos_amigable(resumen_metodos: list) -> pd.DataFrame:
    df = pd.DataFrame(resumen_metodos or [])
    if df.empty:
        return df
    if "Método" in df.columns:
        df = df.rename(columns={"Método": "Tipo de cruce"})
    if "Tipo de cruce" in df.columns:
        df["Tipo de cruce"] = df["Tipo de cruce"].map(etiqueta_metodo_amigable)
    return df


def conteo_revision_adicional(conteo: dict) -> int:
    return sum(
        int(conteo.get(k, 0) or 0)
        for k in CODIGOS_CRUCE_COMPLEMENTARIO
    )


def dataframe_cruces_complementarios(informe: list) -> pd.DataFrame:
    filas = []
    for loc in informe:
        conteo = loc.get("conteo") or {}
        total = conteo_revision_adicional(conteo)
        if total <= 0:
            continue
        filas.append({
            "Localidad": loc["localidad"],
            "Total": total,
            "Por saldo final": int(conteo.get("match_saldo_contrato", 0) or 0),
            "Por contrato": int(conteo.get("k3_unico", 0) or 0),
            "Matriz en cero": int(conteo.get("todos_cero_matriz", 0) or 0),
            "Selección aplicada": int(conteo.get("desempate_manual", 0) or 0),
        })
    df = pd.DataFrame(filas)
    if df.empty:
        return df
    columnas_conteo = [
        "Por saldo final",
        "Por contrato",
        "Matriz en cero",
        "Selección aplicada",
    ]
    visibles = ["Localidad", "Total"] + [
        col for col in columnas_conteo if col in df.columns and int(df[col].sum()) > 0
    ]
    return df[visibles]


def filas_cruce_complementario(detalle: list) -> list[dict]:
    return [
        d
        for d in detalle
        if d.get("Método") in METODOS_CRUCE_COMPLEMENTARIO
    ]


def filas_sin_resolver(detalle: list) -> list[dict]:
    return [d for d in detalle if d.get("Método") in METODOS_SIN_RESOLVER]


def detalle_desempate_visible(detalle) -> str:
    """Oculta diagnósticos técnicos que no ayudan a elegir entre opciones."""
    texto = str(detalle or "").strip()
    if not texto:
        return ""
    texto_norm = unicodedata.normalize("NFD", texto.lower())
    texto_norm = "".join(c for c in texto_norm if unicodedata.category(c) != "Mn")
    if "hay saldos > 0 en matriz" in texto_norm and "saldo final" in texto_norm:
        return ""
    return texto


def numero_fila_excel_para_orden(fila: dict) -> int:
    try:
        return int(float(fila.get("Fila Excel Contratos")))
    except (TypeError, ValueError):
        return 10**9


def incidencias_sin_resolver(detalle: list) -> list[dict]:
    """Contratos pendientes: localidad y fila Excel ascendente para revisión lineal."""
    return sorted(
        filas_sin_resolver(detalle),
        key=lambda f: (
            str(f.get("Localidad") or ""),
            numero_fila_excel_para_orden(f),
            str(f.get("No. de Cto") or ""),
            str(f.get("NOMBRE CONTRATISTA") or "").lower(),
        ),
    )


def _reset_estado_desempate_wizard() -> None:
    st.session_state.desempate_wizard_idx = 0
    st.session_state.desempate_wizard_mapa = {}


def dataframe_resumen_localidades(informe: list) -> pd.DataFrame:
    """Vista compacta: una fila por localidad."""
    if not informe:
        return pd.DataFrame()
    return pd.DataFrame(
        [
            {
                "Localidad": loc["localidad"],
                "Contratos cruzados": f"{loc.get('contratos_ok', 0)}/{loc.get('total_contratos', 0)}",
                "Cruce directo": int((loc.get("conteo") or {}).get("k4_exacto", 0) or 0),
                "Cruce por contrato": int((loc.get("conteo") or {}).get("k3_unico", 0) or 0),
                "Cruce por selección": int((loc.get("conteo") or {}).get("desempate_manual", 0) or 0),
                "CXP del mes": formato_moneda_contable(loc.get("cxp_total", 0)),
            }
            for loc in informe
        ]
    )


def _localidad_requiere_detalle_cruce(loc_info: dict) -> bool:
    """Detalle por localidad solo si hay excepciones o avisos."""
    return (
        loc_info.get("sin_resolver", 0) > 0
        or conteo_revision_adicional(loc_info.get("conteo") or {}) > 0
        or bool(loc_info.get("advertencias_suspendidos"))
    )


def mostrar_informe_cruce_consolidado(
    informe: list,
    titulo_mes: str,
    *,
    cargar_tablas_detalle: bool = True,
) -> None:
    """Resumen conciso del cruce; tablas de fallback solo si hace falta."""
    st.markdown(
        '<p class="section-title">Detalle por localidad</p>',
        unsafe_allow_html=True,
    )
    st.caption(f"Resumen por localidad del cruce aplicado en **{titulo_mes}**.")
    df_loc = dataframe_resumen_localidades(informe)
    if len(df_loc):
        render_tabla_resultados(df_loc)

    with st.expander("Cómo se hizo el cruce", expanded=False):
        st.markdown(
            "- Primero se busca cada contrato en la Matriz con los datos completos.\n"
            "- Si la apropiación no coincide, se revisa el saldo final para ubicar la fila correcta.\n"
            "- Si hay varias opciones posibles, la app lo marca como caso sin coincidencia automática.\n"
            "- Si la Matriz no trae saldo para un contrato, la celda queda vacía.\n"
            "- Las hojas sin contratistas reales no se modifican."
        )

    locales_detalle = [loc for loc in informe if _localidad_requiere_detalle_cruce(loc)]
    if not locales_detalle:
        return
    st.markdown("**Cruces complementarios por localidad**")
    for loc_info in locales_detalle:
        loc = loc_info["localidad"]
        sin_loc = loc_info.get("sin_resolver", 0)
        complemento = conteo_revision_adicional(loc_info.get("conteo") or {})
        avisos = loc_info.get("advertencias_suspendidos") or []
        etiqueta = (
            f"{loc} — {sin_loc} caso(s) sin coincidencia automática"
            if sin_loc
            else f"{loc} — {complemento} cruce(s) complementario(s)"
            if complemento
            else loc
        )
        with st.expander(etiqueta, expanded=sin_loc > 0 or complemento > 0):
            if avisos:
                for aviso in avisos:
                    st.warning(aviso)
            accion_col = str(loc_info.get("accion_columna") or "").strip().lower()
            estado_col = "se creó" if accion_col == "creada" else "se actualizó"
            st.caption(f"Cps por depurar: {estado_col} la columna del mes.")
            if loc_info.get("resumen_metodos"):
                render_tabla_resultados(
                    dataframe_metodos_amigable(loc_info["resumen_metodos"])
                )
            detalle_loc = []
            if cargar_tablas_detalle:
                detalle_loc = filas_cruce_complementario(
                    obtener_cruce_detalle_localidad(loc)
                )
            if detalle_loc:
                st.markdown(
                    "**Contratos con cruce complementario**: no entraron por cruce directo, "
                    "pero la app encontró una coincidencia con una regla adicional."
                )
                df_detalle = pd.DataFrame(detalle_loc).copy()
                if "Método" in df_detalle.columns:
                    df_detalle["Tipo de cruce"] = df_detalle["Método"].map(
                        etiqueta_metodo_amigable
                    )
                cols = [
                    "Tipo de cruce",
                    "NOMBRE CONTRATISTA",
                    "No. de Cto",
                    "APROPIACION DISPONIBLE",
                    "SALDO FINAL (Contratos)",
                    f"Saldo Matriz ({titulo_mes})",
                    "Detalle",
                ]
                cols = [c for c in cols if c in df_detalle.columns]
                render_tabla_resultados(df_detalle[cols])


def resumen_sin_resolver_por_localidad(detalle: list) -> pd.DataFrame:
    """Conteo de casos sin coincidencia automática por localidad."""
    conteo: dict[str, int] = {}
    for fila in filas_sin_resolver(detalle):
        loc = fila.get("Localidad") or "—"
        conteo[loc] = conteo.get(loc, 0) + 1
    if not conteo:
        return pd.DataFrame(columns=["Localidad", "Sin coincidencia automática"])
    return pd.DataFrame(
        [
            {"Localidad": loc, "Sin coincidencia automática": n}
            for loc, n in sorted(conteo.items())
        ]
    )


def dataframe_sin_resolver(detalle: list) -> pd.DataFrame:
    """Solo columnas útiles para revisar excepciones."""
    filas = filas_sin_resolver(detalle)
    if not filas:
        return pd.DataFrame()
    df = pd.DataFrame(filas)
    preferidas = [
        "Localidad",
        "NOMBRE CONTRATISTA",
        "No. de Cto",
        "AÑO SUSCRIPCIÓN",
        "APROPIACION DISPONIBLE",
        "SALDO FINAL (Contratos)",
        "Método",
        "Detalle",
    ]
    cols_matriz = [c for c in df.columns if str(c).startswith("Saldo Matriz")]
    cols = [c for c in preferidas if c in df.columns] + cols_matriz
    resto = [c for c in df.columns if c not in cols]
    return df[cols + resto]


def _clave_base_desempate(fila: dict) -> str:
    return cxp_cruce.clave_fila_contrato(
        fila.get("NOMBRE CONTRATISTA"),
        fila.get("No. de Cto"),
        fila.get("AÑO SUSCRIPCIÓN"),
        fila.get("APROPIACION DISPONIBLE"),
    )


def _opcion_matriz_por_saldo(fila: dict, saldo: float) -> str:
    for cand in fila.get("candidatos_matriz") or []:
        try:
            if float(cand.get("saldo") or 0) == float(saldo):
                return f"opción {formato_valor_simple(cand.get('opcion'))}"
        except (TypeError, ValueError):
            continue
    return f"saldo {formato_moneda_contable(float(saldo))}"


def errores_reuso_opcion_matriz(incidencias: list[dict], mapa: dict[str, float]) -> list[str]:
    """Evita asignar la misma fila/opción de Matriz a dos filas del mismo contrato."""
    usos: dict[tuple[str, float], list[dict]] = {}
    for inc in incidencias:
        clave = clave_desde_detalle(inc)
        if clave not in mapa:
            continue
        candidatos = inc.get("candidatos_matriz") or []
        if len(candidatos) <= 1:
            continue
        try:
            saldo = float(mapa[clave])
        except (TypeError, ValueError):
            continue
        usos.setdefault((_clave_base_desempate(inc), saldo), []).append(inc)

    errores: list[str] = []
    for (_base, saldo), filas in usos.items():
        if len(filas) <= 1:
            continue
        primera = filas[0]
        filas_excel = ", ".join(
            formato_valor_simple(f.get("Fila Excel Contratos")) for f in filas
        )
        errores.append(
            "La "
            f"{_opcion_matriz_por_saldo(primera, saldo)} de Matriz ya está asignada "
            f"a más de una fila del mismo contrato ({filas_excel}). "
            "Elija una opción diferente para cada fila."
        )
    return errores


def _detalle_con_filas_excel_contratos(detalle: list, contratos_act: dict) -> list:
    """Completa la fila real del Excel para distinguir contratos repetidos."""
    if not detalle or not contratos_act:
        return list(detalle or [])

    enriquecido = [dict(fila) for fila in detalle]
    pendientes_por_loc: dict[str, list[int]] = {}
    for idx, fila in enumerate(enriquecido):
        if fila.get("Tipo fila", "Contratos") != "Contratos":
            continue
        if fila.get("Método") not in METODOS_SIN_RESOLVER:
            continue
        if fila.get("Fila Excel Contratos") not in (None, "", "—"):
            continue
        loc = fila.get("Localidad")
        if loc in contratos_act:
            pendientes_por_loc.setdefault(loc, []).append(idx)

    for loc, indices in pendientes_por_loc.items():
        try:
            raw = bytes_contratos_de_salida(contratos_act[loc])
            libro = pd.ExcelFile(BytesIO(raw))
            nombre_hoja = resolver_hoja_cruce_cxp(list(libro.sheet_names))
            df_c = pd.read_excel(
                BytesIO(raw),
                sheet_name=nombre_hoja,
                header=cxp_cruce.HEADER_CONTRATOS,
            )
            col_nombre = cxp_cruce._columna(df_c, "NOMBRE CONTRATISTA")
            col_cto = cxp_cruce._columna(df_c, "No. de Cto", "Número Contrato")
            col_anio = cxp_cruce._columna(
                df_c, "AÑO SUSCRIPCIÓN", "ANO SUSCRIPCION", "Año Suscripción"
            )
            col_aprop = cxp_cruce._columna(
                df_c, "APROPIACION DISPONIBLE", "Apropiación", "Apropiacion"
            )
            if not all([col_nombre, col_cto, col_anio, col_aprop]):
                continue
        except Exception:
            continue

        filas_por_clave: dict[str, list[int]] = {}
        for i, (_, row) in enumerate(df_c.iterrows()):
            nombre = row[col_nombre]
            if pd.isna(nombre) or not str(nombre).strip():
                continue
            clave_base = cxp_cruce.clave_fila_contrato(
                nombre,
                row[col_cto],
                row[col_anio],
                row[col_aprop],
            )
            fila_excel = cxp_cruce._fila_inicio_datos_contratos() + i
            filas_por_clave.setdefault(clave_base, []).append(int(fila_excel))

        usados: dict[str, int] = {}
        for idx in indices:
            fila = enriquecido[idx]
            clave_base = cxp_cruce.clave_fila_contrato(
                fila.get("NOMBRE CONTRATISTA"),
                fila.get("No. de Cto"),
                fila.get("AÑO SUSCRIPCIÓN"),
                fila.get("APROPIACION DISPONIBLE"),
            )
            opciones = filas_por_clave.get(clave_base) or []
            usado = usados.get(clave_base, 0)
            if usado < len(opciones):
                fila["Fila Excel Contratos"] = opciones[usado]
                usados[clave_base] = usado + 1

    return enriquecido


def aplicar_mapa_desempate(mapa: dict[str, float]) -> tuple[bool, list[str]]:
    """Aplica saldos elegidos a Contratos y actualiza el estado de la consolidación."""
    detalle = list(obtener_cruce_detalle())
    contratos_act = dict(st.session_state.get("contratos_actualizados", {}))
    detalle = _detalle_con_filas_excel_contratos(detalle, contratos_act)
    fecha = st.session_state.get("fecha_analisis") or fecha_referencia_analisis()
    titulo_mes = titulo_saldo_corte(fecha)
    snap = _cargar_snapshot_consolidacion()
    informe = list(snap.get("informe") or [])

    localidades_con_pendientes = {
        loc
        for loc in contratos_act
        if claves_pendientes_localidad(detalle, loc)
    }
    if not localidades_con_pendientes:
        return False, ["No hay contratos sin coincidencia automática."]

    errores: list[str] = []
    for loc in sorted(localidades_con_pendientes):
        pendientes = claves_pendientes_localidad(detalle, loc)
        faltan = validar_desempate_completo(
            pendientes,
            mapa,
            [f for f in detalle if f.get("Localidad") == loc],
        )
        if faltan:
            errores.extend([f"**{loc}**: {msg}" for msg in faltan])

    if errores:
        return False, errores

    for loc in sorted(localidades_con_pendientes):
        incidencias_loc = [
            f
            for f in detalle
            if f.get("Localidad") == loc and f.get("Método") in METODOS_SIN_RESOLVER
        ]
        errores.extend(
            [f"**{loc}**: {msg}" for msg in errores_reuso_opcion_matriz(incidencias_loc, mapa)]
        )

    if errores:
        return False, errores

    for loc in sorted(localidades_con_pendientes):
        bytes_nuevos, detalle = aplicar_desempate_en_contratos(
            bytes_contratos_de_salida(contratos_act[loc]),
            fecha,
            mapa,
            detalle,
            loc,
        )
        contratos_act[loc]["bytes_contratos"] = bytes_nuevos
        contratos_act[loc] = _guardar_contratos_en_disco(loc, contratos_act[loc])
        stats_loc = recalcular_estadisticas_localidad(
            [f for f in detalle if f.get("Localidad") == loc],
            titulo_mes,
        )
        contratos_act[loc].update(stats_loc)
        for i, info in enumerate(informe):
            if info["localidad"] == loc:
                informe[i] = {**info, **stats_loc}
                break

    establecer_cruce_detalle(detalle)
    st.session_state.contratos_actualizados = contratos_act

    conteo_global: dict[str, int] = {}
    for info in informe:
        _agregar_conteo_global(conteo_global, info.get("conteo", {}))
    resumen_global = [
        {
            "Método": METODOS_LABEL.get(codigo, codigo),
            "Contratos": cantidad,
        }
        for codigo, cantidad in sorted(conteo_global.items(), key=lambda x: -x[1])
        if cantidad > 0
    ]

    file_stats = list(snap.get("file_stats") or _stats_para_ui())
    for s in file_stats:
        if s.get("Archivo") == "Contratos (Cps por depurar)":
            loc = s.get("Localidad")
            info = next((i for i in informe if i["localidad"] == loc), None)
            if info:
                s["CXP (suma mes)"] = info["cxp_total"]

    st.session_state.cruce_informe = informe
    st.session_state.file_stats = file_stats
    st.session_state.cruce_resumen_global = resumen_global
    _persistir_snapshot_consolidacion()

    _reset_estado_desempate_wizard()
    st.session_state.zip_descarga_listo = False
    _regenerar_zip_descarga_contratos()
    st.session_state[_CLAVE_SCROLL_RESULTADOS] = True
    return True, []


def render_asistente_desempate(detalle: list, titulo_mes: str) -> None:
    """Asistente paso a paso: una incidencia, opciones Matriz con radio, aplicar al final."""
    incidencias = incidencias_sin_resolver(detalle)
    if not incidencias:
        return

    mapa: dict[str, float] = dict(st.session_state.get("desempate_wizard_mapa", {}))
    n = len(incidencias)
    idx = int(st.session_state.get("desempate_wizard_idx", 0))
    idx = max(0, min(idx, n - 1))
    st.session_state.desempate_wizard_idx = idx

    claves_todas = {clave_desde_detalle(inc) for inc in incidencias}
    resueltas = len(claves_todas & set(mapa.keys()))
    st.markdown(
        f'<div class="desempate-progress-text"><strong>{resueltas} de {n}</strong> '
        "incidencias con saldo elegido</div>",
        unsafe_allow_html=True,
    )
    st.progress(resueltas / n if n else 0.0)

    inc = incidencias[idx]
    clave = clave_desde_detalle(inc)
    loc = inc.get("Localidad") or "—"

    saldo_asignar = mapa.get(clave)
    texto_saldo = (
        formato_moneda_contable(saldo_asignar) if saldo_asignar is not None else "—"
    )
    st.markdown(
        f"""
        <div class="desempate-case-header">
            <div class="desempate-kicker">Caso pendiente</div>
            <div class="desempate-case-title">Incidencia {idx + 1} de {n} · {escape(str(loc))}</div>
            <p>Saldo a asignar en <strong>{escape(str(titulo_mes))}</strong>:
            <span class="desempate-saldo">{escape(str(texto_saldo))}</span></p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    ap = inc.get("APROPIACION DISPONIBLE")
    sf = inc.get("SALDO FINAL (Contratos)")
    ap_texto = formato_moneda_ui(ap)
    sf_texto = formato_moneda_ui(sf)
    st.markdown(
        f"""
        <div class="desempate-detail-grid">
            <div class="desempate-detail-item">
                <span>Contratista</span>
                <strong>{escape(formato_valor_simple(inc.get('NOMBRE CONTRATISTA')))}</strong>
            </div>
            <div class="desempate-detail-item">
                <span>No. de contrato</span>
                <strong>{escape(formato_valor_simple(inc.get('No. de Cto')))}</strong>
            </div>
            <div class="desempate-detail-item">
                <span>Fila del Excel de Contratos</span>
                <strong>{escape(formato_valor_simple(inc.get('Fila Excel Contratos')))}</strong>
            </div>
            <div class="desempate-detail-item">
                <span>Apropiación en Contratos</span>
                <strong>{escape(ap_texto)}</strong>
            </div>
            <div class="desempate-detail-item">
                <span>Saldo final en Contratos</span>
                <strong>{escape(sf_texto)}</strong>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    detalle_visible = detalle_desempate_visible(inc.get("Detalle"))
    if detalle_visible:
        st.markdown(
            f'<div class="desempate-note">{escape(detalle_visible)}</div>',
            unsafe_allow_html=True,
        )

    candidatos = inc.get("candidatos_matriz") or []
    st.markdown(
        '<div class="desempate-options-title">Opciones en Matriz</div>',
        unsafe_allow_html=True,
    )
    if candidatos:
        labels: list[str] = []
        valores: list[float] = []
        for cand in candidatos:
            labels.append(formato_moneda_contable(cand.get("saldo") or 0))
            valores.append(float(cand["saldo"]))

        default_idx = None
        if clave in mapa:
            for i, val in enumerate(valores):
                if val == mapa[clave]:
                    default_idx = i
                    break

        eleccion = st.radio(
            "Elija la opción de Matriz",
            options=list(range(len(labels))),
            format_func=lambda i, lbls=labels: lbls[i],
            index=default_idx,
            key=f"wiz_radio_{clave}",
        )
        if eleccion is not None:
            mapa[clave] = valores[eleccion]
    else:
        st.warning(
            detalle_visible
            or "No hay filas candidatas en Matriz. Ingrese el saldo para continuar."
        )
        previo = float(mapa[clave]) if clave in mapa else 0.0
        manual = st.number_input(
            f"Saldo para {titulo_mes}",
            min_value=0.0,
            value=previo,
            step=1.0,
            format="%d",
            key=f"wiz_num_{clave}",
        )
        mapa[clave] = float(manual)

    st.session_state.desempate_wizard_mapa = mapa

    nav_prev, nav_next = st.columns(2)
    with nav_prev:
        if st.button("← Anterior", disabled=idx == 0, use_container_width=True):
            st.session_state.desempate_wizard_idx = idx - 1
            st.rerun()
    with nav_next:
        etiqueta_next = "Siguiente →" if idx < n - 1 else "Fin del listado"
        if st.button(etiqueta_next, disabled=idx >= n - 1, use_container_width=True):
            if clave not in mapa:
                st.warning("Elija una opción de Matriz antes de continuar.")
            else:
                st.session_state.desempate_wizard_idx = idx + 1
                st.rerun()

    completo = claves_todas <= set(mapa.keys())
    errores_reuso = errores_reuso_opcion_matriz(incidencias, mapa)
    faltan = n - len(claves_todas & set(mapa.keys()))
    st.markdown("---")
    if errores_reuso:
        for msg in errores_reuso:
            st.warning(msg)
    elif completo:
        st.success(
            f"Las **{n}** incidencias tienen saldo asignado. "
            "Puede aplicar los cambios y continuar con las descargas."
        )
    else:
        st.info(f"Faltan **{faltan}** incidencia(s) por confirmar (use **Siguiente →**).")

    if st.button(
        "Aplicar selecciones y habilitar descargas",
        type="primary",
        use_container_width=True,
        disabled=not completo or bool(errores_reuso),
        key="btn_aplicar_desempate_wizard",
    ):
        ok, msgs = aplicar_mapa_desempate(mapa)
        if ok:
            st.success("Selecciones aplicadas. Ya puede descargar Contratos y archivos globales.")
            st.rerun()
        else:
            for msg in msgs:
                st.error(msg)


def consolidacion_lista_para_descarga() -> bool:
    """True si no hay pendientes sin resolver."""
    resumen = st.session_state.get(_CLAVE_RESUMEN_LIGERO)
    if resumen is not None:
        return int(resumen.get("sin_resolver", 0)) == 0
    informe = _informe_para_ui()
    return sum(i.get("sin_resolver", 0) for i in informe) == 0


def normalizar(texto: str) -> str:
    texto = texto.lower()
    texto = unicodedata.normalize("NFD", texto)
    return "".join(c for c in texto if unicodedata.category(c) != "Mn")


def palabras_localidad(localidad: str) -> list[str]:
    palabras = re.findall(r"[a-z0-9]+", normalizar(localidad))
    significativas = [p for p in palabras if p not in PALABRAS_IGNORAR and len(p) >= 2]
    return significativas if significativas else palabras


def contiene_palabra_localidad(texto: str, localidad: str) -> bool:
    texto_norm = normalizar(texto)
    return any(palabra in texto_norm for palabra in palabras_localidad(localidad))


def verificar_lectura_matriz(libro: BytesIO) -> None:
    """Comprueba que el Excel desbloqueado tiene la hoja MATRIZ OXP."""
    libro.seek(0)
    pd.read_excel(libro, sheet_name=SHEET_MATRIZ, nrows=1, engine="openpyxl")
    libro.seek(0)


def _error_matriz_protegida(nombre_archivo: str = "") -> ValueError:
    etiqueta = f"Matriz **{nombre_archivo}**" if nombre_archivo else "Matriz"
    return ValueError(
        f"{etiqueta}: está protegida con contraseña. "
        "Ábrala en Excel, quite la contraseña de apertura si existe y súbala de nuevo."
    )


def abrir_matriz_excel(file_bytes: bytes, nombre_archivo: str = "") -> BytesIO:
    """Abre la Matriz desbloqueada (nunca Contratos)."""
    etiqueta = f"Matriz **{nombre_archivo}**" if nombre_archivo else "Matriz"
    raw = BytesIO(file_bytes)
    office = msoffcrypto.OfficeFile(raw)
    if office.is_encrypted():
        raise _error_matriz_protegida(nombre_archivo)
    raw.seek(0)
    try:
        verificar_lectura_matriz(raw)
    except Exception as e:
        raise ValueError(f"{etiqueta}: no se pudo leer ({e})") from e
    return BytesIO(file_bytes)


def _bytes_matriz_sin_reguardar(
    file_bytes: bytes,
    nombre_archivo: str = "",
) -> BytesIO:
    """Abre la Matriz sin pasar por openpyxl.save (preserva caché de fórmulas en col. V)."""
    raw = BytesIO(file_bytes)
    office = msoffcrypto.OfficeFile(raw)
    if office.is_encrypted():
        raise _error_matriz_protegida(nombre_archivo)
    return BytesIO(file_bytes)


def _cuenta_celdas_numericas(valores: list) -> int:
    n = 0
    for v in valores:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        try:
            float(v)
            n += 1
        except (TypeError, ValueError):
            continue
    return n


def _indices_saldo_final_calculado(ws, fila_hdr: int) -> tuple[int, int, int] | None:
    """Columnas Apropiación − Giros − Liberación/Fenecimiento (= fórmula Saldo Final)."""
    col_a = col_g = col_l = None
    for c in range(1, (ws.max_column or 0) + 1):
        titulo = ws.cell(fila_hdr, c).value
        if titulo is None:
            continue
        n = normalizar(str(titulo))
        if n == "apropiacion":
            col_a = c
        elif n == "giros":
            col_g = c
        elif "liberacion" in n or "fenecimiento" in n:
            col_l = c
    if col_a and col_g and col_l:
        return col_a, col_g, col_l
    return None


def _saldo_final_desde_dataframe(df: pd.DataFrame) -> pd.Series | None:
    """Apropiación − Giros − Liberación desde columnas ya cargadas (rápido)."""
    col_a = col_g = col_l = None
    for c in df.columns:
        n = normalizar(str(c))
        if n == "apropiacion":
            col_a = c
        elif n == "giros":
            col_g = c
        elif "liberacion" in n or "fenecimiento" in n:
            col_l = c
    if not (col_a and col_g and col_l):
        return None
    a = pd.to_numeric(df[col_a], errors="coerce").fillna(0)
    g = pd.to_numeric(df[col_g], errors="coerce").fillna(0)
    lib = pd.to_numeric(df[col_l], errors="coerce").fillna(0)
    return a - g - lib


def _calcular_saldo_final_matriz_ws(ws, fila_hdr: int) -> list:
    cols = _indices_saldo_final_calculado(ws, fila_hdr)
    if not cols:
        return []
    col_a, col_g, col_l = cols
    max_r = ws.max_row or fila_hdr
    filas: list = []
    for a_val, g_val, l_val in zip(
        ws.iter_rows(
            min_row=fila_hdr + 1,
            max_row=max_r,
            min_col=col_a,
            max_col=col_a,
            values_only=True,
        ),
        ws.iter_rows(
            min_row=fila_hdr + 1,
            max_row=max_r,
            min_col=col_g,
            max_col=col_g,
            values_only=True,
        ),
        ws.iter_rows(
            min_row=fila_hdr + 1,
            max_row=max_r,
            min_col=col_l,
            max_col=col_l,
            values_only=True,
        ),
    ):
        try:
            filas.append(
                float((a_val[0] if a_val else None) or 0)
                - float((g_val[0] if g_val else None) or 0)
                - float((l_val[0] if l_val else None) or 0)
            )
        except (TypeError, ValueError):
            filas.append(None)
    return filas


def _columna_matriz_data_only(
    libro: BytesIO,
    header: int,
    nombre_columna: str,
) -> list:
    """
    Valores de «Saldo Final» en MATRIZ OXP (data_only).
    Si la caché de fórmulas está vacía (p. ej. tras guardar sin abrir en Excel),
    calcula Apropiación − Giros − Liberación/Fenecimiento.
    """
    from openpyxl import load_workbook

    if header is None:
        return []

    libro.seek(0)
    wb = load_workbook(libro, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_MATRIZ]
        fila_hdr = int(header) + 1
        col_idx = None
        objetivo = normalizar(nombre_columna)
        for c in range(1, (ws.max_column or 0) + 1):
            titulo = ws.cell(fila_hdr, c).value
            if titulo is not None and normalizar(str(titulo)) == objetivo:
                col_idx = c
                break
        valores: list = []
        if col_idx is not None:
            max_r = ws.max_row or fila_hdr
            valores = [
                row[0]
                for row in ws.iter_rows(
                    min_row=fila_hdr + 1,
                    max_row=max_r,
                    min_col=col_idx,
                    max_col=col_idx,
                    values_only=True,
                )
            ]
        umbral = max(3, int(len(valores) * 0.01)) if valores else 3
        if _cuenta_celdas_numericas(valores) < umbral:
            calculados = _calcular_saldo_final_matriz_ws(ws, fila_hdr)
            if _cuenta_celdas_numericas(calculados) > _cuenta_celdas_numericas(valores):
                return calculados
        return valores
    finally:
        wb.close()


def leer_hoja_matriz(
    file_bytes: bytes,
    nombre_archivo: str = "",
    *,
    avance: callable | None = None,
    **kwargs,
) -> pd.DataFrame:
    """Lee la hoja MATRIZ OXP (archivo sin protección por contraseña)."""
    try:
        header_pd = kwargs.get("header", MATRIZ_HEADER_FILA)
        # Sin openpyxl.save: preserva caché de fórmulas en Saldo Final (col. V).
        valores_saldo: list = []
        if avance:
            avance("Matriz · abrir")
        libro = _bytes_matriz_sin_reguardar(file_bytes, nombre_archivo)
        if avance:
            avance("Matriz · leer hoja")
        df = pd.read_excel(
            libro, sheet_name=SHEET_MATRIZ, engine="openpyxl", **kwargs
        )
        if header_pd is not None:
            candidatos = ("Saldo Final", "SALDO FINAL", "Saldo final")
            col_df = next((c for c in candidatos if c in df.columns), None)
            if col_df:
                if avance:
                    avance("Matriz · Saldo Final")
                n = len(df)
                serie = pd.to_numeric(df[col_df], errors="coerce")
                umbral = max(3, int(n * 0.01))
                if int(serie.notna().sum()) < umbral:
                    calc_df = _saldo_final_desde_dataframe(df)
                    if calc_df is not None and int(calc_df.notna().sum()) > int(
                        serie.notna().sum()
                    ):
                        serie = calc_df
                    else:
                        libro.seek(0)
                        valores_saldo = _columna_matriz_data_only(
                            libro, header_pd, "Saldo Final"
                        )
                        serie = pd.to_numeric(
                            pd.Series(valores_saldo[:n]), errors="coerce"
                        )
                        if len(valores_saldo) < n:
                            serie = pd.concat(
                                [
                                    serie,
                                    pd.Series([pd.NA] * (n - len(valores_saldo))),
                                ],
                                ignore_index=True,
                            )
                df = df.copy()
                df[col_df] = serie
        return df
    except ValueError:
        raise
    except Exception as e:
        err = str(e).lower()
        if "zip" in err or "bad magic" in err or "not a zip" in err:
            raise _error_matriz_protegida(nombre_archivo) from e
        raise ValueError(f"No se pudo leer la hoja {SHEET_MATRIZ}: {e}") from e


def es_error_matriz_protegida(mensaje: str) -> bool:
    m = normalizar(mensaje)
    return "protegida" in m and "contrasena" in m


def _probar_matriz_abierta(file_bytes: bytes, nombre_archivo: str = "") -> None:
    """Comprueba que la Matriz se puede leer sin contraseña."""
    try:
        libro = _bytes_matriz_sin_reguardar(file_bytes, nombre_archivo)
        libro.seek(0)
        pd.read_excel(
            libro, sheet_name=SHEET_MATRIZ, engine="openpyxl", nrows=1
        )
    except ValueError:
        raise
    except Exception as e:
        err = str(e).lower()
        if "zip" in err or "bad magic" in err or "not a zip" in err:
            raise _error_matriz_protegida(nombre_archivo) from e
        etiqueta = f"Matriz **{nombre_archivo}**" if nombre_archivo else "Matriz"
        raise ValueError(f"{etiqueta}: no se pudo abrir ({e})") from e


def _valores_columna_a_matriz(file_bytes: bytes, nombre_archivo: str = "") -> list[str]:
    """Columna A desde fila 8 (solo validación de localidad)."""
    from openpyxl import load_workbook

    libro = _bytes_matriz_sin_reguardar(file_bytes, nombre_archivo)
    libro.seek(0)
    wb = load_workbook(libro, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_MATRIZ]
        return [
            str(ws.cell(r, 1).value or "").strip()
            for r in range(FILA_INICIO_MATRIZ, (ws.max_row or FILA_INICIO_MATRIZ) + 1)
            if ws.cell(r, 1).value is not None and str(ws.cell(r, 1).value).strip()
        ]
    finally:
        wb.close()


def texto_localidad_en_matriz(
    file_bytes: bytes,
    nombre_archivo: str = "",
) -> str:
    """Lee columna A desde fila 8 en MATRIZ OXP."""
    return " ".join(_valores_columna_a_matriz(file_bytes, nombre_archivo))


def validar_nombre_contratos(nombre_archivo: str, localidad: str) -> tuple[bool, str]:
    nombre = normalizar(nombre_archivo)
    if KW_CONTRATOS not in nombre:
        return False, f"Contratos: falta «{KW_CONTRATOS}» en **{nombre_archivo}**"
    if not contiene_palabra_localidad(nombre_archivo, localidad):
        return False, (
            f"Contratos: **{nombre_archivo}** no incluye ninguna palabra de la localidad "
            f"**{localidad}**"
        )
    return True, ""


def validar_nombre_matriz(nombre_archivo: str) -> tuple[bool, str]:
    if KW_MATRIZ not in normalizar(nombre_archivo):
        return False, f"Matriz: falta «{KW_MATRIZ}» en el nombre **{nombre_archivo}**"
    return True, ""


def validar_localidad_en_hoja_matriz(
    file_bytes: bytes,
    localidad: str,
    nombre_archivo: str,
) -> tuple[bool, str]:
    try:
        texto = texto_localidad_en_matriz(file_bytes, nombre_archivo)
    except ValueError as e:
        return False, f"**{localidad}** — Matriz **{nombre_archivo}**: {e}"
    except Exception as e:
        return False, f"**{localidad}** — Matriz **{nombre_archivo}**: no se pudo abrir ({e})"

    if not texto.strip():
        return (
            False,
            f"Matriz **{nombre_archivo}**: no se encontró localidad en columna A "
            f"(desde fila {FILA_INICIO_MATRIZ}) en **{SHEET_MATRIZ}**",
        )
    if not contiene_palabra_localidad(texto, localidad):
        return (
            False,
            f"Matriz **{nombre_archivo}**: en **{SHEET_MATRIZ}** (col. A, fila {FILA_INICIO_MATRIZ}+) "
            f"no coincide con la localidad **{localidad}**",
        )
    return True, ""


def validar_matrices_desbloqueadas(cola: list) -> tuple[bool, list[str]]:
    """Comprueba que cada Matriz se puede abrir sin contraseña."""
    errores = []
    for item in cola:
        loc = item["localidad"]
        nm = item["matriz"]["name"]
        try:
            _probar_matriz_abierta(
                bytes_archivo_cola(item["matriz"]),
                item["matriz"]["name"],
            )
        except ValueError as e:
            errores.append(f"**{loc}** — Matriz **{nm}**: {e}")
        except Exception as e:
            errores.append(f"**{loc}** — Matriz **{nm}**: {e}")
    return len(errores) == 0, errores


def _validar_nombres_en_cola(
    cola: list,
    *,
    verificar_texto_localidad: bool = True,
) -> tuple[bool, list[str]]:
    errores: list[str] = []
    for item in cola:
        loc = item["localidad"]
        nc = item["contratos"]["name"]
        nm = item["matriz"]["name"]

        ok_c, msg_c = validar_nombre_contratos(nc, loc)
        if not ok_c:
            errores.append(f"**{loc}** — {msg_c}")

        ok_nm, msg_nm = validar_nombre_matriz(nm)
        if not ok_nm:
            errores.append(f"**{loc}** — {msg_nm}")
        elif verificar_texto_localidad:
            ok_m, msg_m = validar_localidad_en_hoja_matriz(
                bytes_archivo_cola(item["matriz"]), loc, nm
            )
            if not ok_m:
                errores.append(f"**{loc}** — {msg_m}")

    return len(errores) == 0, errores


def validar_cola_archivos(
    cola: list,
    *,
    verificar_texto_localidad: bool = True,
) -> tuple[bool, list[str]]:
    errores = list(_validar_archivos_accesibles(cola))
    if errores:
        return False, errores

    mat_ok, errores_mat = validar_matrices_desbloqueadas(cola)
    if not mat_ok:
        return False, errores_mat

    return _validar_nombres_en_cola(
        cola,
        verificar_texto_localidad=verificar_texto_localidad,
    )


def file_to_buffer(uploaded_file, localidad: str = "", tipo: str = "archivo") -> dict:
    data = uploaded_file.getvalue()
    carpeta = _directorio_archivos_sesion()
    base = sanitizar_nombre_archivo(uploaded_file.name)
    prefijo = sanitizar_nombre_archivo(
        f"{tipo}_{localidad}_{uuid.uuid4().hex[:8]}"
    )
    destino = carpeta / f"{prefijo}_{base}"
    destino.write_bytes(data)
    return {"path": str(destino), "name": uploaded_file.name}


def buffer_to_file(entry: dict):
    return BytesIO(bytes_archivo_cola(entry))


def read_contratos(file_like, name: str, localidad: str):
    try:
        df = pd.read_excel(file_like)
        df = df.copy()
        df["Localidad"] = localidad
        df["Tipo archivo"] = "Contratos plan de choque"
        df["Archivo origen"] = name
        return df
    except Exception as e:
        st.error(f"No se pudo leer Contratos (**{name}**): {e}")
        return None


def read_matriz(file_bytes: bytes, name: str, localidad: str):
    try:
        df = leer_hoja_matriz(file_bytes, name)
    except ValueError as e:
        st.session_state.error_ultima_ejecucion = str(e)
        return None
    except Exception as e:
        st.session_state.error_ultima_ejecucion = f"No se pudo leer Matriz ({name}): {e}"
        return None

    df = df.copy()
    df["Localidad"] = localidad
    df["Tipo archivo"] = "Matriz"
    df["Archivo origen"] = name
    return df


def generate_excel_bytes(dataframe: pd.DataFrame, sheet_name: str = "Datos") -> BytesIO:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        dataframe.to_excel(writer, index=False, sheet_name=sheet_name)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        fmt = workbook.add_format({
            "bold": True, "text_wrap": True, "valign": "top",
            "border": 1, "bg_color": "#1e40af", "font_color": "#ffffff",
        })
        for col_num, value in enumerate(dataframe.columns.values):
            worksheet.write(0, col_num, value, fmt)
            worksheet.set_column(col_num, col_num, 18)
    output.seek(0)
    return output


def localidades_analizadas(stats: list) -> list[str]:
    return sorted({s["Localidad"] for s in stats if s.get("Localidad")})


def construir_avance_plan_de_choque(
    consolidated_df: pd.DataFrame, stats: list, fecha: datetime
) -> pd.DataFrame:
    """Contenido de Avance plan de choque — pendiente reglas de negocio."""
    localidades = localidades_analizadas(stats)
    return pd.DataFrame({
        "Campo": [
            "Archivo",
            "Mes de análisis",
            "Fecha de análisis",
            "Cantidad de localidades",
            "Registros en consolidado",
            "Estado",
        ],
        "Valor": [
            nombre_archivo_salida(ARCHIVO_AVANCE_BASE, fecha, localidades).replace(
                ".xlsx", ""
            ),
            mes_en_espanol(fecha),
            formato_fecha_colombia(fecha, con_hora=True),
            len(localidades),
            len(consolidated_df) if consolidated_df is not None else 0,
            (
                "Pendiente definir análisis de avance "
                "(consolidará Matriz + Contratos actualizados por localidad)"
            ),
        ],
    })


def construir_tabla_resumen(
    consolidated_df: pd.DataFrame, stats: list, fecha: datetime
) -> pd.DataFrame:
    """Tabla de resumen con metadatos y cruce CXP."""
    localidades = localidades_analizadas(stats)
    titulo_mes = titulo_saldo_corte(fecha)
    meta = pd.DataFrame({
        "Campo": [
            "Mes de análisis",
            "Fecha de análisis",
            "Columna en Contratos plan de choque",
            "Cantidad de localidades",
            "Fuentes consolidadas",
        ],
        "Valor": [
            mes_capitalizado(fecha),
            formato_fecha_colombia(fecha, con_hora=True),
            titulo_mes,
            len(localidades),
            "Por localidad: Matriz + Contratos plan de choque actualizados",
        ],
    })
    partes = [meta]
    informe = st.session_state.get("cruce_informe") or []
    if informe:
        partes.append(
            pd.DataFrame([{"Campo": "— Localidades (resumen) —", "Valor": ""}])
        )
        df_loc = dataframe_resumen_localidades(informe)
        partes.append(
            pd.DataFrame(
                [
                    {
                        "Campo": row["Localidad"],
                        "Valor": (
                            f"{row['Contratos cruzados']} contratos cruzados · "
                            f"cruce directo {row['Cruce directo']} · "
                            f"cruce por contrato {row['Cruce por contrato']} · "
                            f"cruce por selección {row['Cruce por selección']} · "
                            f"CXP {row['CXP del mes']}"
                        ),
                    }
                    for _, row in df_loc.iterrows()
                ]
            )
        )
    resumen_global = st.session_state.get("cruce_resumen_global") or []
    if resumen_global:
        partes.append(
            pd.DataFrame([{"Campo": "— Métodos de cruce (global) —", "Valor": ""}])
        )
        rg = dataframe_metodos_amigable(resumen_global)
        rg.columns = ["Campo", "Valor"]
        partes.append(rg)
    contratos_act = st.session_state.get("contratos_actualizados") or {}
    stats_por_loc = {}
    for s in stats:
        loc = s.get("Localidad")
        if not loc:
            continue
        stats_por_loc.setdefault(loc, []).append(s)

    for loc_info in informe:
        loc = loc_info["localidad"]
        if not _localidad_requiere_detalle_cruce(loc_info):
            continue
        filas_loc = stats_por_loc.get(loc, [])
        matriz_nombre = next(
            (s.get("Nombre") for s in filas_loc if s.get("Archivo") == "Matriz"), ""
        )
        contratos_orig = next(
            (s.get("Nombre") for s in filas_loc if "Contratos" in str(s.get("Archivo", ""))),
            "",
        )
        data_cto = contratos_act.get(loc, {})
        contratos_gen = nombre_descarga_contratos_actualizado(
            loc,
            data_cto.get("nombre_contratos") or contratos_orig,
            fecha,
        )
        partes.append(
            pd.DataFrame([
                {
                    "Campo": f"— Detalle {loc} —",
                    "Valor": (
                        f"{loc_info['contratos_ok']}/{loc_info['total_contratos']} cruzados · "
                        f"sin coincidencia automática {loc_info['sin_resolver']}"
                    ),
                },
                {"Campo": "Matriz (origen)", "Valor": matriz_nombre or "—"},
                {"Campo": "Contratos (origen)", "Valor": contratos_orig or "—"},
                {"Campo": "Contratos (actualizado)", "Valor": contratos_gen},
            ])
        )
        if loc_info.get("resumen_metodos"):
            lm = dataframe_metodos_amigable(loc_info["resumen_metodos"])
            lm.columns = ["Campo", "Valor"]
            partes.append(lm)

    if informe and stats_por_loc:
        partes.append(pd.DataFrame([{"Campo": "— Archivos origen —", "Valor": ""}]))
        for loc_info in informe:
            loc = loc_info["localidad"]
            filas_loc = stats_por_loc.get(loc, [])
            matriz_nombre = next(
                (s.get("Nombre") for s in filas_loc if s.get("Archivo") == "Matriz"), "—"
            )
            contratos_orig = next(
                (s.get("Nombre") for s in filas_loc if "Contratos" in str(s.get("Archivo", ""))),
                "—",
            )
            partes.append(
                pd.DataFrame([
                    {
                        "Campo": loc,
                        "Valor": f"Matriz: {matriz_nombre} | Contratos: {contratos_orig}",
                    }
                ])
            )
    detalle = obtener_cruce_detalle()
    filas_sr = filas_sin_resolver(detalle)
    if filas_sr and consolidacion_lista_para_descarga():
        partes.append(
            pd.DataFrame([{"Campo": "— Contratos sin coincidencia automática (histórico) —", "Valor": len(filas_sr)}])
        )
    elif filas_sr:
        partes.append(
            pd.DataFrame([
                {
                    "Campo": "— Contratos sin coincidencia automática —",
                    "Valor": (
                        f"{len(filas_sr)} — casos sin coincidencia automática"
                    ),
                }
            ])
        )

    if stats:
        partes.append(pd.DataFrame([{"Campo": "— Archivos —", "Valor": ""}]))
        partes.append(pd.DataFrame(stats))
    if len(partes) > 1:
        return pd.concat(partes, ignore_index=True, sort=False)
    return pd.DataFrame({
        "Nota": [
            "Pendiente definir Tabla de resumen",
            f"Mes: {mes_en_espanol(fecha)}",
            f"Registros en consolidado: {len(consolidated_df) if consolidated_df is not None else 0}",
        ],
    })


def carpeta_descargas() -> Path:
    """Carpeta Descargas del usuario (Windows en español o inglés)."""
    home = Path.home()
    for nombre in ("Downloads", "Descargas"):
        carpeta = home / nombre
        if carpeta.is_dir():
            return carpeta
    return home


def guardar_archivos_salida(
    consolidated_df: pd.DataFrame, stats: list
) -> list[Path]:
    """
    Guarda siempre 2 archivos en Descargas.
    El nombre incluye mes del análisis y cada localidad, p. ej.:
    - Avance plan de choque Mayo (Usaquén,Kennedy).xlsx
    - Tabla de resumen Mayo (Usaquén,Kennedy).xlsx
    """
    if not stats and (consolidated_df is None or consolidated_df.empty):
        raise ValueError("No hay datos consolidados para exportar.")

    fecha = fecha_referencia_analisis()
    localidades = localidades_analizadas(stats)
    carpeta = carpeta_descargas()
    df_avance = construir_avance_plan_de_choque(consolidated_df, stats, fecha)
    df_resumen = construir_tabla_resumen(consolidated_df, stats, fecha)

    salidas = [
        (
            nombre_archivo_salida(ARCHIVO_AVANCE_BASE, fecha, localidades),
            df_avance,
            "Avance",
        ),
        (
            nombre_archivo_salida(ARCHIVO_RESUMEN_BASE, fecha, localidades),
            df_resumen,
            "Resumen",
        ),
    ]

    rutas = []
    for nombre, df_salida, hoja in salidas:
        ruta = carpeta / nombre
        ruta.write_bytes(generate_excel_bytes(df_salida, sheet_name=hoja).getvalue())
        rutas.append(ruta)

    return rutas


def formulario_completo(localidad, contratos, matriz) -> bool:
    return (
        bool(localidad)
        and localidad != SELECCION_LOCALIDAD
        and contratos is not None
        and matriz is not None
    )


def entrada_desde_formulario(localidad, contratos, matriz) -> dict:
    return {
        "localidad": localidad,
        "contratos": file_to_buffer(contratos, localidad, "contratos"),
        "matriz": file_to_buffer(matriz, localidad, "matriz"),
    }


def item_tiene_contratos_y_matriz(item: dict) -> bool:
    tiene_c = entrada_cola_tiene_archivo(item.get("contratos"))
    tiene_m = entrada_cola_tiene_archivo(item.get("matriz"))
    return tiene_c and tiene_m


def validar_archivos_en_cola(cola: list) -> tuple[bool, list[str]]:
    errores = []
    for item in cola:
        loc = item.get("localidad", "Localidad")
        contratos = item.get("contratos") or {}
        matriz = item.get("matriz") or {}
        if not entrada_cola_tiene_archivo(contratos):
            errores.append(
                _mensaje_archivo_cola_no_disponible(
                    loc,
                    "Contratos plan de choque",
                    contratos,
                )
            )
        if not entrada_cola_tiene_archivo(matriz):
            errores.append(_mensaje_archivo_cola_no_disponible(loc, "Matriz", matriz))
    return len(errores) == 0, errores


def texto_archivo_cola(etiqueta: str, entrada: dict | None) -> str:
    entrada = entrada or {}
    nombre = entrada.get("name") or "Sin archivo"
    if entrada_cola_tiene_archivo(entrada):
        return f"{etiqueta}: {nombre}"
    return f"{etiqueta}: {nombre} (archivo no disponible; cárguelo de nuevo)"


def cola_para_ejecutar(cola: list) -> list:
    """Ítems de la cola con Contratos y Matriz listos para consolidar."""
    return [i for i in cola if item_tiene_contratos_y_matriz(i)]


def puede_ejecutar_cola(cola: list) -> bool:
    """La cola tiene al menos un consolidado completo."""
    return bool(cola) and all(item_tiene_contratos_y_matriz(i) for i in cola)


def _borrar_archivo_cola_en_disco(entrada: dict | None) -> None:
    if not entrada:
        return
    ruta = entrada.get("path")
    if not ruta:
        return
    try:
        Path(ruta).unlink(missing_ok=True)
    except OSError:
        pass


def _rutas_archivos_cola(cola: list) -> set[str]:
    rutas: set[str] = set()
    for item in cola:
        for clave in ("contratos", "matriz"):
            ruta = (item.get(clave) or {}).get("path")
            if ruta:
                rutas.add(str(Path(ruta)))
    return rutas


def _borrar_archivo_cola_si_no_esta_en_uso(
    entrada: dict | None,
    rutas_en_uso: set[str],
) -> None:
    ruta = (entrada or {}).get("path")
    if ruta and str(Path(ruta)) in rutas_en_uso:
        return
    _borrar_archivo_cola_en_disco(entrada)


def _limpiar_estado_ejecucion_por_cambio_cola() -> None:
    st.session_state.cola_ejecucion = []
    st.session_state.pendiente_consolidacion = False
    st.session_state.ejecutar_consolidacion_ahora = False
    st.session_state.consolidacion_en_curso = False
    st.session_state.pop("_pc_validacion_entrada_fallo", None)
    st.session_state.pop("_pc_validacion_entrada_mensaje", None)
    st.session_state.pop("_pc_validacion_entrada_errores", None)
    st.session_state.pop("_pc_mostrar_formulario_correccion", None)
    st.session_state.pop("consolidacion_work", None)


def quitar_de_cola(localidad: str, *, mantener_modo_correccion: bool = False) -> None:
    cola_actual = list(st.session_state.get("cola_localidades", []))
    eliminados = [i for i in cola_actual if i.get("localidad") == localidad]
    restantes = [i for i in cola_actual if i.get("localidad") != localidad]
    rutas_restantes = _rutas_archivos_cola(restantes)
    for item in eliminados:
        _borrar_archivo_cola_si_no_esta_en_uso(item.get("contratos"), rutas_restantes)
        _borrar_archivo_cola_si_no_esta_en_uso(item.get("matriz"), rutas_restantes)
    st.session_state.cola_localidades = restantes
    _limpiar_estado_ejecucion_por_cambio_cola()
    if restantes and mantener_modo_correccion:
        st.session_state["_pc_mostrar_cola_correccion"] = True
    elif restantes and all(item_tiene_contratos_y_matriz(i) for i in restantes):
        st.session_state.pop("_pc_mostrar_cola_correccion", None)
    elif restantes:
        st.session_state["_pc_mostrar_cola_correccion"] = True
    else:
        st.session_state.pop("_pc_mostrar_cola_correccion", None)


def limpiar_resultado_consolidado():
    st.session_state.processed = False
    st.session_state.pop("consolidated_df", None)
    st.session_state.file_stats = []
    st.session_state.last_processed_at = None
    st.session_state.error_ultima_ejecucion = None
    st.session_state.errores_ejecucion = []
    st.session_state.fecha_analisis = None
    st.session_state.cruce_informe = []
    _borrar_cruce_detalle_en_disco()
    _borrar_snapshot_consolidacion()
    st.session_state.pop(_CLAVE_MOSTRAR_RESULTADOS, None)
    st.session_state.contratos_actualizados = {}
    st.session_state.cruce_resumen_global = []
    st.session_state.titulo_saldo_corte = ""
    st.session_state.pop("consolidacion_work", None)
    st.session_state.pop("zip_descarga_contratos", None)
    st.session_state.pop("zip_descarga_listo", None)
    st.session_state.pop(_CLAVE_RUTA_REPORTE, None)
    st.session_state.pop("reporte_ejecucion", None)
    _reset_estado_desempate_wizard()


def iniciar_nuevo_reporte() -> None:
    for item in st.session_state.get("cola_localidades", []):
        _borrar_archivo_cola_en_disco(item.get("contratos"))
        _borrar_archivo_cola_en_disco(item.get("matriz"))
    st.session_state.cola_localidades = []
    st.session_state.cola_ejecucion = []
    st.session_state.pendiente_consolidacion = False
    st.session_state.ejecutar_consolidacion_ahora = False
    st.session_state.consolidacion_en_curso = False
    st.session_state.pop("_pc_validacion_entrada_fallo", None)
    st.session_state.pop("_pc_mostrar_cola_correccion", None)
    st.session_state.pop("_pc_mostrar_formulario_correccion", None)
    st.session_state.upload_key = int(st.session_state.get("upload_key", 0)) + 1
    limpiar_resultado_consolidado()
    _purgar_uploaders_obsoletos()


def _agregar_conteo_global(acumulado: dict, conteo: dict) -> None:
    for codigo, cantidad in conteo.items():
        acumulado[codigo] = acumulado.get(codigo, 0) + cantidad


def _guardar_reporte_en_sesion(reporte: ReporteEjecucion) -> None:
    if not reporte.tiene_casos():
        st.session_state.pop("reporte_ejecucion", None)
        st.session_state.pop(_CLAVE_RUTA_REPORTE, None)
        return
    df = reporte.a_dataframe()
    payload = {
        "texto": reporte.generar_texto(),
        "tabla": df.to_dict("records"),
        "resumen": reporte.resumen,
        "generado": datetime.now().isoformat(timespec="seconds"),
    }
    ruta = _directorio_archivos_sesion() / "reporte_ejecucion.pkl"
    with open(ruta, "wb") as f:
        pickle.dump(payload, f, protocol=4)
    st.session_state[_CLAVE_RUTA_REPORTE] = str(ruta)
    st.session_state.pop("reporte_ejecucion", None)


def _cargar_reporte_ejecucion() -> dict | None:
    legacy = st.session_state.get("reporte_ejecucion")
    if legacy and legacy.get("tabla"):
        return legacy
    ruta = st.session_state.get(_CLAVE_RUTA_REPORTE)
    if ruta and Path(ruta).is_file():
        with open(ruta, "rb") as f:
            return pickle.load(f)
    return None


def mostrar_reporte_tecnico_admin() -> None:
    """Solo casos no previstos del sistema (para quien mantiene el código)."""
    payload = _cargar_reporte_ejecucion()
    if not payload or not payload.get("tabla"):
        return

    with st.expander("Casos no previstos (para soporte técnico)", expanded=False):
        st.caption(payload.get("resumen", ""))
        nombre_archivo = (
            f"casos_no_previstos_{payload.get('generado', 'ejecucion')}.txt"
        ).replace(":", "-")

        st.download_button(
            "Descargar reporte (.txt)",
            data=payload.get("texto", ""),
            file_name=nombre_archivo,
            mime="text/plain",
            key="dl_reporte_casos_no_previstos",
            use_container_width=True,
        )
        st.dataframe(
            pd.DataFrame(payload["tabla"]),
            use_container_width=True,
            hide_index=True,
        )


# Barra por pasos iguales (más movimiento visible); texto: localidad + fase.
_PASOS_VALIDACION = 4
_PASOS_POR_LOCALIDAD = 14  # 3 Matriz + 9 Cruce/Excel + 2 cierre


def _barra_nueva(num_localidades: int) -> dict:
    n = max(int(num_localidades), 1)
    return {"paso": 0, "total": _PASOS_VALIDACION + n * _PASOS_POR_LOCALIDAD}


def _texto_fase_barra(
    fase: str,
    localidad: str | None = None,
    indice: int = 0,
    total: int = 0,
) -> str:
    if localidad and total:
        return f"{localidad} ({indice + 1}/{total}) · {fase}"
    if localidad:
        return f"{localidad} · {fase}"
    return fase


def _barra_tick(
    barra: dict,
    progress,
    fase: str,
    *,
    localidad: str | None = None,
    indice: int = 0,
    total: int = 0,
) -> None:
    barra["paso"] = int(barra.get("paso", 0)) + 1
    valor = barra["paso"] / max(int(barra.get("total", 1)), 1)
    _actualizar_barra_progreso(
        progress,
        min(valor, 1.0),
        _texto_fase_barra(fase, localidad, indice, total),
    )


def _actualizar_barra_progreso(progress, valor: float, texto: str) -> None:
    """Actualiza la barra (sin bloquear la ejecución en Streamlit Cloud)."""
    if progress is None:
        return
    progress.progress(min(max(valor, 0.0), 1.0), text=texto)


def _omitir_formulario_entrada() -> bool:
    """Solo durante un run activo (Continuar o paso siguiente), no al abrir la página."""
    return bool(st.session_state.get("ejecutar_consolidacion_ahora"))


def _sanear_sesion_consolidacion() -> None:
    """Evita quedar colgado si la sesión quedó a medias tras F5 o un cierre."""
    if st.session_state.get("consolidacion_work") and not st.session_state.get(
        "ejecutar_consolidacion_ahora"
    ):
        st.session_state.consolidacion_en_curso = False
        st.session_state.pendiente_consolidacion = False
    if st.session_state.get("pendiente_consolidacion") and not st.session_state.get(
        "ejecutar_consolidacion_ahora"
    ):
        st.session_state.pendiente_consolidacion = False


def _reporte_con_casos_guardados(casos: list) -> ReporteEjecucion:
    reporte = ReporteEjecucion()
    for datos in casos:
        reporte.casos.append(CasoNoPrevisto(**datos))
    return reporte


def _procesar_localidad_en_work(
    work: dict,
    item: dict,
    reporte: ReporteEjecucion,
    progress,
    indice: int,
    total: int,
) -> None:
    """Cruza una localidad y acumula el resultado en work."""
    ahora = work["ahora"]
    titulo_mes = work["titulo_mes"]
    localidad = item["localidad"]
    barra = work["_barra"]

    def tick(fase: str) -> None:
        _barra_tick(
            barra,
            progress,
            fase,
            localidad=localidad,
            indice=indice,
            total=total,
        )

    try:
        df_matriz = leer_hoja_matriz(
            bytes_archivo_cola(item["matriz"]),
            item["matriz"]["name"],
            header=MATRIZ_HEADER_FILA,
            avance=tick,
        )
    except ValueError as e:
        work["errores"].append(f"**{localidad}** — Matriz: {e}")
        reporte.desde_excepcion(
            e,
            localidad=localidad,
            archivo=item["matriz"]["name"],
            fase="lectura_matriz",
        )
        return
    except Exception as e:
        work["errores"].append(f"**{localidad}** — Matriz: no se pudo leer ({e})")
        reporte.desde_excepcion(
            e,
            localidad=localidad,
            archivo=item["matriz"]["name"],
            fase="lectura_matriz",
        )
        return

    try:
        resultado = procesar_localidad_cxp(
            bytes_archivo_cola(item["contratos"]),
            df_matriz,
            localidad,
            ahora,
            item["contratos"]["name"],
            item["matriz"]["name"],
            avance=tick,
        )
    except ValueError as e:
        work["errores"].append(f"**{localidad}** — Contratos: {e}")
        reporte.desde_excepcion(
            e,
            localidad=localidad,
            archivo=item["contratos"]["name"],
            fase="procesamiento_contratos",
        )
        return
    except Exception as e:
        work["errores"].append(
            f"**{localidad}** — Contratos: no se pudo procesar ({e})"
        )
        reporte.desde_excepcion(
            e,
            localidad=localidad,
            archivo=item["contratos"]["name"],
            fase="procesamiento_contratos",
        )
        return

    tick("Excel · aplicado")
    registrar_resultado_localidad(reporte, item, resultado)
    _agregar_conteo_global(work["conteo_global"], resultado["conteo"])
    work["informe_localidades"].append({
        "localidad": localidad,
        "columna_mes": resultado["columna_mes"],
        "accion_columna": resultado["accion_columna"],
        "total_contratos": resultado["total_contratos"],
        "contratos_ok": resultado["contratos_ok"],
        "sin_resolver": resultado["sin_resolver"],
        "cxp_total": resultado["cxp_total"],
        "resumen_metodos": resultado["resumen_metodos"],
        "conteo": resultado["conteo"],
        "advertencias_suspendidos": resultado.get("advertencias_suspendidos", []),
    })
    work["detalle_global"].extend(resultado["detalle"])
    work["contratos_actualizados"][localidad] = _guardar_contratos_en_disco(
        localidad, resultado
    )
    work["stats"].extend([
        {
            "Localidad": localidad,
            "Archivo": f"Contratos ({resultado.get('hoja_cruce', 'Cps/Caja por depurar')})",
            "Nombre": item["contratos"]["name"],
            "Filas": resultado["total_contratos"],
            "CXP (suma mes)": resultado["cxp_total"],
            f"Columna {titulo_mes}": resultado["accion_columna"],
        },
        {
            "Localidad": localidad,
            "Archivo": "Matriz",
            "Nombre": item["matriz"]["name"],
            "Filas": len(df_matriz),
        },
    ])

    tick("Localidad · lista")


def _aplicar_work_a_sesion(work: dict) -> bool:
    errores = work["errores"]
    informe = work["informe_localidades"]
    total = len(work["cola"])
    titulo_mes = work["titulo_mes"]
    ahora = work["ahora"]

    if errores:
        st.session_state.errores_ejecucion = list(errores)
        st.session_state.pop("consolidacion_work", None)
        st.session_state.processed = False
        return False

    if len(informe) != total:
        st.session_state.errores_ejecucion = [
            f"**{item['localidad']}** — no se pudo consolidar (revise los mensajes anteriores)."
            for item in work["cola"]
            if item["localidad"] not in {x["localidad"] for x in informe}
        ]
        st.session_state.pop("consolidacion_work", None)
        st.session_state.processed = False
        return False

    conteo_global = work["conteo_global"]
    detalle_global = work["detalle_global"]
    resumen_global = [
        {
            "Método": METODOS_LABEL.get(codigo, codigo),
            "Contratos": cantidad,
        }
        for codigo, cantidad in sorted(conteo_global.items(), key=lambda x: -x[1])
        if cantidad > 0
    ]

    st.session_state.cruce_informe = informe
    establecer_cruce_detalle(detalle_global)
    st.session_state.contratos_actualizados = work["contratos_actualizados"]
    st.session_state.cruce_resumen_global = resumen_global
    st.session_state.file_stats = work["stats"]
    st.session_state.processed = True
    st.session_state.fecha_analisis = ahora
    st.session_state.last_processed_at = formato_fecha_colombia(ahora, con_hora=True)
    st.session_state.titulo_saldo_corte = titulo_mes
    _reset_estado_desempate_wizard()
    st.session_state.zip_descarga_listo = False
    _regenerar_zip_descarga_contratos()
    _persistir_snapshot_consolidacion()
    st.session_state[_CLAVE_MOSTRAR_RESULTADOS] = True
    st.session_state[_CLAVE_SCROLL_RESULTADOS] = True
    return True


def ejecutar_consolidacion(
    cola,
    reporte: ReporteEjecucion,
    progress=None,
    fraccion_inicio: float = 0.0,
    fraccion_fin: float = 1.0,
):
    del fraccion_inicio, fraccion_fin
    stats, errores = [], []
    informe_localidades = []
    detalle_global = []
    contratos_actualizados = {}
    conteo_global: dict[str, int] = {}
    total = len(cola) or 1
    ahora = datetime.now()
    titulo_mes = titulo_saldo_corte(ahora)
    barra = _barra_nueva(total)

    for i, item in enumerate(cola):
        localidad = item["localidad"]

        def tick(fase: str) -> None:
            _barra_tick(
                barra,
                progress,
                fase,
                localidad=localidad,
                indice=i,
                total=total,
            )

        try:
            df_matriz = leer_hoja_matriz(
                bytes_archivo_cola(item["matriz"]),
                item["matriz"]["name"],
                header=MATRIZ_HEADER_FILA,
                avance=tick,
            )
        except ValueError as e:
            errores.append(f"**{localidad}** — Matriz: {e}")
            reporte.desde_excepcion(
                e,
                localidad=localidad,
                archivo=item["matriz"]["name"],
                fase="lectura_matriz",
            )
            continue
        except Exception as e:
            errores.append(f"**{localidad}** — Matriz: no se pudo leer ({e})")
            reporte.desde_excepcion(
                e,
                localidad=localidad,
                archivo=item["matriz"]["name"],
                fase="lectura_matriz",
            )
            continue

        try:
            resultado = procesar_localidad_cxp(
                bytes_archivo_cola(item["contratos"]),
                df_matriz,
                localidad,
                ahora,
                item["contratos"]["name"],
                item["matriz"]["name"],
                avance=tick,
            )
        except ValueError as e:
            errores.append(f"**{localidad}** — Contratos: {e}")
            reporte.desde_excepcion(
                e,
                localidad=localidad,
                archivo=item["contratos"]["name"],
                fase="procesamiento_contratos",
            )
            continue
        except Exception as e:
            errores.append(f"**{localidad}** — Contratos: no se pudo procesar ({e})")
            reporte.desde_excepcion(
                e,
                localidad=localidad,
                archivo=item["contratos"]["name"],
                fase="procesamiento_contratos",
            )
            continue

        tick("Excel · aplicado")
        registrar_resultado_localidad(reporte, item, resultado)
        _agregar_conteo_global(conteo_global, resultado["conteo"])
        informe_localidades.append({
            "localidad": localidad,
            "columna_mes": resultado["columna_mes"],
            "accion_columna": resultado["accion_columna"],
            "total_contratos": resultado["total_contratos"],
            "contratos_ok": resultado["contratos_ok"],
            "sin_resolver": resultado["sin_resolver"],
            "cxp_total": resultado["cxp_total"],
            "resumen_metodos": resultado["resumen_metodos"],
            "conteo": resultado["conteo"],
            "advertencias_suspendidos": resultado.get("advertencias_suspendidos", []),
        })
        detalle_global.extend(resultado["detalle"])
        contratos_actualizados[localidad] = _guardar_contratos_en_disco(
            localidad, resultado
        )

        stats.extend([
            {
                "Localidad": localidad,
                "Archivo": f"Contratos ({resultado.get('hoja_cruce', 'Cps/Caja por depurar')})",
                "Nombre": item["contratos"]["name"],
                "Filas": resultado["total_contratos"],
                "CXP (suma mes)": resultado["cxp_total"],
                f"Columna {titulo_mes}": resultado["accion_columna"],
            },
            {
                "Localidad": localidad,
                "Archivo": "Matriz",
                "Nombre": item["matriz"]["name"],
                "Filas": len(df_matriz),
            },
        ])

        tick("Localidad · lista")

    _barra_tick(barra, progress, "Consolidación · terminada")
    if progress is not None:
        progress.empty()

    if errores:
        st.session_state.errores_ejecucion = errores
        limpiar_resultado_consolidado()
        return False

    if len(informe_localidades) != total:
        limpiar_resultado_consolidado()
        return False

    resumen_global = [
        {
            "Método": METODOS_LABEL.get(codigo, codigo),
            "Contratos": cantidad,
        }
        for codigo, cantidad in sorted(conteo_global.items(), key=lambda x: -x[1])
        if cantidad > 0
    ]

    st.session_state.cruce_informe = informe_localidades
    establecer_cruce_detalle(detalle_global)
    st.session_state.contratos_actualizados = contratos_actualizados
    st.session_state.cruce_resumen_global = resumen_global
    st.session_state.file_stats = stats
    st.session_state.processed = True
    st.session_state.fecha_analisis = ahora
    st.session_state.last_processed_at = formato_fecha_colombia(ahora, con_hora=True)
    st.session_state.titulo_saldo_corte = titulo_mes
    _reset_estado_desempate_wizard()
    _persistir_snapshot_consolidacion()
    st.session_state[_CLAVE_MOSTRAR_RESULTADOS] = True
    st.session_state[_CLAVE_SCROLL_RESULTADOS] = True
    return True


def _consolidacion_corriendo() -> bool:
    return bool(
        st.session_state.get("ejecutar_consolidacion_ahora")
        or st.session_state.get("consolidacion_en_curso")
    )


def _limpiar_estado_consolidacion_bloqueado() -> None:
    """Evita quedar con el botón deshabilitado si una ejecución anterior se interrumpió."""
    if st.session_state.get("consolidacion_en_curso") and not st.session_state.get(
        "ejecutar_consolidacion_ahora"
    ):
        st.session_state.consolidacion_en_curso = False


def _consumir_disparador_consolidacion() -> bool:
    """True si el usuario pulsó Continuar en el paso anterior."""
    if st.session_state.pop("pendiente_consolidacion", False):
        return True
    if st.session_state.pop("iniciar_consolidacion", False):
        return True
    return False


def _resolver_cola_para_ejecutar() -> list:
    cola = st.session_state.get("cola_ejecucion") or []
    if not cola:
        base = st.session_state.get("cola_localidades") or []
        if base:
            cola = cola_para_ejecutar(base)
            st.session_state.cola_ejecucion = cola
    return _asegurar_cola_en_disco(cola) if cola else []


def _ejecutar_consolidacion_si_pendiente(progress=None) -> None:
    if not st.session_state.get("ejecutar_consolidacion_ahora"):
        return
    work = st.session_state.get("consolidacion_work")
    disparo = _consumir_disparador_consolidacion()
    if not disparo and not work:
        st.session_state.ejecutar_consolidacion_ahora = False
        st.session_state.consolidacion_en_curso = False
        return
    if disparo:
        cola = _resolver_cola_para_ejecutar()
        if not cola:
            st.session_state.ejecutar_consolidacion_ahora = False
            st.session_state.consolidacion_en_curso = False
            if progress is not None:
                progress.empty()
            st.error(
                "No hay cola de ejecución. Pulse **Ejecutar consolidación** de nuevo "
                "y luego **Continuar**."
            )
            return
        procesar_consolidacion(cola, progress=progress, reiniciar=True)
    else:
        cola = _asegurar_cola_en_disco(work["cola"])
        procesar_consolidacion(cola, progress=progress, reiniciar=False)


def procesar_consolidacion(
    cola_run: list,
    progress=None,
    *,
    reiniciar: bool = True,
):
    n = len(cola_run)
    if n == 0:
        st.error("La cola está vacía. Añada localidades con Contratos y Matriz.")
        return
    st.session_state.consolidacion_en_curso = True
    if progress is None:
        progress = st.progress(0, text="Preparando consolidación…")
    try:
        work = st.session_state.get("consolidacion_work")
        if reiniciar or work is None:
            limpiar_resultado_consolidado()
            cola_run = _asegurar_cola_en_disco(cola_run)
            barra = _barra_nueva(n)
            _barra_tick(barra, progress, "Validación · archivos")
            errores_acceso = _validar_archivos_accesibles(cola_run)
            if errores_acceso:
                nombres_ok, errores_nombres = False, errores_acceso
            else:
                _barra_tick(barra, progress, "Validación · Matriz")
                mat_ok, errores_mat = validar_matrices_desbloqueadas(cola_run)
                if not mat_ok:
                    nombres_ok, errores_nombres = False, errores_mat
                else:
                    _barra_tick(barra, progress, "Validación · nombres")
                    nombres_ok, errores_nombres = _validar_nombres_en_cola(
                        cola_run,
                        verificar_texto_localidad=False,
                    )
            if nombres_ok:
                _barra_tick(barra, progress, "Validación · lista")
            if not nombres_ok:
                st.session_state["_pc_validacion_entrada_fallo"] = True
                st.session_state["_pc_mostrar_cola_correccion"] = True
                st.session_state.ejecutar_consolidacion_ahora = False
                st.session_state.pendiente_consolidacion = False
                st.session_state.consolidacion_en_curso = False
                reporte = ReporteEjecucion()
                reporte.cerrar(False)
                if any(es_error_matriz_protegida(e) for e in errores_nombres):
                    mensaje_fallo = (
                        "Revisión detenida. La Matriz parece tener contraseña de apertura."
                    )
                else:
                    mensaje_fallo = (
                        "Revisión detenida. Hay una entrada de la cola que no se pudo validar."
                    )
                st.session_state["_pc_validacion_entrada_mensaje"] = mensaje_fallo
                st.session_state["_pc_validacion_entrada_errores"] = list(errores_nombres)
                if progress is not None:
                    progress.empty()
                return

            ahora = datetime.now()
            work = {
                "cola": cola_run,
                "idx": 0,
                "reporte_casos": [],
                "stats": [],
                "informe_localidades": [],
                "detalle_global": [],
                "contratos_actualizados": {},
                "conteo_global": {},
                "errores": [],
                "ahora": ahora,
                "titulo_mes": titulo_saldo_corte(ahora),
                "_barra": barra,
            }
            st.session_state.consolidacion_work = work
        else:
            work["cola"] = _asegurar_cola_en_disco(work["cola"])
            if "_barra" not in work:
                work["_barra"] = _barra_nueva(len(work["cola"]))
            st.session_state.consolidacion_work = work

        work = st.session_state.consolidacion_work
        total = len(work["cola"])

        while work["idx"] < total:
            idx = work["idx"]
            reporte_paso = ReporteEjecucion()
            _procesar_localidad_en_work(
                work, work["cola"][idx], reporte_paso, progress, idx, total
            )
            work["reporte_casos"].extend(c.a_dict() for c in reporte_paso.casos)
            work["idx"] = idx + 1

        _barra_tick(work.get("_barra", _barra_nueva(total)), progress, "Consolidación · terminada")
        progress.empty()

        reporte = _reporte_con_casos_guardados(work["reporte_casos"])
        exito = _aplicar_work_a_sesion(work)
        localidades_ok = len(st.session_state.get("cruce_informe", []))
        reporte.cerrar(exito, localidades_ok, n)
        _guardar_reporte_en_sesion(reporte)
        st.session_state.pop("consolidacion_work", None)

        if exito and st.session_state.processed:
            titulo_mes = st.session_state.get("titulo_saldo_corte", "")
            sin_res = sum(
                i.get("sin_resolver", 0)
                for i in st.session_state.get("cruce_informe", [])
            )
            msg = (
                f"**{n}** localidad(es) consolidadas · columna **{titulo_mes}** "
                f"(Cps, Suspendidos, Próximos, Trámites, Liquidados, Estrategias)."
            )
            if sin_res:
                msg += f" **{sin_res}** contrato(s) sin coincidencia automática."
            _vaciar_cola_tras_consolidar()
            st.success(msg)
        else:
            errores_ej = st.session_state.pop("errores_ejecucion", [])
            if any(es_error_matriz_protegida(e) for e in errores_ej):
                st.error(
                    "La Matriz tiene contraseña de apertura. "
                    "Ábrala en Excel, quite esa contraseña si existe y súbala de nuevo."
                )
            else:
                st.error("No se consolidaron las localidades correctamente.")
            for detalle in errores_ej:
                st.markdown(f"- {detalle}")

        # Evita duplicar el download_button si luego se pinta el panel completo.
        if not (
            exito
            and st.session_state.get(_CLAVE_MOSTRAR_RESULTADOS)
        ):
            mostrar_reporte_tecnico_admin()
    finally:
        st.session_state.consolidacion_en_curso = False
        # Si aún hay trabajo pendiente (st.rerun al siguiente paso), mantener el flag.
        if st.session_state.get("consolidacion_work") is None:
            st.session_state.ejecutar_consolidacion_ahora = False


def render_aviso_validacion_entrada() -> None:
    if not st.session_state.get("_pc_validacion_entrada_fallo"):
        return
    mensaje_validacion = st.session_state.get(
        "_pc_validacion_entrada_mensaje",
        "Revisión detenida. Hay una entrada de la cola que no se pudo validar.",
    )
    errores_validacion = st.session_state.get("_pc_validacion_entrada_errores", [])
    texto_validacion = (
        f"**{mensaje_validacion}**\n\n"
        "Quite la entrada marcada abajo, cárguela de nuevo y vuelva a ejecutar."
    )
    if errores_validacion:
        detalle = "\n".join(f"- {detalle}" for detalle in errores_validacion)
        texto_validacion = f"{texto_validacion}\n\n{detalle}"
    st.warning(texto_validacion)


if not st.session_state.get("acceso_autorizado"):
    if contrasena_acceso_esperada() is None:
        st.session_state.acceso_autorizado = True
    else:
        render_portada_acceso()
        st.stop()

render_mensaje_bienvenida_pendiente()

@st.cache_resource(show_spinner=False)
def _dependencias_consolidacion():
    """Una sola carga por proceso del servidor (no en cada F5)."""
    import importlib
    import msoffcrypto
    import msoffcrypto.exceptions as ms_exceptions
    import pandas as pd

    import cxp_cruce
    cxp_cruce = importlib.reload(cxp_cruce)

    from cxp_cruce import (
        METODOS_LABEL,
        aplicar_desempate_en_contratos,
        clave_desde_detalle,
        claves_pendientes_localidad,
        procesar_localidad_cxp,
        recalcular_estadisticas_localidad,
        resolver_hoja_cruce_cxp,
        titulo_saldo_corte,
        validar_desempate_completo,
    )
    from reporte_ejecucion import (
        CasoNoPrevisto,
        ReporteEjecucion,
        registrar_resultado_localidad,
    )

    return {
        "pd": pd,
        "msoffcrypto": msoffcrypto,
        "ms_exceptions": ms_exceptions,
        "cxp_cruce": cxp_cruce,
        "METODOS_LABEL": METODOS_LABEL,
        "aplicar_desempate_en_contratos": aplicar_desempate_en_contratos,
        "clave_desde_detalle": clave_desde_detalle,
        "claves_pendientes_localidad": claves_pendientes_localidad,
        "procesar_localidad_cxp": procesar_localidad_cxp,
        "recalcular_estadisticas_localidad": recalcular_estadisticas_localidad,
        "resolver_hoja_cruce_cxp": resolver_hoja_cruce_cxp,
        "titulo_saldo_corte": titulo_saldo_corte,
        "validar_desempate_completo": validar_desempate_completo,
        "CasoNoPrevisto": CasoNoPrevisto,
        "ReporteEjecucion": ReporteEjecucion,
        "registrar_resultado_localidad": registrar_resultado_localidad,
    }


def _necesita_dependencias_pesadas() -> bool:
    if st.session_state.get("ejecutar_consolidacion_ahora"):
        return True
    if st.session_state.get("cola_localidades") or st.session_state.get("consolidacion_work"):
        return True
    if st.session_state.get("processed") and st.session_state.get(
        _CLAVE_MOSTRAR_RESULTADOS, False
    ):
        return True
    if st.session_state.get("acceso_autorizado"):
        # Formulario / añadir a cola: hace falta pandas, cxp_cruce, etc.
        if st.session_state.get("processed") and not st.session_state.get(
            _CLAVE_MOSTRAR_RESULTADOS, False
        ):
            return False  # F5 rápido: solo resumen tras consolidar
        return True
    return False


def _inicializar_dependencias_modulo() -> None:
    global pd, msoffcrypto, ms_exceptions, cxp_cruce, METODOS_LABEL
    global aplicar_desempate_en_contratos, clave_desde_detalle
    global claves_pendientes_localidad, procesar_localidad_cxp
    global recalcular_estadisticas_localidad, resolver_hoja_cruce_cxp, titulo_saldo_corte
    global validar_desempate_completo, CasoNoPrevisto, ReporteEjecucion
    global registrar_resultado_localidad
    if globals().get("_DEPS_MODULO_LISTAS"):
        return
    dep = _dependencias_consolidacion()
    pd = dep["pd"]
    msoffcrypto = dep["msoffcrypto"]
    ms_exceptions = dep["ms_exceptions"]
    cxp_cruce = dep["cxp_cruce"]
    METODOS_LABEL = dep["METODOS_LABEL"]
    aplicar_desempate_en_contratos = dep["aplicar_desempate_en_contratos"]
    clave_desde_detalle = dep["clave_desde_detalle"]
    claves_pendientes_localidad = dep["claves_pendientes_localidad"]
    procesar_localidad_cxp = dep["procesar_localidad_cxp"]
    recalcular_estadisticas_localidad = dep["recalcular_estadisticas_localidad"]
    resolver_hoja_cruce_cxp = dep["resolver_hoja_cruce_cxp"]
    titulo_saldo_corte = dep["titulo_saldo_corte"]
    validar_desempate_completo = dep["validar_desempate_completo"]
    CasoNoPrevisto = dep["CasoNoPrevisto"]
    ReporteEjecucion = dep["ReporteEjecucion"]
    registrar_resultado_localidad = dep["registrar_resultado_localidad"]
    globals()["_DEPS_MODULO_LISTAS"] = True


if _necesita_dependencias_pesadas():
    _inicializar_dependencias_modulo()

_limpiar_estado_consolidacion_bloqueado()
_sanear_sesion_consolidacion()
if not st.session_state.get("ejecutar_consolidacion_ahora"):
    if st.session_state.get(_CLAVE_MOSTRAR_RESULTADOS):
        _compactar_sesion_vista_completa()
    else:
        _compactar_sesion_para_f5()

_barra_consolidacion = None
if st.session_state.get("ejecutar_consolidacion_ahora"):
    _barra_consolidacion = st.progress(0, text="Iniciando consolidación…")

# ── Título ─────────────────────────────────────────────────────────────────────
st.markdown('<h1 class="app-title">Plan de Choque</h1>', unsafe_allow_html=True)
st.markdown(
    '<p class="app-subtitle">Bogotá — consolidación por localidad</p>',
    unsafe_allow_html=True,
)

_omitir_formulario = _omitir_formulario_entrada()

if _omitir_formulario:
    cola_ej = st.session_state.get("cola_ejecucion") or []
    if not cola_ej and st.session_state.get("consolidacion_work"):
        cola_ej = st.session_state.consolidacion_work.get("cola", [])
    aviso_consolidacion = st.empty()
    if cola_ej:
        nombres = ", ".join(item["localidad"] for item in cola_ej)
        aviso_consolidacion.info(
            f"Consolidación en curso ({len(cola_ej)} localidad/es): **{nombres}**. "
            "No cierre esta pestaña hasta ver el resultado."
        )
    else:
        aviso_consolidacion.info(
            "Consolidación en curso. No cierre esta pestaña hasta ver el resultado."
        )
    _ejecutar_consolidacion_si_pendiente(_barra_consolidacion)
    if not _consolidacion_corriendo():
        aviso_consolidacion.empty()
    if st.session_state.get("_pc_validacion_entrada_fallo"):
        aviso_consolidacion.empty()
        _omitir_formulario = False
    st.divider()

render_aviso_validacion_entrada()

cola_correccion = st.session_state.get("cola_localidades", [])
_modo_correccion_cola = bool(
    st.session_state.get("_pc_mostrar_cola_correccion") and cola_correccion
)
if _modo_correccion_cola:
    st.markdown(
        '<p class="section-title">Entradas en cola por corregir</p>',
        unsafe_allow_html=True,
    )
    st.caption("Elimine la entrada equivocada, vacíe la cola o agregue una entrada corregida.")
    for i, item in enumerate(cola_correccion):
        loc = item["localidad"]
        with st.container(border=True):
            st.markdown(f"**{i + 1}. {loc}**")
            st.caption(texto_archivo_cola("Contratos", item.get("contratos")))
            st.caption(texto_archivo_cola("Matriz", item.get("matriz")))
            if st.button(
                "Eliminar entrada",
                key=f"quitar_cola_correccion_{loc}",
                use_container_width=True,
                help=f"Eliminar {loc} de la cola",
            ):
                quitar_de_cola(loc, mantener_modo_correccion=True)
                if not st.session_state.get("cola_localidades"):
                    st.session_state.pop("_pc_mostrar_cola_correccion", None)
                limpiar_resultado_consolidado()
                if not st.session_state.get("cola_localidades"):
                    st.session_state.pop("_pc_mostrar_cola_correccion", None)
                    st.session_state.pop("_pc_mostrar_formulario_correccion", None)
                st.rerun()
    c_agregar_corr, c_vaciar_corr = st.columns(2)
    with c_agregar_corr:
        if st.button(
            "Agregar entrada",
            type="secondary",
            use_container_width=True,
            key="btn_mostrar_formulario_correccion",
        ):
            st.session_state["_pc_mostrar_formulario_correccion"] = True
            st.rerun()
    with c_vaciar_corr:
        if st.button(
            "Vaciar cola y empezar de nuevo",
            type="secondary",
            use_container_width=True,
            key="btn_vaciar_cola_correccion",
        ):
            st.session_state.cola_localidades = []
            st.session_state.pop("_pc_mostrar_cola_correccion", None)
            st.session_state.pop("_pc_mostrar_formulario_correccion", None)
            limpiar_resultado_consolidado()
            st.rerun()

uk = st.session_state.upload_key

_trabajo_pausado = st.session_state.get("consolidacion_work")
if not _omitir_formulario and _trabajo_pausado:
    _idx = _trabajo_pausado.get("idx", 0)
    _total = len(_trabajo_pausado.get("cola", []))
    with st.container(border=True):
        st.warning(
            f"Quedó una consolidación interrumpida (**{_idx}/{_total}** localidades procesadas)."
        )
        c_reanudar, c_descartar = st.columns(2)
        with c_reanudar:
            if st.button(
                "Reanudar consolidación",
                type="primary",
                use_container_width=True,
                key="btn_reanudar_consolidacion",
            ):
                st.session_state.ejecutar_consolidacion_ahora = True
                st.session_state.consolidacion_en_curso = True
                st.session_state.pendiente_consolidacion = False
                st.rerun()
        with c_descartar:
            if st.button(
                "Descartar y empezar de nuevo",
                use_container_width=True,
                key="btn_descartar_consolidacion",
            ):
                st.session_state.pop("consolidacion_work", None)
                st.session_state.pendiente_consolidacion = False
                st.session_state.ejecutar_consolidacion_ahora = False
                st.rerun()

mostrar_formulario_entrada = not _omitir_formulario and (
    not _modo_correccion_cola
    or bool(st.session_state.get("_pc_mostrar_formulario_correccion"))
)

if mostrar_formulario_entrada:
    # ── Formulario ─────────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown('<p class="form-card-title">Entrada por localidad</p>', unsafe_allow_html=True)
        st.caption(
            "Proporcione el archivo de **Contratos plan de choque** y su **Matriz** "
            "correspondiente por localidad. En Excel, quite los **filtros/autofiltros** "
            "antes de subirlos."
        )

        st.markdown('<p class="field-label">Localidad</p>', unsafe_allow_html=True)
        localidad = st.selectbox(
            "Localidad",
            options=[SELECCION_LOCALIDAD] + LOCALIDADES,
            index=0,
            label_visibility="collapsed",
            key=f"select_localidad_{uk}",
        )

        st.divider()

        st.markdown(
            '<p class="field-label"><span class="field-num">1</span> Contratos plan de choque</p>',
            unsafe_allow_html=True,
        )
        archivo_contratos = st.file_uploader(
            "Contratos plan de choque",
            type=["xlsx", "xls"],
            accept_multiple_files=False,
            label_visibility="collapsed",
            key=f"uploader_contratos_{uk}",
            help="Un solo archivo Excel (.xlsx o .xls).",
        )
        if archivo_contratos:
            st.markdown(
                f'<p class="file-ok">✓ {archivo_contratos.name}</p>',
                unsafe_allow_html=True,
            )

        st.markdown(
            '<p class="field-label"><span class="field-num">2</span> Matriz</p>',
            unsafe_allow_html=True,
        )
        archivo_matriz = st.file_uploader(
            "Matriz",
            type=["xlsx", "xls"],
            accept_multiple_files=False,
            label_visibility="collapsed",
            key=f"uploader_matriz_{uk}",
            help=(
                "Excel .xlsx o .xls con hoja MATRIZ OXP."
            ),
        )
        if archivo_matriz:
            st.markdown(
                f'<p class="file-ok">✓ {archivo_matriz.name}</p>',
                unsafe_allow_html=True,
            )

        form_ok = formulario_completo(localidad, archivo_contratos, archivo_matriz)

        add_clicked = st.button(
            "Agregar entrada" if _modo_correccion_cola else "Añadir a cola de consolidados",
            type="secondary",
            use_container_width=True,
            help=(
                "Agrega la entrada corregida a la cola."
                if _modo_correccion_cola
                else "Guarda la localidad y los archivos en la cola. Luego puede cargar la siguiente."
            ),
        )

    if add_clicked:
        if not form_ok:
            st.warning(
                "Complete la localidad y los dos archivos antes de añadir el consolidado a la cola."
            )
        elif any(
            item["localidad"] == localidad for item in st.session_state.cola_localidades
        ):
            st.warning(f"**{localidad}** ya está en la cola. Quítela o elija otra localidad.")
        else:
            st.session_state.cola_localidades.append(
                entrada_desde_formulario(localidad, archivo_contratos, archivo_matriz)
            )
            st.session_state.upload_key += 1
            st.session_state.pop("_pc_mostrar_formulario_correccion", None)
            _purgar_uploaders_obsoletos()
            st.toast(f"{localidad} añadido a la cola", icon="➕")
            st.rerun()

# ── Cola pendiente ─────────────────────────────────────────────────────────────
cola = st.session_state.cola_localidades
if (
    not _omitir_formulario
    and cola
    and not st.session_state.get("_pc_mostrar_cola_correccion")
):
    st.markdown(
        f'<p class="section-title">Cola de consolidados ({len(cola)})</p>',
        unsafe_allow_html=True,
    )
    st.caption(
        "Revise las entradas antes de ejecutar. Si una localidad o archivo no corresponde, elimine esa entrada y cárguela de nuevo."
    )
    for i, item in enumerate(cola):
        loc = item["localidad"]
        contratos_nombre = escape(item["contratos"]["name"])
        matriz_nombre = escape(item["matriz"]["name"])
        st.markdown('<div class="queue-compact-row">', unsafe_allow_html=True)
        c_num, c_loc, c_con, c_mat, c_btn = st.columns([0.35, 1.25, 2.2, 2.2, 0.85])
        with c_num:
            st.markdown(
                f'<span class="queue-index">{i + 1}</span>',
                unsafe_allow_html=True,
            )
        with c_loc:
            st.markdown(
                f'<p class="queue-localidad">{escape(loc)}</p>',
                unsafe_allow_html=True,
            )
        with c_con:
            st.markdown(
                f'<p class="queue-file-label">Contratos</p>'
                f'<p class="queue-file-name">{contratos_nombre}</p>',
                unsafe_allow_html=True,
            )
        with c_mat:
            st.markdown(
                f'<p class="queue-file-label">Matriz</p>'
                f'<p class="queue-file-name">{matriz_nombre}</p>',
                unsafe_allow_html=True,
            )
        with c_btn:
            if st.button(
                "Quitar",
                key=f"quitar_cola_{loc}",
                use_container_width=True,
                help=f"Eliminar {loc} de la cola",
            ):
                quitar_de_cola(loc)
                if not st.session_state.get("cola_localidades"):
                    st.session_state.pop("_pc_mostrar_cola_correccion", None)
                limpiar_resultado_consolidado()
                st.toast(f"{loc} eliminado de la cola")
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    if st.button("Vaciar cola", type="secondary"):
        st.session_state.cola_localidades = []
        st.session_state.pop("_pc_mostrar_cola_correccion", None)
        st.session_state.pop("_pc_mostrar_formulario_correccion", None)
        limpiar_resultado_consolidado()
        st.rerun()

run_clicked = False
hay_cola_pendiente = bool(st.session_state.get("cola_localidades"))
if (
    st.session_state.get("processed")
    and not hay_cola_pendiente
    and not _consolidacion_corriendo()
):
    if st.button(
        "Iniciar nuevo reporte",
        type="primary",
        use_container_width=True,
        key="btn_iniciar_nuevo_reporte",
        help="Limpia el resultado actual y vuelve al formulario de carga.",
    ):
        iniciar_nuevo_reporte()
        st.rerun()
elif not _consolidacion_corriendo():
    st.divider()
    run_clicked = st.button(
        "Ejecutar consolidación",
        type="primary",
        use_container_width=True,
        key="btn_ejecutar_consolidacion",
        help="Procesa todos los consolidados de la cola.",
    )

if run_clicked:
    cola_actual = st.session_state.cola_localidades
    if not cola_actual:
        st.warning(
            "Añada al menos un consolidado a la cola (localidad, Contratos y Matriz) "
            "antes de ejecutar."
        )
    else:
        archivos_ok, errores_archivos = validar_archivos_en_cola(cola_actual)
        if not archivos_ok:
            st.session_state["_pc_validacion_entrada_fallo"] = True
            st.session_state["_pc_mostrar_cola_correccion"] = True
            st.session_state["_pc_validacion_entrada_mensaje"] = (
                "Revisión detenida. Hay entradas en cola sin archivo disponible."
            )
            st.session_state["_pc_validacion_entrada_errores"] = errores_archivos
            st.rerun()
        elif not puede_ejecutar_cola(cola_actual):
            st.warning(
                "Cada localidad en la cola debe incluir Contratos plan de choque y Matriz."
            )
        else:
            cola_ejec = cola_para_ejecutar(cola_actual)
            st.session_state.cola_ejecucion = cola_ejec
            st.session_state.pendiente_consolidacion = True
            st.session_state.ejecutar_consolidacion_ahora = True
            st.session_state.pop("_pc_validacion_entrada_fallo", None)
            st.session_state.pop("_pc_mostrar_cola_correccion", None)
            st.session_state.pop("_pc_mostrar_formulario_correccion", None)
            st.session_state.pop("consolidacion_work", None)
            st.rerun()

def _render_panel_resultados_completos() -> None:
    """Tablas, desempate y descargas (solo cuando el usuario lo pide o tras consolidar)."""
    snap = _cargar_snapshot_consolidacion()
    stats = snap.get("file_stats") or []
    informe = snap.get("informe") or []
    st.session_state.cruce_resumen_global = snap.get("cruce_resumen_global") or []
    titulo_mes = snap.get("titulo_saldo_corte") or st.session_state.get("titulo_saldo_corte", "")
    etiqueta_cxp = titulo_mes or "Saldo de corte"
    fecha_analisis = snap.get("fecha_analisis") or st.session_state.get("fecha_analisis")
    procesado_en = snap.get("last_processed_at") or st.session_state.get("last_processed_at", "")

    st.markdown('<div id="resultado-consolidado-anchor"></div>', unsafe_allow_html=True)
    render_scroll_resultados_si_pendiente()

    total_contratos = sum(i.get("total_contratos", 0) for i in informe)
    total_ok = sum(i.get("contratos_ok", 0) for i in informe)
    sin_resolver = sum(i.get("sin_resolver", 0) for i in informe)
    cruce_directo = sum(int((i.get("conteo") or {}).get("k4_exacto", 0) or 0) for i in informe)
    revision_adicional = sum(conteo_revision_adicional(i.get("conteo") or {}) for i in informe)
    contratos_act = st.session_state.get("contratos_actualizados", {})

    if sin_resolver > 0:
        detalle_todo = _detalle_con_filas_excel_contratos(
            obtener_cruce_detalle(),
            dict(contratos_act),
        )
        st.markdown(
            '<p class="section-title">Contratos sin coincidencia automática</p>',
            unsafe_allow_html=True,
        )
        st.markdown(
            f"""
            <div class="desempate-callout">
                <div class="desempate-kicker">Acción requerida</div>
                <div class="desempate-callout-title">Contratos sin coincidencia automática</div>
                <p>Hay <strong>{sin_resolver}</strong> contrato(s) sin coincidencia automática en
                <strong>{escape(str(titulo_mes))}</strong>. Complete estos casos para habilitar
                las descargas de Contratos actualizados y archivos globales.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if detalle_todo:
            df_resumen_sr = resumen_sin_resolver_por_localidad(detalle_todo)
            if len(df_resumen_sr) > 1:
                st.markdown("**Por localidad**")
                render_tabla_resultados(df_resumen_sr)
            st.caption(
                "Revise cada caso en pantalla y avance con **Siguiente**. "
                "Al terminar, aplique las selecciones."
            )
            render_asistente_desempate(detalle_todo, titulo_mes)
        else:
            st.warning(
                "No se pudo cargar el detalle de los contratos pendientes. "
                "Vuelva a ejecutar la consolidación para reconstruirlo."
            )
        return

    st.markdown('<p class="section-title">Resultado consolidado</p>', unsafe_allow_html=True)
    if titulo_mes or procesado_en or fecha_analisis:
        partes_fecha = []
        if titulo_mes:
            partes_fecha.append(f"Corte **{titulo_mes}**")
        if fecha_analisis:
            partes_fecha.append(formato_fecha_colombia(fecha_analisis))
        if procesado_en:
            partes_fecha.append(f"Procesado {procesado_en}")
        st.caption(" · ".join(partes_fecha))
    mostrar_reporte_tecnico_admin()

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">Localidades</div>'
            f'<div class="metric-value metric-value-sm">{len(informe)}</div></div>',
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">Contratos cruzados</div>'
            f'<div class="metric-value metric-value-sm">{total_ok}/{total_contratos}</div></div>',
            unsafe_allow_html=True,
        )
    with c3:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">Cruce directo</div>'
            f'<div class="metric-value">{cruce_directo}</div></div>',
            unsafe_allow_html=True,
        )
    with c4:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">Cruce complementario</div>'
            f'<div class="metric-value">{revision_adicional}</div></div>',
            unsafe_allow_html=True,
        )
    requiere_detalle = informe_requiere_detalle_cruce(informe)
    mostrar_informe_cruce_consolidado(
        informe,
        titulo_mes,
        cargar_tablas_detalle=requiere_detalle,
    )
    descargas_ok = consolidacion_lista_para_descarga()

    if stats:
        with st.expander("Archivos de entrada (Matriz y Contratos)", expanded=False):
            df_stats = pd.DataFrame(stats)
            if "CXP (suma mes)" in df_stats.columns:
                df_stats["CXP (suma mes)"] = df_stats["CXP (suma mes)"].map(
                    formato_moneda_contable
                )
                df_stats = df_stats.rename(columns={"CXP (suma mes)": "CXP del mes"})
            cols_show = [
                c
                for c in [
                    "Localidad",
                    "Archivo",
                    "Nombre",
                    "Filas",
                    "CXP del mes",
                ]
                if c in df_stats.columns
            ]
            render_tabla_resultados(df_stats[cols_show])

    contratos_act = st.session_state.get("contratos_actualizados", {})
    if contratos_act:
        fecha_dl = st.session_state.get("fecha_analisis") or fecha_referencia_analisis()
        n_loc = len(contratos_act)
        st.markdown(
            '<p class="section-title">Contratos plan de choque actualizados</p>',
            unsafe_allow_html=True,
        )
        st.caption(
            f"Un ZIP con {n_loc} archivo(s) Excel (uno por localidad). "
            f"Abra el ZIP y use el .xlsx dentro. "
            f"El mes en el nombre se actualiza a «- {mes_capitalizado(fecha_dl)}» "
            "(conserva el texto antes del guion)."
        )
        if not descargas_ok:
            st.caption(
                "Disponible cuando no haya contratos sin coincidencia automática."
            )
        zip_info = st.session_state.get("zip_descarga_contratos")
        nombre_dl = (zip_info or {}).get("nombre") or "contratos.zip"
        mime_dl = (zip_info or {}).get("mime") or "application/zip"
        ruta_zip = (zip_info or {}).get("path")
        if descargas_ok and ruta_zip and Path(ruta_zip).is_file():
            datos_dl = _leer_binario_desde_ruta(
                ruta_zip, Path(ruta_zip).stat().st_mtime
            )
            st.download_button(
                label="Descargar Contratos actualizados (ZIP)",
                data=datos_dl,
                file_name=nombre_dl,
                mime=mime_dl,
                key="dl_contratos_todas",
                use_container_width=True,
            )
        elif descargas_ok:
            st.warning(
                "No se encontró el ZIP en el servidor. Vuelva a consolidar o aplique "
                "las selecciones para regenerarlo."
            )
        with st.expander("Archivos incluidos en la descarga"):
            for loc, data in sorted(contratos_act.items(), key=lambda x: x[0]):
                st.markdown(
                    f"- **{loc}:** "
                    f"`{nombre_descarga_contratos_actualizado(loc, data.get('nombre_contratos', ''), fecha_dl)}`"
                )
    st.markdown('<p class="section-title">Archivos globales de salida</p>', unsafe_allow_html=True)
    if not descargas_ok:
        st.caption(
            "Bloqueados mientras haya contratos sin coincidencia automática. "
            "Los archivos globales solo se generan con datos 100% completos."
        )
    else:
        st.caption(
            "Reúnen la información de **todas** las localidades: "
            "Matriz y Contratos plan de choque actualizados (con selecciones aplicadas). "
            "Se guardan en Descargas."
        )
    if st.button(
        "Descargar archivos de salida",
        use_container_width=True,
        key="btn_descargar_excel",
        disabled=not descargas_ok,
    ):
        try:
            df_export = dataframe_consolidado()
            rutas = guardar_archivos_salida(df_export, stats)
            st.session_state.ultima_descarga = [str(r) for r in rutas]
            st.toast("2 archivos guardados en Descargas", icon="✅")
            lista = "\n".join(f"- `{r.name}`" for r in rutas)
            st.success(
                "Se guardaron **2** archivos en Descargas (siempre los mismos, "
                "sin importar cuántas localidades procesó):\n\n" + lista
            )
        except (OSError, ValueError) as e:
            st.error(f"No se pudo guardar en Descargas: {e}")

    if st.button(
        "Ocultar detalle de resultados",
        use_container_width=True,
        key="btn_ocultar_resultados_completos",
    ):
        st.session_state[_CLAVE_MOSTRAR_RESULTADOS] = False
        st.rerun()


# ── Resultados ─────────────────────────────────────────────────────────────────
if st.session_state.processed:
    resumen_ligero = st.session_state.get(_CLAVE_RESUMEN_LIGERO) or {}
    if not resumen_ligero:
        snap_tmp = _cargar_snapshot_consolidacion()
        informe_tmp = snap_tmp.get("informe") or []
        if informe_tmp:
            resumen_ligero = _resumen_ligero_desde_informe(informe_tmp)
            st.session_state[_CLAVE_RESUMEN_LIGERO] = resumen_ligero

    if not st.session_state.get(_CLAVE_MOSTRAR_RESULTADOS):
        st.markdown('<p class="section-title">Resultado consolidado</p>', unsafe_allow_html=True)
        n_loc = resumen_ligero.get("n_localidades", 0)
        sin_r = resumen_ligero.get("sin_resolver", 0)
        titulo_mes = resumen_ligero.get("titulo_mes") or st.session_state.get(
            "titulo_saldo_corte", ""
        )
        st.success(
            f"Consolidación lista: **{n_loc}** localidad(es), "
            f"**{resumen_ligero.get('total_ok', 0)}/{resumen_ligero.get('total_contratos', 0)}** "
            f"contratos cruzados"
            + (f", **{sin_r}** sin coincidencia automática" if sin_r else "")
            + (f". Corte **{titulo_mes}**." if titulo_mes else ".")
        )
        st.caption(
            "Abra el detalle para revisar tablas, casos informativos y descargas."
        )
        if st.button(
            "Mostrar resultados completos",
            type="primary",
            use_container_width=True,
            key="btn_mostrar_resultados_completos",
        ):
            st.session_state[_CLAVE_MOSTRAR_RESULTADOS] = True
            st.session_state[_CLAVE_SCROLL_RESULTADOS] = True
            _inicializar_dependencias_modulo()
            st.rerun()
    else:
        _inicializar_dependencias_modulo()
        _render_panel_resultados_completos()
