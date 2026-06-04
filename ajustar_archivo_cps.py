"""Ajusta encabezados SALDO (azul/amarillo) en Cps por depurar de un .xlsx de Contratos."""
from __future__ import annotations

import re
import sys
import zipfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

import cxp_cruce as cxp

DEFAULT_ENTRADA = Path(
    r"c:\Users\f1rac\Downloads\Copia de 08. Contratos plan de choque Kennedy 2026 - Junio.xlsx"
)
DEFAULT_SALIDA = DEFAULT_ENTRADA.with_name(
    DEFAULT_ENTRADA.stem + " (encabezados Cps).xlsx"
)


def ajustar_contratos_cps(entrada: Path, salida: Path) -> None:
    with entrada.open("rb") as f:
        raw = f.read()

    resultado = cxp._finalizar_xlsx_contratos(raw)
    salida.write_bytes(resultado)
    print(f"Guardado: {salida}")

    data = salida.read_bytes()
    wb = load_workbook(BytesIO(data))
    ws = wb[cxp.resolver_hoja_cruce_cxp(wb.sheetnames)]
    st = zipfile.ZipFile(BytesIO(data)).read("xl/styles.xml").decode()
    block = re.search(r"<cellXfs[^>]*>(.*)</cellXfs>", st, re.DOTALL)
    xfs = re.findall(r"<xf\b[^>]*>", block.group(1) if block else "")
    fills = re.findall(r"<fill>.*?</fill>", st, re.DOTALL)
    idx = wb.sheetnames.index(ws.title) + 1
    sheet = zipfile.ZipFile(BytesIO(data)).read(f"xl/worksheets/sheet{idx}.xml").decode()
    print("Verificación fila 3 (SALDO por mes):")
    for col, _ in cxp._listar_columnas_corte_mes_cps(ws):
        coord = ws.cell(3, col).coordinate
        tit = ws.cell(3, col).value
        m = re.search(rf'<c r="{re.escape(coord)}" s="(\d+)"', sheet)
        sid = int(m.group(1)) if m else -1
        xf = xfs[sid] if 0 <= sid < len(xfs) else ""
        fid_m = re.search(r'fillId="(\d+)"', xf)
        fid = int(fid_m.group(1)) if fid_m else -1
        rgb_m = re.search(r'rgb="([^"]+)"', fills[fid]) if 0 <= fid < len(fills) else None
        apply = 'applyFill="1"' in xf
        print(
            f"  {coord} {str(tit)[:28]!r} applyFill={apply} "
            f"rgb={rgb_m.group(1) if rgb_m else '?'}"
        )


if __name__ == "__main__":
    entrada = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_ENTRADA
    salida = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_SALIDA
    ajustar_contratos_cps(entrada, salida)
